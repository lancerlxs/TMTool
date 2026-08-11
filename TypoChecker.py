"""
汉字错别字检查工具 — 基于pycorrector数据，检查Excel指定列中的错别字并修改
运行环境：Windows 7 + Python 3.8
依赖：PyQt5, openpyxl, pycorrector(仅用其数据文件)
"""

import sys
import os
from datetime import datetime

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QLineEdit,
                             QFileDialog, QTableWidget, QTableWidgetItem,
                             QHeaderView, QComboBox, QMessageBox,
                             QProgressBar, QGroupBox, QFormLayout)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor


def _load_typo_pairs():
    """从 pycorrector 数据文件加载错别字对照表（错→对）"""
    pairs = {}
    # 手动补充的公文常见错别字
    manual = {
        "帐号": "账号", "帐户": "账户", "帐目": "账目", "帐单": "账单", "帐务": "账务",
        "部份": "部分", "成份": "成分", "身分": "身份", "布署": "部署", "烦琐": "繁琐",
        "供献": "贡献", "到塌": "倒塌", "溶合": "融合", "记念": "纪念", "连系": "联系",
        "好象": "好像", "想象": "想象", "想像": "想象", "录象": "录像", "画象": "画像",
        "暗然": "黯然", "渡假": "度假", "欢渡": "欢度", "拼博": "拼搏",
        "做为": "作为", "按装": "安装", "按排": "安排", "复盖": "覆盖",
    }
    pairs.update(manual)

    # 从 pycorrector 数据文件加载同音/同形错别字
    try:
        import pycorrector as _pc
        data_dir = os.path.join(os.path.dirname(_pc.__file__), "data")
        # 同音字（常见误用）
        pinyin_file = os.path.join(data_dir, "same_pinyin.txt")
        if os.path.isfile(pinyin_file):
            with open(pinyin_file, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    parts = line.split("\t")
                    if len(parts) >= 2:
                        correct_char = parts[0]
                        wrong_chars = parts[1].split() + (parts[2].split() if len(parts) > 2 else [])
                        for wc in wrong_chars:
                            if len(wc) == 1 and wc != correct_char:
                                pairs[wc] = correct_char
        # 同形字
        stroke_file = os.path.join(data_dir, "same_stroke.txt")
        if os.path.isfile(stroke_file):
            with open(stroke_file, encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    chars = line.split()
                    if len(chars) >= 2:
                        for i in range(1, len(chars)):
                            if len(chars[i]) == 1:
                                pairs[chars[i]] = chars[0]
    except Exception:
        pass

    return pairs


TYPO_PAIRS = _load_typo_pairs()


QSS = """
QWidget { background-color: #ecf0f1; font-family: 'Microsoft YaHei', Arial; font-size: 14px; }
QMainWindow { background-color: #ecf0f1; }
QLabel { color: #2c3e50; }
QLabel#TitleLabel { font-size: 22px; font-weight: bold; color: #1a5276; padding: 8px; }
QGroupBox {
    border: 2px solid #aed6f1; border-radius: 8px; margin-top: 12px; padding-top: 16px;
    font-weight: bold; color: #1a5276; background-color: #ffffff;
}
QGroupBox::title { subcontrol-origin: margin; left: 14px; padding: 0 6px; }
QLineEdit, QComboBox {
    border: 1px solid #bdc3c7; border-radius: 4px; padding: 6px; background-color: white;
}
QLineEdit:focus { border: 2px solid #3498db; }
QPushButton {
    border-radius: 5px; padding: 8px 20px; font-weight: bold; font-size: 14px; border: none;
}
QPushButton#GreenBtn { background-color: #27ae60; color: white; }
QPushButton#GreenBtn:hover { background-color: #2ecc71; }
QPushButton#GreenBtn:disabled { background-color: #bdc3c7; }
QPushButton#BlueBtn { background-color: #2980b9; color: white; }
QPushButton#BlueBtn:hover { background-color: #3498db; }
QPushButton#OrangeBtn { background-color: #e67e22; color: white; }
QPushButton#OrangeBtn:hover { background-color: #f39c12; }
QTableWidget {
    background-color: white; border: 1px solid #bdc3c7; border-radius: 4px; gridline-color: #ecf0f1;
    font-size: 14px;
}
QHeaderView::section {
    background-color: #d6eaf8; color: #1a5276; font-weight: bold; padding: 6px;
    border: none; border-right: 1px solid #bdc3c7; border-bottom: 1px solid #bdc3c7; font-size: 14px;
}
QTableWidget::item:selected { background-color: #3498db; color: white; }
QProgressBar { border: 1px solid #bdc3c7; border-radius: 4px; text-align: center; }
QProgressBar::chunk { background-color: #27ae60; border-radius: 3px; }
"""


class TypoCheckerWindow(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("汉字错别字检查工具")
        self.setGeometry(80, 60, 1300, 800)
        self.workbook = None
        self.typos_found = []
        self._log_content = ""
        self._init_ui()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        self.setStyleSheet(QSS)
        layout = QVBoxLayout(central)
        layout.setSpacing(10)
        layout.setContentsMargins(15, 10, 15, 10)

        title = QLabel("汉字错别字检查工具")
        title.setObjectName("TitleLabel")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # 文件选择
        file_group = QGroupBox("文件选择")
        file_layout = QFormLayout()
        self.file_path = QLineEdit()
        self.file_path.setReadOnly(True)
        self.file_path.setPlaceholderText("选择Excel文件...")
        btn_browse = QPushButton("选择文件")
        btn_browse.setObjectName("BlueBtn")
        btn_browse.clicked.connect(self.browse_file)
        h = QHBoxLayout(); h.addWidget(self.file_path); h.addWidget(btn_browse)
        file_layout.addRow("Excel文件:", h)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setEnabled(False)
        file_layout.addRow("工作表:", self.sheet_combo)

        self.col_input = QLineEdit()
        self.col_input.setPlaceholderText("输入列字母，如 A,B,C 或 A-C")
        file_layout.addRow("检查列:", self.col_input)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # 按钮
        btn_row = QHBoxLayout()
        self.check_btn = QPushButton("开始检查")
        self.check_btn.setObjectName("GreenBtn")
        self.check_btn.clicked.connect(self.start_check)
        self.check_btn.setEnabled(False)
        btn_row.addWidget(self.check_btn)

        self.fix_btn = QPushButton("一键修改")
        self.fix_btn.setObjectName("OrangeBtn")
        self.fix_btn.clicked.connect(self.fix_all)
        self.fix_btn.setEnabled(False)
        btn_row.addWidget(self.fix_btn)

        self.export_btn = QPushButton("导出结果")
        self.export_btn.setObjectName("BlueBtn")
        self.export_btn.clicked.connect(self.export_result)
        self.export_btn.setEnabled(False)
        btn_row.addWidget(self.export_btn)

        self.log_btn = QPushButton("查看日志")
        self.log_btn.setObjectName("BlueBtn")
        self.log_btn.clicked.connect(self.show_log)
        self.log_btn.setEnabled(False)
        btn_row.addWidget(self.log_btn)
        layout.addLayout(btn_row)

        self.progress = QProgressBar()
        layout.addWidget(self.progress)

        # 结果表格
        table_group = QGroupBox("检查结果（双击'是否修改'列可切换）")
        table_layout = QVBoxLayout()
        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(["工作表", "行", "列", "原文本", "错别字", "正确字", "是否修改"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setMinimumHeight(350)
        self.table.cellDoubleClicked.connect(self.on_cell_double_click)
        table_layout.addWidget(self.table)
        table_group.setLayout(table_layout)
        layout.addWidget(table_group)

        self.status_label = QLabel(f"已加载错别字字典: {len(TYPO_PAIRS)} 条对照规则")
        self.status_label.setStyleSheet("color: #7f8c8d; font-size: 13px; padding: 4px;")
        layout.addWidget(self.status_label)

    def browse_file(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx)")
        if not path:
            return
        self.file_path.setText(path)
        self.sheet_combo.clear()
        self.sheet_combo.setEnabled(False)
        self.check_btn.setEnabled(False)
        try:
            from openpyxl import load_workbook
            self.workbook = load_workbook(path, read_only=True)
            for name in self.workbook.sheetnames:
                self.sheet_combo.addItem(name)
            self.sheet_combo.setEnabled(True)
            self.check_btn.setEnabled(True)
            self.status_label.setText(f"已加载: {len(self.workbook.sheetnames)} 个工作表")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载失败: {e}")

    def _parse_columns(self, col_str):
        cols = set()
        for part in col_str.replace('，', ',').split(','):
            part = part.strip().upper()
            if '-' in part:
                a, b = part.split('-')
                for c in range(ord(a), ord(b) + 1):
                    cols.add(c - ord('A') + 1)
            elif len(part) == 1 and 'A' <= part <= 'Z':
                cols.add(ord(part) - ord('A') + 1)
        return sorted(cols)

    def start_check(self):
        path = self.file_path.text()
        if not path or not os.path.isfile(path):
            QMessageBox.warning(self, "提示", "请先选择Excel文件")
            return
        col_str = self.col_input.text().strip()
        if not col_str:
            QMessageBox.warning(self, "提示", "请输入要检查的列")
            return
        cols = self._parse_columns(col_str)
        if not cols:
            QMessageBox.warning(self, "提示", "列格式不正确")
            return
        sheet_name = self.sheet_combo.currentText()
        if not sheet_name:
            return

        self.typos_found = []
        self.table.setRowCount(0)
        self.progress.setValue(0)
        self.status_label.setText("正在检查...")

        from openpyxl import load_workbook
        try:
            wb = load_workbook(path, read_only=True)
            ws = wb[sheet_name]
            max_row = ws.max_row or 0
            found = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row), 1):
                for col_idx in cols:
                    if col_idx - 1 < len(row):
                        text = str(row[col_idx - 1].value) if row[col_idx - 1].value else ""
                        if not text:
                            continue
                        for wrong, correct in TYPO_PAIRS.items():
                            if wrong in text:
                                self.typos_found.append({
                                    'sheet': sheet_name, 'row': row_idx, 'col': col_idx,
                                    'text': text, 'wrong': wrong, 'correct': correct, 'modify': True
                                })
                                found += 1
                if max_row > 0:
                    self.progress.setValue(int(row_idx / max_row * 100))
                QApplication.processEvents()
            wb.close()
            self._display_results()
            self.progress.setValue(100)
            if found == 0:
                self.status_label.setText(f"检查完成：共 {max_row} 行，未发现错别字 ✓")
            else:
                self.status_label.setText(f"发现 {found} 处错别字")
                self.fix_btn.setEnabled(True)
                self.export_btn.setEnabled(True)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"检查失败: {e}")

    def _display_results(self):
        self.table.setRowCount(len(self.typos_found))
        for i, t in enumerate(self.typos_found):
            self.table.setItem(i, 0, QTableWidgetItem(t['sheet']))
            self.table.setItem(i, 1, QTableWidgetItem(str(t['row'])))
            self.table.setItem(i, 2, QTableWidgetItem(chr(t['col'] + 64)))
            self.table.setItem(i, 3, QTableWidgetItem(t['text']))
            w = QTableWidgetItem(t['wrong']); w.setForeground(QColor('#e74c3c')); w.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            self.table.setItem(i, 4, w)
            c = QTableWidgetItem(t['correct']); c.setForeground(QColor('#27ae60')); c.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            self.table.setItem(i, 5, c)
            m = QTableWidgetItem("✓ 修改"); m.setForeground(QColor('#27ae60'))
            self.table.setItem(i, 6, m)

    def on_cell_double_click(self, row, col):
        if col != 6:
            return
        item = self.table.item(row, 6)
        if item.text().startswith("✓"):
            item.setText("✗ 跳过"); item.setForeground(QColor('#e74c3c')); self.typos_found[row]['modify'] = False
        else:
            item.setText("✓ 修改"); item.setForeground(QColor('#27ae60')); self.typos_found[row]['modify'] = True

    def fix_all(self):
        to_fix = [t for t in self.typos_found if t['modify']]
        if not to_fix:
            return
        reply = QMessageBox.question(self, "确认", f"将修改 {len(to_fix)} 处错别字，确认？", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        from openpyxl import load_workbook
        try:
            wb = load_workbook(self.file_path.text())
            ws = wb[self.typos_found[0]['sheet']]
            modified = 0; logs = []
            for t in to_fix:
                cell = ws.cell(row=t['row'], column=t['col'])
                old = str(cell.value) if cell.value else ""
                new = old.replace(t['wrong'], t['correct'])
                if new != old:
                    cell.value = new; modified += 1
                    logs.append(f"[{t['sheet']}] {chr(t['col']+64)}{t['row']}: '{t['wrong']}' → '{t['correct']}'")
            base, ext = os.path.splitext(self.file_path.text())
            new_path = f"{base}_已修改_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
            wb.save(new_path); wb.close()
            self._log_content = (f"错别字修改日志\n原文件: {self.file_path.text()}\n修改时间: {datetime.now()}\n修改: {modified} 处\n保存为: {new_path}\n{'='*60}\n" + "\n".join(logs))
            self.log_btn.setEnabled(True)
            self.status_label.setText(f"已修改 {modified} 处 → {os.path.basename(new_path)}")
            QMessageBox.information(self, "完成", f"已修改 {modified} 处\n保存为: {new_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"修改失败: {e}")

    def export_result(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出", f"错别字结果_{datetime.now().strftime('%Y%m%d')}.xlsx", "*.xlsx")
        if not path:
            return
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.append(["工作表","行","列","原文本","错别字","正确字","是否修改"])
        for t in self.typos_found:
            ws.append([t['sheet'], t['row'], chr(t['col']+64), t['text'], t['wrong'], t['correct'], "修改" if t['modify'] else "跳过"])
        wb.save(path); QMessageBox.information(self, "成功", f"已导出: {path}")

    def show_log(self):
        from PyQt5.QtWidgets import QTextEdit, QDialog, QVBoxLayout
        dlg = QDialog(self); dlg.setWindowTitle("修改日志"); dlg.setMinimumSize(700, 500)
        te = QTextEdit(); te.setReadOnly(True); te.setPlainText(self._log_content)
        dlg.layout_ = QVBoxLayout(dlg); dlg.layout_.addWidget(te); dlg.exec_()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TypoCheckerWindow()
    window.show()
    sys.exit(app.exec_())

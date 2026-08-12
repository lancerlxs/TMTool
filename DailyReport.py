"""
电子卷宗随案生成中心 - 日报编写程序
功能：日历选择日期、填写日报数据、自动计算日/周/月合计、导出Excel
"""

import sys
import os
from datetime import datetime, timedelta

from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QCalendarWidget, QTableWidget,
                             QTableWidgetItem, QHeaderView, QPushButton,
                             QFileDialog, QMessageBox, QSplitter, QLabel,
                             QAbstractItemView, QFrame, QDialog, QComboBox,
                             QTreeWidget, QTreeWidgetItem, QLineEdit, QFormLayout)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor, QBrush, QTextCharFormat

from daily_report_data import (
    DEPARTMENTS, COL_COUNT, COL_NAMES, ARCHIVE_GROUPS,
    load_all_data, get_day_data, set_day_data,
    calc_daily_total, calc_weekly_total, calc_monthly_total,
    format_date_with_week, format_title_with_month,
    get_week_number,
    get_default_archive_category, load_categories, save_categories,
    get_day_categories, set_day_categories,
    load_baselines, save_baselines, has_any_data,
    get_archive_group_data, set_archive_group_value,
)


class DailyReportWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("电子卷宗随案生成中心 - 日报编写")
        self.setGeometry(80, 60, 1500, 900)
        self.data = load_all_data()
        self.categories = load_categories()
        self.baselines = load_baselines()
        self.current_date = datetime.now().strftime('%Y%m%d')
        self._building_table = False
        self._updating_totals = False
        self._init_ui()
        self._mark_calendar_dates()
        # 首次使用检查
        self._check_first_use()
        self._load_date(self.current_date)

    def _check_first_use(self):
        """首次使用检查：无数据则自动初始化"""
        if not has_any_data():
            dlg = InitDialog(self.baselines, auto=True, parent=self)
            if dlg.exec_() == QDialog.Accepted:
                self.baselines = dlg.get_baselines()
                save_baselines(self.baselines)

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)

        # === 左侧：日历 ===
        left = QWidget()
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)

        lbl = QLabel("选择日期")
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet("font-size: 16px; font-weight: bold; color: #1a5276; padding: 4px;")
        left_layout.addWidget(lbl)

        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        self.calendar.setFirstDayOfWeek(Qt.Monday)
        self.calendar.setMaximumDate(QDate.currentDate())  # 不能选择未来日期
        self.calendar.clicked.connect(self._on_date_clicked)
        self.calendar.currentPageChanged.connect(self._mark_calendar_dates)
        self.calendar.setStyleSheet("""
            QCalendarWidget { font-size: 14px; }
            QCalendarWidget QToolButton { font-size: 14px; padding: 4px 8px; color: #1a5276; }
            QCalendarWidget QToolButton#qt_calendar_prevmonth,
            QCalendarWidget QToolButton#qt_calendar_nextmonth { font-size: 16px; }
            QCalendarWidget QWidget#qt_calendar_navigationbar { background-color: #d6eaf8; }
            QCalendarWidget QAbstractItemView { selection-background-color: #2980b9; selection-color: white; font-size: 13px; }
            QCalendarWidget QHeaderView::section { background-color: #aed6f1; font-weight: bold; }
        """)
        left_layout.addWidget(self.calendar)

        # 初始化按钮
        self.init_btn = QPushButton("初始化设置")
        self.init_btn.setStyleSheet(
            "background-color: #e67e22; color: white; font-size: 14px; font-weight: bold; "
            "padding: 8px; border-radius: 4px;")
        self.init_btn.clicked.connect(self._show_init_dialog)
        left_layout.addWidget(self.init_btn)

        # 查询历史按钮
        self.history_btn = QPushButton("查询历史数据")
        self.history_btn.setStyleSheet(
            "background-color: #2980b9; color: white; font-size: 14px; font-weight: bold; "
            "padding: 8px; border-radius: 4px;")
        self.history_btn.clicked.connect(self._show_history)
        left_layout.addWidget(self.history_btn)

        # 保存按钮
        self.save_btn = QPushButton("保存数据")
        self.save_btn.setStyleSheet(
            "background-color: #27ae60; color: white; font-size: 15px; font-weight: bold; "
            "padding: 10px; border-radius: 4px;")
        self.save_btn.clicked.connect(self._save_current_data)
        left_layout.addWidget(self.save_btn)

        # 导出按钮
        self.export_btn = QPushButton("导出 Excel")
        self.export_btn.setStyleSheet(
            "background-color: #238636; color: white; font-size: 15px; font-weight: bold; "
            "padding: 10px; border-radius: 4px;")
        self.export_btn.clicked.connect(self._export_excel)
        left_layout.addWidget(self.export_btn)

        # 当前日期信息
        self.date_info = QLabel("")
        self.date_info.setAlignment(Qt.AlignCenter)
        self.date_info.setStyleSheet("font-size: 13px; color: #555; padding: 4px;")
        left_layout.addWidget(self.date_info)

        main_layout.addWidget(left)

        # === 右侧：表格 ===
        right = QWidget()
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)

        self.title_label = QLabel("")
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet(
            "font-size: 20px; font-weight: bold; color: #1a5276; "
            "font-family: 'SimHei','黑体','Microsoft YaHei'; padding: 6px;")
        right_layout.addWidget(self.title_label)

        # 表格：12列 = 序号(0) + 部门(1) + 8数据列(2-9) + 归档类别(10) + 归档数据(11)
        self.table = QTableWidget()
        self.table.setRowCount(22)
        self.table.setColumnCount(12)
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(True)
        self.table.setStyleSheet("""
            QTableWidget { gridline-color: #333; font-size: 13px; }
            QTableWidget::item { padding: 2px; }
            QHeaderView::section { background-color: #d6eaf8; font-weight: bold; font-size: 13px; padding: 3px; border: 1px solid #333; }
        """)
        self.table.itemChanged.connect(self._on_item_changed)
        right_layout.addWidget(self.table)

        main_layout.addWidget(right, 1)

        self._setup_table_structure()

    def _setup_table_structure(self):
        """设置表格固定结构（表头、序号、部门、归档类别按组合并等）"""
        self._building_table = True

        # 列宽
        self.table.setColumnWidth(0, 40)   # 序号
        self.table.setColumnWidth(1, 60)   # 部门
        for c in range(2, 10):
            self.table.setColumnWidth(c, 85)  # 8个数据列
        self.table.setColumnWidth(10, 55)  # 归档类别
        self.table.setColumnWidth(11, 85)  # 归档数据

        # === Row 0: 标题行 (合并 col 0-11) ===
        self.table.setSpan(0, 0, 1, 12)
        item = QTableWidgetItem("电子卷宗随案生成中心工作量统计表")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("SimHei", 14, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        self.table.setItem(0, 0, item)

        # === Row 1: 日期 + 分类表头 ===
        self.table.setSpan(1, 0, 1, 2)
        self.date_cell = QTableWidgetItem("日期：")
        self.date_cell.setTextAlignment(Qt.AlignCenter)
        self.date_cell.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        self.date_cell.setFlags(Qt.NoItemFlags)
        self.table.setItem(1, 0, self.date_cell)

        # 立案 (col 2-4 merged)
        self.table.setSpan(1, 2, 1, 3)
        item = QTableWidgetItem("立案")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        self.table.setItem(1, 2, item)

        # 结案 (col 5-9 merged)
        self.table.setSpan(1, 5, 1, 5)
        item = QTableWidgetItem("结案")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        self.table.setItem(1, 5, item)

        # 归档/案 (col 10-11 merged, span 2 rows)
        self.table.setSpan(1, 10, 2, 2)
        item = QTableWidgetItem("归档/案")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        self.table.setItem(1, 10, item)

        # === Row 2: 列子表头 ===
        headers = ['序号', '部门', '接收/案', '扫描质检/页', '上传/册',
                   '接收/案', '扫描质检/页', '上传/册', '卷宗整理/册', '装订/册']
        for c, h in enumerate(headers):
            item = QTableWidgetItem(h)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Microsoft YaHei", 10))
            item.setFlags(Qt.NoItemFlags)
            self.table.setItem(2, c, item)

        # === Rows 3-18: 16个部门 ===
        for i, (seq, dept, arc) in enumerate(DEPARTMENTS):
            row = 3 + i
            # 序号
            item = QTableWidgetItem(str(seq))
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            self.table.setItem(row, 0, item)
            # 部门
            item = QTableWidgetItem(dept)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            self.table.setItem(row, 1, item)
            # 8个数据列 (col 2-9) 可编辑
            for c in range(2, 10):
                item = QTableWidgetItem("")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, c, item)
            # 归档类别列 (col 10) - 只读，后续按组合并
            item = QTableWidgetItem(arc)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            self.table.setItem(row, 10, item)
            # 归档数据列 (col 11) - 只读，后续按组合并（数据从组级读取）
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            self.table.setItem(row, 11, item)

        # === 归档类别列(col 10)和数据列(col 11)按组合并 ===
        for grp_name, start_row, count in ARCHIVE_GROUPS:
            grp_row = 3 + start_row
            # col 10: 类别标签合并（可编辑）
            self.table.setSpan(grp_row, 10, count, 1)
            cat_item = self.table.item(grp_row, 10)
            if cat_item:
                cat_item.setText(grp_name)
                cat_item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
                # 可编辑
                cat_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            # col 11: 归档数据合并（可编辑）
            self.table.setSpan(grp_row, 11, count, 1)
            data_item = self.table.item(grp_row, 11)
            if data_item:
                data_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)

        # === Row 19: 日合计 ===
        self.table.setSpan(19, 0, 1, 2)
        item = QTableWidgetItem("日合计")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        item.setBackground(QBrush(QColor(255, 255, 200)))
        self.table.setItem(19, 0, item)
        for c in range(2, 12):
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            item.setBackground(QBrush(QColor(255, 255, 200)))
            item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            self.table.setItem(19, c, item)

        # === Row 20: 周合计 ===
        self.table.setSpan(20, 0, 1, 2)
        item = QTableWidgetItem("周合计")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        item.setBackground(QBrush(QColor(200, 230, 255)))
        self.table.setItem(20, 0, item)
        for c in range(2, 12):
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            item.setBackground(QBrush(QColor(200, 230, 255)))
            item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            self.table.setItem(20, c, item)

        # === Row 21: 月合计 ===
        self.table.setSpan(21, 0, 1, 2)
        item = QTableWidgetItem("月合计")
        item.setTextAlignment(Qt.AlignCenter)
        item.setFont(QFont("Microsoft YaHei", 11, QFont.Bold))
        item.setFlags(Qt.NoItemFlags)
        item.setBackground(QBrush(QColor(200, 255, 200)))
        self.table.setItem(21, 0, item)
        for c in range(2, 12):
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(Qt.NoItemFlags)
            item.setBackground(QBrush(QColor(200, 255, 200)))
            item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            self.table.setItem(21, c, item)

        self.table.setRowCount(22)
        self._building_table = False

    def _get_prev_baselines(self, date_str):
        """
        返回初始基准值，供 calc_weekly_total/calc_monthly_total 使用。
        这些函数内部会自动累计本周/本月所有已填天数的数据。
        - 月初第一天：月合计基准归零
        """
        dt = datetime.strptime(date_str, '%Y%m%d')
        is_first_of_month = (dt.day == 1)

        # 初始基准值（用户通过初始化对话框设置的）
        init_baselines = load_baselines()
        weekly = list(init_baselines.get('weekly', [0] * COL_COUNT))

        if is_first_of_month:
            # 每月第一天：月合计基准归零
            monthly = [0] * COL_COUNT
        else:
            monthly = list(init_baselines.get('monthly', [0] * COL_COUNT))

        return {'weekly': weekly, 'monthly': monthly}

    def _on_date_clicked(self, qdate):
        """日历点击日期"""
        date_str = qdate.toString("yyyyMMdd")
        self._load_date(date_str)

    def _load_date(self, date_str):
        """加载指定日期的数据到表格"""
        # 切换日期前先保存当前数据
        if self.current_date and not self._building_table:
            day_data = self._read_current_data()
            set_day_data(self.data, self.current_date, day_data)

        self.current_date = date_str
        self._building_table = True

        # 动态更新基准值：从上一天数据计算周/月合计作为当天起点
        self.baselines = self._get_prev_baselines(date_str)

        # 更新标题
        self.title_label.setText(format_title_with_month(date_str))

        # 更新日期显示
        self.date_cell.setText(f"日期：{format_date_with_week(date_str)}")
        self.date_info.setText(f"当前选择: {format_date_with_week(date_str)}")

        # 加载数据 (9列: 0-7对应col 2-9, 8对应归档数据)
        day_data = get_day_data(self.data, date_str)
        for i in range(len(DEPARTMENTS)):
            row = 3 + i
            # 8个数据列 (col 2-9)
            for c in range(8):
                val = day_data[i][c]
                item = self.table.item(row, 2 + c)
                if item:
                    item.setText(str(val) if val is not None else "")

        # 加载归档数据（按组显示在合并单元格中）
        for grp_name, start_row, count in ARCHIVE_GROUPS:
            grp_row = 3 + start_row
            # 从组内所有行求和
            grp_sum = 0
            for j in range(start_row, start_row + count):
                if j < len(day_data) and day_data[j][8] is not None:
                    try:
                        grp_sum += int(day_data[j][8])
                    except (ValueError, TypeError):
                        pass
            item = self.table.item(grp_row, 11)
            if item:
                item.setText(str(grp_sum) if grp_sum else "")

        # 计算合计
        self._update_totals()

        # 根据日期设置单元格可编辑性（只能编辑当天及之前的日期）
        self._update_cell_editability(date_str)
        self._building_table = False

    def _update_cell_editability(self, date_str):
        """根据日期设置单元格可编辑性：只能编辑当天及之前的日期"""
        today = datetime.now().strftime('%Y%m%d')
        is_editable = (date_str <= today)

        if is_editable:
            edit_flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable
        else:
            edit_flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable

        # 设置数据行(3-18)的可编辑性
        for i in range(len(DEPARTMENTS)):
            row = 3 + i
            # 8个数据列 (col 2-9)
            for c in range(2, 10):
                item = self.table.item(row, c)
                if item:
                    item.setFlags(edit_flags)

        # 归档类别(col 10)和归档数据(col 11)按组设置
        for grp_name, start_row, count in ARCHIVE_GROUPS:
            grp_row = 3 + start_row
            # col 10: 归档类别
            cat_item = self.table.item(grp_row, 10)
            if cat_item:
                cat_item.setFlags(edit_flags)
            # col 11: 归档数据
            data_item = self.table.item(grp_row, 11)
            if data_item:
                data_item.setFlags(edit_flags)

    def _read_current_data(self):
        """从表格读取当前数据，返回 16行x9列 的二维列表"""
        day_data = []
        for i in range(len(DEPARTMENTS)):
            row = 3 + i
            row_data = []
            # 8个数据列 (col 2-9)
            for c in range(8):
                item = self.table.item(row, 2 + c)
                text = item.text().strip() if item else ""
                try:
                    row_data.append(int(text))
                except (ValueError, TypeError):
                    row_data.append(None)
            # 归档数据 (col 11) - 从合并单元格读取组值，分配到组第一行
            row_data.append(None)  # 先占位
            day_data.append(row_data)

        # 处理归档数据：从合并单元格读取组值
        for grp_name, start_row, count in ARCHIVE_GROUPS:
            grp_row = 3 + start_row
            item = self.table.item(grp_row, 11)
            text = item.text().strip() if item else ""
            try:
                grp_val = int(text) if text else None
            except (ValueError, TypeError):
                grp_val = None
            # 设置到组内第一行
            day_data[start_row][8] = grp_val
            # 其余行归档数据为None
            for j in range(1, count):
                idx = start_row + j
                if idx < len(day_data):
                    day_data[idx][8] = None

        return day_data

    def _on_item_changed(self, item):
        """表格内容改变时自动保存并更新合计"""
        if self._building_table or self._updating_totals:
            return
        row = item.row()
        col = item.column()
        # 数据行(3-18)的数据列(2-9) 或 归档类别(10) 或 归档数据(11)
        if row < 3 or row > 18:
            return
        if col not in list(range(2, 10)) + [10, 11]:
            return
        # 归档类别修改时，更新分组名称
        if col == 10:
            self._update_archive_group_names()
        # 保存数据
        day_data = self._read_current_data()
        set_day_data(self.data, self.current_date, day_data)
        # 更新合计
        self._update_totals()

    def _save_current_data(self):
        """保存当前日期的数据，并显示提示"""
        day_data = self._read_current_data()
        set_day_data(self.data, self.current_date, day_data)
        self.date_info.setText(f"✓ 已保存 {format_date_with_week(self.current_date)} 的数据")
        self.date_info.setStyleSheet("font-size: 13px; color: #27ae60; padding: 4px; font-weight: bold;")
        # 3秒后恢复默认样式
        from PyQt5.QtCore import QTimer
        QTimer.singleShot(3000, lambda: (
            self.date_info.setText(f"当前选择: {format_date_with_week(self.current_date)}"),
            self.date_info.setStyleSheet("font-size: 13px; color: #555; padding: 4px;")
        ))

    def _update_archive_group_names(self):
        """归档类别名称被编辑后，同步更新到categories中"""
        # 读取当前3个归档组的类别名
        new_names = []
        for grp_name, start_row, count in ARCHIVE_GROUPS:
            grp_row = 3 + start_row
            item = self.table.item(grp_row, 10)
            name = item.text().strip() if item else grp_name
            new_names.append(name)
        # 保存到categories（按日期存储）
        cats = load_categories()
        cats[self.current_date] = {
            'group_names': new_names
        }
        save_categories(cats)

    def _update_totals(self):
        """更新日合计、周合计、月合计"""
        self._updating_totals = True
        day_data = self._read_current_data()

        # 日合计：按列求和 (9列)
        daily = calc_daily_total(day_data)
        for c in range(8):
            item = self.table.item(19, 2 + c)
            if item:
                item.setText(str(daily[c]) if daily[c] else "")
        # 归档数据合计 (index 8 -> col 11)
        item = self.table.item(19, 11)
        if item:
            item.setText(str(daily[8]) if daily[8] else "")

        # 周合计 = 初始基准值 + 本周所有已填天数的合计（含当天）
        weekly = calc_weekly_total(self.data, self.current_date, self.baselines)
        for c in range(8):
            item = self.table.item(20, 2 + c)
            if item:
                item.setText(str(weekly[c]) if weekly[c] else "")
        # 周合计 col 11: 初始基准归档 + 本周归档日合计
        item = self.table.item(20, 11)
        if item:
            item.setText(str(weekly[8]) if weekly[8] else "")

        # 月合计 = 上月末合计 + 本月所有已填天数的合计（含当天）
        monthly = calc_monthly_total(self.data, self.current_date, self.baselines)
        for c in range(8):
            item = self.table.item(21, 2 + c)
            if item:
                item.setText(str(monthly[c]) if monthly[c] else "")
        # 月合计 col 11: 基准归档 + 本月归档日合计
        item = self.table.item(21, 11)
        if item:
            item.setText(str(monthly[8]) if monthly[8] else "")
        self._updating_totals = False

    def _show_init_dialog(self):
        """显示初始化对话框"""
        dlg = InitDialog(self.baselines, auto=False, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            self.baselines = dlg.get_baselines()
            save_baselines(self.baselines)
            self._update_totals()

    def _export_excel(self):
        """导出为Excel文件"""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        except ImportError:
            QMessageBox.warning(self, "提示", "需要安装 openpyxl 库。\n请运行: pip install openpyxl")
            return

        default_name = f"电子卷宗随案生成中心工作量统计表_{self.current_date}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel", os.path.join(os.path.expanduser('~'), 'Desktop', default_name),
            "Excel文件 (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        title_font = Font(name='SimHei', size=14, bold=True)
        header_font = Font(name='Microsoft YaHei', size=11, bold=True)
        data_font = Font(name='Microsoft YaHei', size=10)
        yellow_fill = PatternFill(start_color='FFFFF0', end_color='FFFFF0', fill_type='solid')
        blue_fill = PatternFill(start_color='C8E6FF', end_color='C8E6FF', fill_type='solid')
        green_fill = PatternFill(start_color='C8FFC8', end_color='C8FFC8', fill_type='solid')

        def set_cell(r, c, val, font=data_font, fill=None, align=center):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = font
            cell.alignment = align
            cell.border = border
            if fill:
                cell.fill = fill
            return cell

        # Row 3: 标题 (B3:M3)
        ws.merge_cells('B3:M3')
        set_cell(3, 2, format_title_with_month(self.current_date), title_font)

        # Row 4: 日期 + 分类
        ws.merge_cells('B4:C4')
        set_cell(4, 2, format_date_with_week(self.current_date), header_font)
        ws.merge_cells('D4:F4')
        set_cell(4, 4, '立案', header_font)
        ws.merge_cells('G4:K4')
        set_cell(4, 7, '结案', header_font)
        ws.merge_cells('L4:M5')
        set_cell(4, 12, '归档/案', header_font)

        # Row 5: 列头
        h5 = ['序号', '部门', '接收/案:', '扫描质检/页', '上传/册',
              '接收/案:', '扫描质检/页', '上传/册', '卷宗整理/册', '装订/册']
        for i, h in enumerate(h5):
            set_cell(5, 2 + i, h, header_font)

        # 数据行 (rows 6-21)
        day_data = self._read_current_data()
        for i, (seq, dept, arc) in enumerate(DEPARTMENTS):
            r = 6 + i
            set_cell(r, 2, seq)
            set_cell(r, 3, dept)
            # 8个数据列 (D-K, cols 4-11)
            for c in range(8):
                v = day_data[i][c]
                set_cell(r, 4 + c, v)
            # 归档类别 (L列) - 先填值，后面合并
            set_cell(r, 12, arc if arc else '', header_font)
            # 归档数据 (M列) - 先填值
            set_cell(r, 13, day_data[i][8])

        # 归档类别合并 (L列) 和 归档数据合并 (M列) 按组
        for grp_name, start_idx, count in ARCHIVE_GROUPS:
            start_r = 6 + start_idx
            end_r = start_r + count - 1
            if count > 1:
                # L列：归档类别合并
                ws.merge_cells(start_row=start_r, start_column=12,
                               end_row=end_r, end_column=12)
                # M列：归档数据合并
                ws.merge_cells(start_row=start_r, start_column=13,
                               end_row=end_r, end_column=13)

        # 日合计 (row 22)
        ws.merge_cells('B22:C22')
        daily = calc_daily_total(day_data)
        set_cell(22, 2, '日合计', header_font, yellow_fill)
        for c in range(8):
            set_cell(22, 4 + c, daily[c] if daily[c] else None, data_font, yellow_fill)
        set_cell(22, 12, None, data_font, yellow_fill)
        set_cell(22, 13, daily[8] if daily[8] else None, data_font, yellow_fill)

        # 周合计 (row 23)
        ws.merge_cells('B23:C23')
        weekly = calc_weekly_total(self.data, self.current_date, self.baselines)
        set_cell(23, 2, '周合计', header_font, blue_fill)
        for c in range(8):
            set_cell(23, 4 + c, weekly[c] if weekly[c] else None, data_font, blue_fill)
        set_cell(23, 12, None, data_font, blue_fill)
        set_cell(23, 13, weekly[8] if weekly[8] else None, data_font, blue_fill)

        # 月合计 (row 24)
        ws.merge_cells('B24:C24')
        monthly = calc_monthly_total(self.data, self.current_date, self.baselines)
        set_cell(24, 2, '月合计', header_font, green_fill)
        for c in range(8):
            set_cell(24, 4 + c, monthly[c] if monthly[c] else None, data_font, green_fill)
        set_cell(24, 12, None, data_font, green_fill)
        set_cell(24, 13, monthly[8] if monthly[8] else None, data_font, green_fill)

        # 设置列宽
        for col_letter in 'BCDEFGHIJKLM':
            ws.column_dimensions[col_letter].width = 13

        wb.save(path)
        QMessageBox.information(self, "导出成功", f"文件已保存到:\n{path}")

    def _mark_calendar_dates(self):
        """在日历上标记已有数据的日期（绿色背景）"""
        green_fmt = QTextCharFormat()
        green_fmt.setBackground(QColor(144, 238, 144))
        green_fmt.setForeground(QColor(0, 100, 0))
        green_fmt.setFontWeight(QFont.Bold)
        default_fmt = QTextCharFormat()
        cur_year = self.calendar.yearShown()
        cur_month = self.calendar.monthShown()
        import calendar as cal_mod
        days_in_month = cal_mod.monthrange(cur_year, cur_month)[1]
        for d in range(1, days_in_month + 1):
            qd = QDate(cur_year, cur_month, d)
            self.calendar.setDateTextFormat(qd, default_fmt)
        for date_str in self.data:
            try:
                dt = datetime.strptime(date_str, '%Y%m%d')
                if dt.year == cur_year and dt.month == cur_month:
                    day_data = self.data[date_str]
                    has_data = False
                    for row in day_data:
                        for v in row:
                            if v is not None:
                                has_data = True
                                break
                        if has_data:
                            break
                    if has_data:
                        qd = QDate(dt.year, dt.month, dt.day)
                        self.calendar.setDateTextFormat(qd, green_fmt)
            except Exception:
                pass

    def _show_history(self):
        """显示历史数据查询对话框"""
        # 先保存当前数据
        day_data = self._read_current_data()
        set_day_data(self.data, self.current_date, day_data)
        dlg = HistoryQueryDialog(self.data, self.baselines, self)
        dlg.exec_()


class InitDialog(QDialog):
    """初始化对话框：按初始化报表格式，表格形式编辑日/周/月合计的9个数据列"""
    def __init__(self, baselines, auto=False, parent=None):
        super().__init__(parent)
        self.baselines = baselines
        self.setWindowTitle("初始化设置")
        self.setGeometry(150, 100, 900, 280)
        self._auto = auto
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        if self._auto:
            info = QLabel("首次使用，请设置初始基准值（可直接修改表格中的数据）：")
        else:
            info = QLabel("设置周合计、月合计的初始值：")
        info.setStyleSheet("font-size: 14px; font-weight: bold; padding: 6px;")
        layout.addWidget(info)

        # 表格：2行(周合计/月合计) x 9列数据
        self.table = QTableWidget()
        self.table.setRowCount(2)
        self.table.setColumnCount(10)  # 1标签列 + 9数据列
        self.table.verticalHeader().setVisible(False)
        self.table.setWordWrap(True)
        self.table.setStyleSheet("""
            QTableWidget { gridline-color: #333; font-size: 13px; }
            QTableWidget::item { padding: 3px; }
            QHeaderView::section { background-color: #d6eaf8; font-weight: bold; font-size: 12px; padding: 3px; border: 1px solid #333; }
        """)

        # 列宽
        self.table.setColumnWidth(0, 70)
        for c in range(1, 10):
            self.table.setColumnWidth(c, 90)

        # 表头行
        headers = ['合计类型', '接收/案', '扫描质检/页', '上传/册',
                   '接收/案', '扫描质检/页', '上传/册', '卷宗整理/册', '装订/册', '归档/案']
        for c, h in enumerate(headers):
            item = QTableWidgetItem(h)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            item.setFlags(Qt.NoItemFlags)
            if c == 0:
                item.setBackground(QBrush(QColor(214, 234, 248)))
            self.table.setHorizontalHeaderItem(c, item)

        # 行标签和基准值
        weekly = self.baselines.get('weekly', [0] * COL_COUNT)
        monthly = self.baselines.get('monthly', [0] * COL_COUNT)
        row_labels = ['周合计', '月合计']
        row_colors = [QColor(200, 230, 255), QColor(200, 255, 200)]
        data_rows = [weekly, monthly]

        for r in range(2):
            # 标签列
            item = QTableWidgetItem(row_labels[r])
            item.setTextAlignment(Qt.AlignCenter)
            item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
            item.setFlags(Qt.NoItemFlags)
            item.setBackground(QBrush(row_colors[r]))
            self.table.setItem(r, 0, item)
            # 9个数据列
            for c in range(COL_COUNT):
                val = data_rows[r][c] if c < len(data_rows[r]) else 0
                item = QTableWidgetItem(str(val) if val else "")
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(QBrush(row_colors[r]))
                self.table.setItem(r, 1 + c, item)

        layout.addWidget(self.table)

        # 按钮
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("background-color: #238636; color: white; padding: 8px 20px; font-size: 14px;")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("padding: 8px 20px; font-size: 14px;")
        cancel_btn.clicked.connect(self.reject)
        if self._auto:
            cancel_btn.setText("跳过")
        btn_layout.addStretch()
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def get_baselines(self):
        """获取用户设置的基准值"""
        weekly = [0] * COL_COUNT
        monthly = [0] * COL_COUNT
        for c in range(COL_COUNT):
            try:
                weekly[c] = int(self.table.item(0, 1 + c).text().strip())
            except (ValueError, TypeError, AttributeError):
                pass
            try:
                monthly[c] = int(self.table.item(1, 1 + c).text().strip())
            except (ValueError, TypeError, AttributeError):
                pass
        return {'weekly': weekly, 'monthly': monthly}


class HistoryQueryDialog(QDialog):
    """历史数据查询对话框：按初始化报表格式显示，以周为单位"""
    def __init__(self, data, baselines=None, parent=None):
        super().__init__(parent)
        self.data = data
        self.baselines = baselines or {'weekly': [0] * COL_COUNT, 'monthly': [0] * COL_COUNT}
        self.parent_win = parent
        self.setWindowTitle("历史数据查询")
        self.setGeometry(80, 60, 1200, 700)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)

        # 顶部：月份筛选
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("选择月份："))
        self.month_combo = QComboBox()
        self.month_combo.setMinimumWidth(150)
        months = set()
        for date_str in sorted(self.data.keys()):
            try:
                dt = datetime.strptime(date_str, '%Y%m%d')
                months.add(f"{dt.year}年{dt.month}月")
            except Exception:
                pass
        self.month_combo.addItem("全部月份")
        for m in sorted(months, reverse=True):
            self.month_combo.addItem(m)
        self.month_combo.currentIndexChanged.connect(self._refresh_table)
        filter_layout.addWidget(self.month_combo)

        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet("font-size: 13px; color: #1a5276; font-weight: bold; padding: 4px;")
        filter_layout.addWidget(self.stats_label)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 滚动区域，每周一个表格
        from PyQt5.QtWidgets import QScrollArea
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setStyleSheet("QScrollArea { border: none; }")

        self.container = QWidget()
        self.container_layout = QVBoxLayout(self.container)
        self.container_layout.setContentsMargins(5, 5, 5, 5)
        self.container_layout.setSpacing(15)
        self.scroll.setWidget(self.container)
        layout.addWidget(self.scroll)

        # 底部按钮
        btn_layout = QHBoxLayout()
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("padding: 6px 20px;")
        close_btn.clicked.connect(self.close)
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        self._week_tables = []  # [(week_label, table, dates_in_week)]
        self._refresh_table()

    def _refresh_table(self):
        """刷新历史数据，按周分组显示"""
        # 清除旧内容
        while self.container_layout.count():
            child = self.container_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self._week_tables = []

        selected_month = self.month_combo.currentText()

        # 收集有数据的日期
        valid_dates = []
        for date_str in sorted(self.data.keys()):
            day_data = get_day_data(self.data, date_str)
            has_data = False
            for row in day_data:
                for v in row:
                    if v is not None:
                        has_data = True
                        break
                if has_data:
                    break
            if not has_data:
                continue
            if selected_month != "全部月份":
                try:
                    dt = datetime.strptime(date_str, '%Y%m%d')
                    month_str = f"{dt.year}年{dt.month}月"
                    if month_str != selected_month:
                        continue
                except Exception:
                    continue
            valid_dates.append(date_str)

        if not valid_dates:
            lbl = QLabel("暂无历史数据")
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet("font-size: 16px; color: #999; padding: 40px;")
            self.container_layout.addWidget(lbl)
            self.stats_label.setText("共 0 条记录")
            return

        # 按周分组（周一~周日）
        weeks = {}
        for date_str in valid_dates:
            dt = datetime.strptime(date_str, '%Y%m%d')
            monday = dt - timedelta(days=dt.weekday())
            week_key = monday.strftime('%Y%m%d')
            if week_key not in weeks:
                weeks[week_key] = []
            weeks[week_key].append(date_str)

        total_weeks = len(weeks)
        total_days = len(valid_dates)

        # 按周倒序创建表格
        for week_key in sorted(weeks.keys(), reverse=True):
            dates = sorted(weeks[week_key])
            dt_start = datetime.strptime(dates[0], '%Y%m%d')
            dt_end = datetime.strptime(dates[-1], '%Y%m%d')
            week_num = get_week_number(dates[0])
            week_label_text = f"{dt_start.year}年{dt_start.month}月 第{week_num}周  ({dates[0][:4]}-{dates[0][4:6]}-{dates[0][6:]} ~ {dates[-1][:4]}-{dates[-1][4:6]}-{dates[-1][6:]})"

            # 周标题
            week_lbl = QLabel(week_label_text)
            week_lbl.setStyleSheet(
                "font-size: 14px; font-weight: bold; color: #1a5276; "
                "padding: 4px 8px; background-color: #d6eaf8; border-radius: 3px;")
            self.container_layout.addWidget(week_lbl)

            # 每周一个表格：2行(周合计/月合计) × 10列(标签+9数据)
            tbl = QTableWidget()
            tbl.setRowCount(2)
            tbl.setColumnCount(10)
            tbl.verticalHeader().setVisible(False)
            tbl.setWordWrap(True)
            tbl.setStyleSheet("""
                QTableWidget { gridline-color: #333; font-size: 13px; }
                QTableWidget::item { padding: 3px; }
                QHeaderView::section { background-color: #d6eaf8; font-weight: bold; font-size: 12px; padding: 3px; border: 1px solid #333; }
            """)
            tbl.setColumnWidth(0, 70)
            for c in range(1, 10):
                tbl.setColumnWidth(c, 90)
            tbl.setMaximumHeight(120)

            # 表头
            headers = ['合计类型', '接收/案', '扫描质检/页', '上传/册',
                       '接收/案', '扫描质检/页', '上传/册', '卷宗整理/册', '装订/册', '归档/案']
            for c, h in enumerate(headers):
                item = QTableWidgetItem(h)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
                item.setFlags(Qt.NoItemFlags)
                tbl.setHorizontalHeaderItem(c, item)

            row_colors = [QColor(200, 230, 255), QColor(200, 255, 200)]
            row_labels = ['周合计', '月合计']

            # 取该周最后一天计算合计
            last_date = dates[-1]
            weekly = calc_weekly_total(self.data, last_date, self.baselines)
            monthly = calc_monthly_total(self.data, last_date, self.baselines)
            totals = [weekly, monthly]

            for r in range(2):
                # 标签
                item = QTableWidgetItem(row_labels[r])
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
                item.setFlags(Qt.NoItemFlags)
                item.setBackground(QBrush(row_colors[r]))
                tbl.setItem(r, 0, item)
                # 9个数据列
                for c in range(COL_COUNT):
                    val = totals[r][c]
                    item = QTableWidgetItem(str(val) if val else "0")
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(Qt.NoItemFlags)
                    item.setBackground(QBrush(row_colors[r]))
                    tbl.setItem(r, 1 + c, item)

            # 存储引用，用于双击跳转
            self._week_tables.append((week_key, tbl, dates))
            tbl.cellDoubleClicked.connect(self._on_week_cell_clicked)
            self.container_layout.addWidget(tbl)

            # 该周每日明细
            for date_str in reversed(dates):
                day_data = get_day_data(self.data, date_str)
                detail_parts = []
                for i, (seq, dept, arc) in enumerate(DEPARTMENTS):
                    row_sum = sum(v for v in day_data[i] if v is not None)
                    if row_sum > 0:
                        detail_parts.append(f"{dept}:{row_sum}")
                detail = ", ".join(detail_parts) if detail_parts else "无数据"
                detail_lbl = QLabel(f"  {format_date_with_week(date_str)}: {detail}")
                detail_lbl.setStyleSheet("font-size: 12px; color: #555; padding: 1px 8px;")
                self.container_layout.addWidget(detail_lbl)

            # 分隔线
            line = QFrame()
            line.setFrameShape(QFrame.HLine)
            line.setStyleSheet("color: #ccc;")
            self.container_layout.addWidget(line)

        self.container_layout.addStretch()
        self.stats_label.setText(f"共 {total_weeks} 周 / {total_days} 天数据")

    def _on_week_cell_clicked(self, row, col):
        """双击周表格跳转到该周第一天"""
        tbl = self.sender()
        for week_key, t, dates in self._week_tables:
            if t is tbl and dates:
                date_str = dates[0]
                if self.parent_win:
                    dt = datetime.strptime(date_str, '%Y%m%d')
                    qd = QDate(dt.year, dt.month, dt.day)
                    self.parent_win.calendar.setSelectedDate(qd)
                    self.parent_win.calendar.setCurrentPage(dt.year, dt.month)
                    self.parent_win._load_date(date_str)
                    self.close()
                break


if __name__ == "__main__":
    app = QApplication(sys.argv)
    try:
        if hasattr(Qt, 'AA_EnableHighDpiScaling'):
            QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
        if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
            QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    except Exception:
        pass

    window = DailyReportWindow()
    window.show()
    window.raise_()
    window.activateWindow()
    sys.exit(app.exec_())

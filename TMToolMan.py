# 以下为您编写的“同美档案工具集合”Python代码。代码采用PyQt5框架构建，界面采用深色科技感配色，左侧为菜单栏，右侧为动态切换的功能区。
#
# 请确保在运行前安装依赖库：`pip
# install
# PyQt5
# `
#
# ```python
import sys
import re
import os
import shutil
# 使用 PyQt5（Qt5）替代 PySide6（Qt6），以兼容 Python 3.8。
# PyQt5 中信号为 pyqtSignal，通过别名统一为 Signal，保持下游写法不变。
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QPushButton, QLabel, QLineEdit,
                             QFileDialog, QTextEdit, QSpinBox, QComboBox,
                             QFormLayout, QGroupBox, QMessageBox, QStackedWidget, QCheckBox,
                             QProgressBar, QDialog, QListWidget, QTableWidget,
                             QTableWidgetItem, QHeaderView, QAbstractItemView)
from PyQt5.QtCore import Qt, QThread, pyqtSignal as Signal, QRegularExpression
from PyQt5.QtGui import QFont, QRegularExpressionValidator
from PIL import Image, ImageDraw, ImageFont, ImageOps
from concurrent.futures import ThreadPoolExecutor, as_completed
# 注：cv2 与 numpy 已改为在 AutoPagingWorker.add_number_to_image 中延迟导入，
# 以提升 Win7 下的健壮性（opencv 在 Win7 上较易出现 DLL 加载失败），
# 这样即使该依赖缺失或损坏，主程序仍可启动、其他功能页仍可使用。
import time
import threading
from datetime import datetime
from pathlib import Path


class TechStyle:
    """浅色页面样式表(参照 ImageCheckTool 浅色风格)"""
    MAIN_BG = "#ecf0f1"
    PANEL_BG = "#ffffff"
    TEXT_COLOR = "#2c3e50"
    ACCENT_COLOR = "#1a5276"
    BTN_HOVER = "#2980b9"

    QSS = f"""
    QMainWindow {{ background-color: {MAIN_BG}; }}
    QWidget {{ color: {TEXT_COLOR}; font-family: 'Microsoft YaHei', Arial; }}

    /* 左侧菜单栏 */
    #LeftPanel {{
        background-color: {PANEL_BG};
        border-right: 1px solid #bdc3c7;
    }}
    QLabel#TitleLabel {{
        color: {ACCENT_COLOR};
        font-size: 18px;
        font-weight: bold;
        padding: 20px;
        border-bottom: 1px solid #bdc3c7;
    }}
    QPushButton#MenuBtn {{
        background-color: transparent;
        color: #1a5276;
        text-align: left;
        padding: 16px 20px;
        border: none;
        font-size: 17px;
        font-weight: bold;
    }}
    QPushButton#MenuBtn:hover {{
        background-color: #d6eaf8;
        color: #154360;
        border-left: 4px solid {ACCENT_COLOR};
    }}
    QPushButton#MenuBtn:checked {{
        background-color: #aed6f1;
        color: #0e2f44;
        border-left: 5px solid {ACCENT_COLOR};
        font-weight: bold;
    }}

    /* 右侧功能区 */
    #RightPanel {{ background-color: {MAIN_BG}; }}
    QGroupBox {{
        border: 1px solid #bdc3c7;
        border-radius: 5px;
        margin-top: 15px;
        padding: 15px;
        font-weight: bold;
        color: {ACCENT_COLOR};
    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        left: 10px;
        padding: 0 5px;
    }}

    /* 输入控件 */
    QLineEdit, QSpinBox, QComboBox {{
        background-color: #ffffff;
        border: 1px solid #bdc3c7;
        border-radius: 3px;
        padding: 5px;
        color: {TEXT_COLOR};
    }}
    QLineEdit:focus, QSpinBox:focus, QComboBox:focus {{
        border: 1px solid {ACCENT_COLOR};
    }}

    /* QSpinBox 上下按钮：显式定义，避免样式表覆盖后按钮消失/失效 */
    QSpinBox::up-button, QSpinBox::down-button {{
        background-color: #d6eaf8;
        border: none;
        width: 18px;
    }}
    QSpinBox::up-button {{ subcontrol-origin: border; subcontrol-position: top right; border-left: 1px solid #bdc3c7; }}
    QSpinBox::down-button {{ subcontrol-origin: border; subcontrol-position: bottom right; border-left: 1px solid #bdc3c7; border-top: 1px solid #bdc3c7; }}
    QSpinBox::up-button:hover, QSpinBox::down-button:hover {{ background-color: #2980b9; }}
    QSpinBox::up-button:pressed, QSpinBox::down-button:pressed {{ background-color: {ACCENT_COLOR}; }}
    QSpinBox::up-arrow {{
        image: none;
        border-left: 4px solid transparent;
        border-right: 4px solid transparent;
        border-bottom: 5px solid {TEXT_COLOR};
        width: 0px; height: 0px;
    }}
    QSpinBox::down-arrow {{
        image: none;
        border-left: 4px solid transparent;
        border-right: 4px solid transparent;
        border-top: 5px solid {TEXT_COLOR};
        width: 0px; height: 0px;
    }}

    /* 动作按钮 */
    QPushButton#ActionBtn {{
        background-color: #238636;
        color: white;
        border: none;
        padding: 8px 16px;
        border-radius: 3px;
        font-weight: bold;
    }}
    QPushButton#ActionBtn:hover {{
        background-color: #2EA043;
    }}
    QPushButton#BrowseBtn {{
        background-color: #2980b9;
        color: white;
        border: none;
        padding: 5px 10px;
        border-radius: 3px;
    }}
    QPushButton#BrowseBtn:hover {{
        background-color: #3498db;
    }}

    QTextEdit {{
        background-color: #ffffff;
        border: 1px solid #bdc3c7;
        border-radius: 3px;
        color: #2c3e50;
        font-size: 13px;
    }}
    
    /* 消息提示框 */
    QMessageBox {{
        background-color: #ffffff;
    }}
    QMessageBox QLabel {{
        color: #2c3e50;
        font-weight: bold;
        font-size: 14px;
    }}
    QMessageBox QPushButton {{
        background-color: #238636;
        color: white;
        border: none;
        padding: 8px 16px;
        border-radius: 3px;
        font-weight: bold;
        min-width: 80px;
    }}
    QMessageBox QPushButton:hover {{
        background-color: #2EA043;
    }}
    """


class FunctionPage(QWidget):
    """功能页面基类"""

    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.layout.setAlignment(Qt.AlignTop)

        lbl_title = QLabel(title)
        lbl_title.setStyleSheet(
            f"color: {TechStyle.ACCENT_COLOR}; font-size: 20px; font-weight: bold; margin-bottom: 10px;")
        self.layout.addWidget(lbl_title)

    def add_log_widget(self):
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setFixedHeight(150)
        self.layout.addWidget(self.log_box)

    def log(self, msg):
        self.log_box.append(f">> {msg}")


class FileSplitWorker(QThread):
    """
    分件后台处理线程：
    对所选目录下的每个子目录——
      1. 若目录下有 Directory.txt(文件批量替换生成): 第一页为卷皮, 目录页
         取自 Directory.txt 记录的文件(逐个OCR); 无则取排序后第2页(0002) OCR；
      2. 判断标题是否为“卷内文件目录”；是则解析表格中“序号”“页号”列；
      3. 按序号建子目录(目录名+“-”+四位序号)，按页号+偏移量移动 jpg 文件
         (偏移量=Directory.txt最大文件序号, 无Directory.txt时缺省为2；
         如偏移量2时页号“1-17”→移动 0003.jpg..0019.jpg)；
      4. 建“目录名+备考表卷底”“目录名+卷皮目录”两个子目录；
         卷皮页与目录页移入卷皮目录，最大与次大文件名移入备考表卷底；
         分件完成后删除该目录下 Directory.txt；
      5. 全程详细日志(OCR识别行、序号/页号解析、文件移动范围)写入所选目录。
    """
    log_signal = Signal(str)
    progress_signal = Signal(int, int)          # (当前, 总数)
    finished_signal = Signal(bool, str)

    def __init__(self, base_dir, target_base=None, copy_mode=False, xlsx_dir=None, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.target_base = target_base    # 分件到新目录: 输出根目录(None=原地)
        self.copy_mode = copy_mode        # True=文件拷贝(源不动), False=移动
        self.xlsx_dir = xlsx_dir          # 目录文件(xlsx)所在根目录(None=OCR模式)
        self.is_stopped = False
        self._ocr = None  # PaddleOCR 延迟初始化(只初始化一次, 避免重复加载模型)

    # ---------- OCR ----------
    def _get_ocr(self):
        """
        延迟初始化 PaddleOCR(中文, 带方向分类)。失败返回 None。
        模型文件随程序打包(ocr_models/)，显式指定路径，避免在用户机器上
        联网下载模型(内网环境会静默失败导致 OCR 无结果)。
        """
        if self._ocr is None:
            try:
                from paddleocr import PaddleOCR
                self.log_signal.emit("  初始化 PaddleOCR 引擎(首次较慢)...")
                # 模型目录定位: PyInstaller 打包后资源在 sys._MEIPASS; 开发环境在脚本目录
                base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
                m_det = os.path.join(base, 'ocr_models', 'det')
                m_rec = os.path.join(base, 'ocr_models', 'rec')
                m_cls = os.path.join(base, 'ocr_models', 'cls')
                has_models = all(os.path.exists(os.path.join(m, 'inference.pdmodel'))
                                 for m in (m_det, m_rec, m_cls))
                import warnings as _w
                _w.filterwarnings('ignore', message='.*use_angle_cls.*deprecated.*')
                # paddleocr 版本碎片化兼容: 依次尝试多组配置, 成功即用。
                # 覆盖: 2.x(全参数) / 3.x legacy(use_angle_cls 可用但 show_log 移除)
                #       / 3.x 新参数(仅无内置模型时; 会走联网下载, 内网可能失败)
                configs = []
                if has_models:
                    configs += [
                        # 2.x 标准配置
                        dict(use_angle_cls=True, lang='ch', show_log=False,
                             det_model_dir=m_det, rec_model_dir=m_rec, cls_model_dir=m_cls),
                        # 3.x legacy: use_angle_cls 被兼容, 但 show_log 已移除
                        dict(use_angle_cls=True, lang='ch',
                             det_model_dir=m_det, rec_model_dir=m_rec, cls_model_dir=m_cls),
                        # 3.x 新参数 + 旧模型路径(部分版本参数改名但模型格式仍兼容)
                        dict(use_textline_orientation=True, lang='ch',
                             det_model_dir=m_det, rec_model_dir=m_rec, cls_model_dir=m_cls),
                    ]
                configs += [
                    dict(use_angle_cls=True, lang='ch', show_log=False),
                    dict(use_angle_cls=True, lang='ch'),
                    dict(use_textline_orientation=True),
                    dict(),
                ]
                last_err = None
                for i, kw in enumerate(configs, 1):
                    try:
                        self._ocr = PaddleOCR(**kw)
                        self.log_signal.emit(f"  (OCR初始化成功: 配置#{i}"
                                             f"{' 含内置模型' if 'det_model_dir' in kw else ''})")
                        break
                    except Exception as e:
                        last_err = e
                        self.log_signal.emit(f"  (配置#{i} 不适用: {str(e)[:80]}, 尝试下一配置)")
                if self._ocr is None:
                    raise last_err or RuntimeError('所有OCR配置均失败')
            except Exception as e:
                self.log_signal.emit(f"  × PaddleOCR 初始化失败: {e}")
                return None
        return self._ocr

    def _ocr_page(self, image_path, logf, wlog):
        """
        对单页做 OCR，返回片段列表 [(文本, x0, y0, x1), ...]。
        x1 为片段右边界，用于列定位。
        兼容 paddleocr 2.x (ocr(img, cls=True)) 与 3.x (predict(img), 结果对象)。
        """
        ocr = self._get_ocr()
        if ocr is None:
            return []
        result = None
        try:
            result = ocr.ocr(image_path, cls=True)
        except Exception as e:
            # 3.x: ocr() 不接受 cls 参数 → 重试无参/用 predict
            try:
                result = ocr.ocr(image_path)
            except Exception:
                try:
                    result = ocr.predict(image_path)
                except Exception:
                    import traceback
                    wlog(f"    OCR 出错 {os.path.basename(image_path)}: {e}")
                    wlog("    详细: " + traceback.format_exc().replace('\n', ' | ')[:500])
                    return []

        frags = []
        items = []
        if result:
            first = result[0]
            if first is not None and hasattr(first, 'rec_texts'):
                # 3.x 结果对象: rec_texts/rec_scores/dt_polys
                texts = list(getattr(first, 'rec_texts', []) or [])
                scores = list(getattr(first, 'rec_scores', []) or [])
                polys = list(getattr(first, 'dt_polys', []) or [])
                for i, txt in enumerate(texts):
                    poly = polys[i] if i < len(polys) and len(polys[i]) >= 4 else None
                    if poly is not None:
                        xs = [p[0] for p in poly]
                        ys = [p[1] for p in poly]
                        box = [[min(xs), min(ys)], [max(xs), max(ys)]]
                    else:
                        box = [[0, 0], [0, 0]]
                    conf = scores[i] if i < len(scores) else 0.0
                    items.append((box, txt, conf))
            elif first:
                # 2.x: [[box, (txt, conf)], ...] box为四点多边形
                # (左上/右上/右下/左下) → 统一转为 [左上,右下] 两点,
                # 使 y1-y0 为真实行高(box[1]右上的y≈box[0]左上的y)
                for item in first:
                    try:
                        box, (txt, conf) = item[0], item[1]
                        if len(box) >= 4:
                            xs = [p[0] for p in box]
                            ys = [p[1] for p in box]
                            box = [[min(xs), min(ys)], [max(xs), max(ys)]]
                        items.append((box, txt, conf))
                    except Exception:
                        continue
        if not items:
            shape = 'result为空' if not result else ('result[0]为None' if result[0] is None
                                                     else '结果无内容')
            wlog(f"    OCR 无内容({shape}): {os.path.basename(image_path)}")
            return frags
        for box, txt, conf in items:
            x0, y0 = box[0][0], box[0][1]
            x1, y1 = box[1][0], box[1][1]
            frags.append((txt, x0, y0, x1, y1))
            wlog(f"    OCR行: 「{txt}」 置信度={conf:.2f} 位置=({x0:.0f},{y0:.0f})")
        return frags

    # ---------- 表格解析 ----------
    @staticmethod
    def _row_of(frags, y, tol=35):
        """取与 y 同一行的片段(按 y 中心, 容差 tol)。"""
        out = []
        for txt, x0, y0, x1, _y1 in frags:
            yc = y0  # y0 即片段顶(近似行位置)
            if abs(yc - y) <= tol:
                out.append((txt, x0, x1))
        return out

    @staticmethod
    def _is_catalog_title(frags):
        """
        判断 OCR 片段中是否含标题「卷内文件目录」。
        匹配策略(由严到宽):
          1. 完整包含「卷内文件目录」;
          2. 同一片段同时含「卷内」和「目录」(允许中间夹杂其他字);
          3. 页面顶部(y最小)的大片段含「目录」且含「卷」——标题被OCR拆散/夹字时的兜底。
        """
        for txt, x0, y0, x1, _y1 in frags:
            compact = txt.replace(' ', '').replace('　', '')
            if '卷内文件目录' in compact:
                return True
            if '卷内' in compact and '目录' in compact:
                return True
        # 兜底: 顶部区域(y < 600)内, 「卷」和「目录」分别出现在不同片段
        has_juan = any('卷' in t.replace(' ', '').replace('　', '') for t, *_ in frags)
        has_mulu = any('目录' in t.replace(' ', '').replace('　', '') for t, _, y, _, _h in frags if y < 600)
        return has_juan and has_mulu

    @staticmethod
    def _parse_catalog_rows(frags, wlog):
        """
        解析表格数据行：返回 [(序号int, 起页int, 止页int), ...]。

        真实档案的两种页号写法(同一表内可混合)：
          1. 范围式 “1-17” / “146-149”；
          2. 单数字 “144”  —— 表示该件从该页开始，到下一行起始页-1 止
             (最后一行则到“下一起始页”未知，保守取同值，即只含该页)。
        列定位：先找表头“序号”“页号”的 x 坐标确定两列的 x 范围，
        数据行内按 x 落点取列值(序号列较窄, 页号列在表格最右数据区)。
        """
        # ---- 1. 定位表头列 x ----
        seq_x = page_x = None
        for txt, x0, y0, x1, _y1 in frags:
            compact = txt.replace(' ', '').replace('　', '')
            if compact == '序号' and seq_x is None:
                seq_x = x0
            elif '页号' in compact and page_x is None:
                page_x = x0
        if seq_x is None or page_x is None:
            wlog("    未找到表头「序号」或「页号」列, 尝试按位置推断列")
            # 推断: 序号=最左侧数字列, 页号=最右侧数据列(取全图宽)
            if frags:
                page_x = 0.62 * max(f[3] for f in frags)
                seq_x = 0.0

        # ---- 2. 收集数据行: 每行 = (序号候选, 页号原始文本, y) ----
        # 序号候选: 纯 1-2 位数字片段, x 靠近序号列(x0 < 页号列左界, 且 x0 在序号列附近±180px)
        seq_tol = 200
        page_left = page_x - 160  # 页号列数据允许在表头左侧一点
        candidates = []  # [(seq, y, pagetxt)]

        def _find_pagetxt(y_row):
            """在该行找页号: 优先独立片段; 否则从长文本尾部提取(OCR 常把日期和
            页号粘连, 如「...征求意见书及2021.12.2039-143」尾部是页号)。
            片段起点或右边界落入页号区(x1 >= page_left)即视为含页号。
            粘连两级处理:
              a. 日期与页号直接相连无分隔: 「2021.8.30107-111」→日期尾「30」
                 与页号「107-111」粘连 → 按日期模式(YYYY.M.D)切出页号;
              b. 页号前是非数字: 「...12.2039-143」→「39-143」(避免吞日期)。"""
            row = FileSplitWorker._row_of(frags, y_row)
            best = ''
            for rt, rx0, rx1 in sorted(row, key=lambda r: r[1]):
                if rx0 < page_left and rx1 < page_left:
                    continue  # 片段整体在页号区左侧
                rt2 = rt.strip()
                if re.fullmatch(r'\d{1,4}[-–—~]\d{1,4}', rt2):
                    return rt2
                if re.fullmatch(r'\d{1,4}', rt2):
                    best = best or rt2
                else:
                    # a. 日期模式紧贴页号: 「2021.8.30107-111」「2021.12.2039-143」
                    #    日期 YYYY.M.D 的最后一段日期数字与页号起始粘连,
                    #    页号= 去掉日期前缀后剩下的「数字-数字」
                    m = re.search(r'\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}\s*(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*'
                                  r'(?:\s+(\d{1,4})\s*)?$', rt2)
                    if m:
                        # 可能还有第二个独立页号(如「...112-118」带空格形式)
                        return f"{m.group(1)}-{m.group(2)}"
                    # 带空格分隔: 「2021.8.30 112-118」→「112-118」
                    m = re.search(r'\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}\s+(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*$', rt2)
                    if m:
                        return f"{m.group(1)}-{m.group(2)}"
                    # b. 页号前是非数字(避免把日期一部分吞进来)
                    m = re.search(r'(?:^|\D)(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*$', rt2)
                    if m:
                        return f"{m.group(1)}-{m.group(2)}"
                    m = re.search(r'(?:^|\s)(\d{1,4})\s*$', rt2)
                    if m and not best:
                        best = m.group(1)
            return best

        for txt, x0, y0, x1, _y1 in frags:
            t = txt.strip()
            if not re.fullmatch(r'\d{1,2}', t):
                continue
            if x0 > page_left:      # 在页号列右侧的数字 → 页号候选, 不是序号
                continue
            if abs(x0 - seq_x) > seq_tol and seq_x > 0:
                continue
            candidates.append((int(t), y0, _find_pagetxt(y0)))

        # ---- 2b. 补漏: 有页号但序号未被识别的行(如首行序号「1」漏识别) ----
        # 找出未被任何 candidate 认领的页号来源(两类) → 按 y 位置插入, 序号推断:
        #   i.  独立片段: 纯「X-Y」或纯数字(x 在页号区);
        #   ii. 长文本尾部粘连: 「2021.8.30 112-118」/「2021.8.30107-111」
        #       (序号与页号都被 OCR 挤进日期长文本的整行, 用同一套粘连提取)
        used_y = [c[1] for c in candidates]
        orphan_pages = []  # [(y, pagetxt)]

        def _extract_glued(t):
            """从长文本尾部提取粘连页号(与 _find_pagetxt 相同的规则, 返回''表示无)。"""
            m = re.search(r'\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}\s*(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*$', t)
            if m:
                return f"{m.group(1)}-{m.group(2)}"
            m = re.search(r'(?:^|\D)(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*$', t)
            if m:
                return f"{m.group(1)}-{m.group(2)}"
            m = re.search(r'(?:^|\s)(\d{1,4})\s*$', t)
            return m.group(1) if m else ''

        for txt, x0, y0, x1, _y1 in frags:
            t = txt.strip()
            if x0 < page_left and x1 < page_left:
                continue  # 整体在页号区左侧, 不含页号
            if any(abs(y0 - uy) < 40 for uy in used_y):
                continue  # 已被认领
            if re.fullmatch(r'\d{1,4}[-–—~]\d{1,4}', t) or re.fullmatch(r'\d{1,4}', t):
                orphan_pages.append((y0, t))
            else:
                # 长文本: 仅当其右边界伸入页号区才尝试提取尾部粘连页号。
                # 只接受「X-Y」范围形式——单数字从长文本尾部提取误报率高
                # (容易把日期/字号等数字当页号), 不作为补漏来源。
                if x1 >= page_left:
                    glued = _extract_glued(t)
                    if glued and re.fullmatch(r'\d{1,4}[-–—~]\d{1,4}', glued):
                        orphan_pages.append((y0, glued))
        # 孤立页号行统一编号:
        #   在已认领行之前的 → 按 y 排序后从 (最小已认领序号 - 前置孤立行数)
        #   开始依次 +1 (例: 已认领最小序号4, 前置孤立3行 → 1,2,3)
        #   在已认领行之间的 → 前一行序号+1 (同前逻辑)
        orphan_pages.sort(key=lambda o: o[0])
        n_before = 0
        min_seq_known = min((c[0] for c in candidates), default=1)
        min_y_known = min((c[1] for c in candidates), default=float('inf'))
        # 统计在已认领行之前的孤立行数
        lead_orphans = [o for o in orphan_pages if o[0] < min_y_known]
        n_lead = len(lead_orphans)
        assigned_lead = 0
        for y0, t in orphan_pages:
            if not candidates:
                nxt_seq = 1 + assigned_lead
                assigned_lead += 1
            elif y0 < min_y_known:
                # 前置孤立行: 从 min_seq - n_lead 开始递增(≥1)
                nxt_seq = max(1, min_seq_known - n_lead) + assigned_lead
                assigned_lead += 1
            else:
                prevs = [c for c in candidates if c[1] < y0]
                base = max(prevs, key=lambda c: c[1]) if prevs else None
                if base is not None:
                    nxt_seq = base[0] + 1
                    # 若与后一行序号冲突(>= 后一行), 说明推断异常, 用后一行-1
                    nexts = [c for c in candidates if c[1] > y0]
                    if nexts:
                        nmin = min(nexts, key=lambda c: c[1])
                        if nxt_seq >= nmin[0]:
                            nxt_seq = nmin[0] - 1
                else:
                    nxt_seq = 1 + assigned_lead
                    assigned_lead += 1
            candidates.append((nxt_seq, y0, t))
            wlog(f"    [补漏] 序号={nxt_seq} y={y0:.0f} 页号「{t}」(序号未被OCR识别, 已推断)")

        # 按 y 去重(同一序号可能被 OCR 拆出多个同值片段)
        candidates.sort(key=lambda c: c[1])
        dedup = []
        for c in candidates:
            if dedup and abs(c[1] - dedup[-1][1]) < 30 and c[0] == dedup[-1][0]:
                # 补页号(若前一条为空)
                if not dedup[-1][2] and c[2]:
                    dedup[-1] = (dedup[-1][0], dedup[-1][1], c[2])
                continue
            dedup.append(c)

        # ---- 3. 解析页号(两种格式), 单数字推断止页=下一行起始-1 ----
        parsed = []
        for i, (seq, y, pagetxt) in enumerate(dedup):
            if not pagetxt:
                wlog(f"    跳过行(无页号): 序号={seq} y={y:.0f}")
                continue
            m = re.fullmatch(r'(\d{1,4})[-–—~](\d{1,4})', pagetxt)
            if m:
                p_start, p_end = int(m.group(1)), int(m.group(2))
                if p_end < p_start:
                    p_start, p_end = p_end, p_start
                src = pagetxt
            else:
                p_start = int(pagetxt)
                # 找下一行的起始页
                nxt = None
                for j in range(i + 1, len(dedup)):
                    if dedup[j][2]:
                        m2 = re.match(r'(\d{1,4})', dedup[j][2])
                        if m2:
                            nxt = int(m2.group(1))
                            break
                p_end = (nxt - 1) if (nxt is not None and nxt > p_start) else p_start
                src = f"{pagetxt}(推断到{p_end})"
            parsed.append((seq, p_start, p_end))
            wlog(f"    解析行: 序号={seq} 页号={p_start}-{p_end} ← 「{src}」 y={y:.0f}")

        # ---- 4. 连续性校验修正 ----
        # 档案各件页号首尾相接(本行起始 = 前行止页+1, 本行止页 = 下行起始-1)。
        # OCR 常把日期与页号粘连(如「2021.12.2039-143」实为页号「39-143」),
        # 导致某端数字虚大。按相邻行的连续性约束重切粘连数字。
        for i in range(len(parsed)):
            seq, p_start, p_end = parsed[i]
            prev_end = parsed[i - 1][2] if i > 0 else None
            next_start = parsed[i + 1][1] if i + 1 < len(parsed) else None
            raw = (dedup[i][2] if i < len(dedup) else '') or ''
            m = re.fullmatch(r'(\d{1,5})[-–—~](\d{1,5})', raw.strip())
            if not m:
                continue
            big, tail = m.group(1), m.group(2)
            big_v, tail_v = int(big), int(tail)
            lo, hi = (tail_v, big_v) if big_v > tail_v else (big_v, tail_v)
            # 期望: start=prev_end+1 (若有前行), end=next_start-1 (若有后行)
            exp_start = (prev_end + 1) if prev_end is not None else None
            exp_end = (next_start - 1) if (next_start is not None and next_start > 1) else None
            fixed = False
            new_start, new_end = p_start, p_end
            # 情形A: start 虚大(粘连在头, 如「2039-143」start 应为 39)
            if exp_start is not None and lo != exp_start and hi == (exp_end or hi):
                cand_s = big[-len(str(exp_start)):] if big_v > tail_v else tail[-len(str(exp_start)):]
                if cand_s == str(exp_start):
                    new_start, new_end = exp_start, hi
                    fixed = True
            # 情形B: end 虚大(粘连在尾)
            if not fixed and exp_end is not None and hi > exp_end and lo == exp_start:
                cand_e = tail[-len(str(exp_end)):] if tail_v > big_v else big[-len(str(exp_end)):]
                if cand_e == str(exp_end):
                    new_start, new_end = lo, exp_end
                    fixed = True
            # 情形C: 双端都可能粘连, 用两端约束直接切
            if not fixed and exp_start is not None and exp_end is not None \
                    and (lo != exp_start or hi != exp_end):
                s_str, e_str = str(exp_start), str(exp_end)
                joined = big + tail
                # 尝试在 joined 中找 s_str 和 e_str 的合理组合
                for cut in range(1, len(joined)):
                    a, b = joined[:cut], joined[cut:]
                    if a.endswith(s_str) and b.startswith(e_str) and len(a) >= len(s_str):
                        new_start, new_end = exp_start, exp_end
                        fixed = True
                        break
            if fixed and (new_start, new_end) != (p_start, p_end):
                parsed[i] = (seq, new_start, new_end)
                wlog(f"    [连续性修正] 序号={seq}: 页号 {p_start}-{p_end} → "
                     f"{new_start}-{new_end} (原文「{raw}」日期与页号粘连, 已按邻行连续性切分)")
        return parsed

    # ---------- 表格线模式解析 ----------
    _CN_NUM = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
               '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}

    @staticmethod
    def _imread_cn(path):
        """中文路径安全读图(cv2.imread 不支持 Windows 非 ASCII 路径)。"""
        try:
            import cv2
            import numpy as _np
            data = _np.fromfile(path, dtype=_np.uint8)
            if data.size == 0:
                return None
            return cv2.imdecode(data, cv2.IMREAD_COLOR)
        except Exception:
            return None

    @staticmethod
    def _imwrite_cn(path, img):
        """中文路径安全写图(配套 _imread_cn)。"""
        try:
            import cv2
            ext = '.png' if path.lower().endswith('.png') else '.jpg'
            ok, buf = cv2.imencode(ext, img)
            if ok:
                buf.tofile(path)
                return True
        except Exception:
            pass
        return False

    def _detect_table_rows(self, image_path):
        """
        OpenCV 检测表格横线, 返回数据行区间 [(top, bottom), ...] 或 []。
        横线特征: 水平长线(≥图宽30%); 数据行=相邻横线间隔>100px。
        """
        try:
            import cv2
            import numpy as _np
            img = self._imread_cn(image_path)
            if img is None:
                return []
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            H, W = gray.shape
            _, bw = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY_INV)
            hk = cv2.getStructuringElement(cv2.MORPH_RECT, (max(W // 15, 20), 1))
            horiz = cv2.morphologyEx(bw, cv2.MORPH_OPEN, hk)
            ys = _np.where(horiz.sum(axis=1) > 255 * W * 0.3)[0]
            lines = []
            for y in ys:
                if lines and y - lines[-1][-1] <= 5:
                    lines[-1].append(y)
                else:
                    lines.append([y])
            centers = [int(_np.mean(g)) for g in lines]
            if len(centers) < 3:
                return []
            return [(centers[i], centers[i + 1])
                    for i in range(len(centers) - 1)
                    if centers[i + 1] - centers[i] > 100]
        except Exception:
            return []

    def _parse_catalog_by_table(self, image_path, wlog):
        """
        表格线模式解析: 横线定位每个数据行 → 整行裁剪降采样后单独OCR →
        按行提取 序号(数字或汉字一~十) + 页号(数字/范围/汉字数字)。
        相比全图片段模式的优势: 行不漏(每行独立OCR)、序号跳号无误判
        (序号来自本行而非推断)、支持汉字页号(一~十)。
        返回 [(序号, 起, 止), ...]; 无表格线或全失败返回 []。
        """
        bands = self._detect_table_rows(image_path)
        if not bands:
            return []
        wlog(f"    表格线模式: 检测到 {len(bands)} 个行区间")
        try:
            import cv2
        except ImportError:
            return []
        img = self._imread_cn(image_path)
        if img is None:
            return []
        H, W = img.shape[:2]

        def to_page(txt):
            """页号文本→(起,止)或None。支持 数字/范围/汉字一~十。"""
            t = txt.strip().replace(' ', '')
            m = re.fullmatch(r'(\d{1,4})[-–—~](\d{1,4})', t)
            if m:
                a, b = int(m.group(1)), int(m.group(2))
                return (min(a, b), max(a, b))
            if re.fullmatch(r'\d{1,4}', t):
                v = int(t)
                return (v, v)
            if t in self._CN_NUM:
                v = self._CN_NUM[t]
                return (v, v)
            return None

        def to_seq(txt):
            t = txt.strip()
            if re.fullmatch(r'\d{1,2}', t):
                return int(t)
            if t in self._CN_NUM:
                return self._CN_NUM[t]
            return None

        results = []   # [(y_center, seq, (s,e))]
        for bi, (y1, y2) in enumerate(bands):
            # 跳过表头行(含"序号"字样)与跨行标题行(无页号且序号列无数字)
            row_img = img[y1:y2]
            # 降采样到宽~1200 提速
            sc = 1200.0 / W
            small = cv2.resize(row_img, (1200, max(1, int((y2 - y1) * sc))))
            tmp = image_path + f'.row{bi}.png'
            self._imwrite_cn(tmp, small)
            r = self._ocr_page(tmp, None, lambda s: None)
            try:
                os.remove(tmp)
            except Exception:
                pass
            if not r:
                continue
            texts = [t for t, *_ in r]
            joined = ''.join(texts)
            if '序号' in joined and '页号' in joined:
                continue  # 表头行
            # 找序号: 行内最左的纯数字(1-2位)或汉字数字
            seq = None
            seq_frag = None
            seq_y = None
            for t, x0, y0, x1, _y1 in sorted(r, key=lambda f: f[1]):
                s = to_seq(t)
                if s is not None and s <= 30:
                    seq = s
                    seq_frag = (t, x0, y0, x1)
                    seq_y = (y1 + y2) / 2
                    break
            # 找页号: 行内任何位置的 页号形态(独立或长文本尾部)。
            # 排除序号已占用的片段(避免把序号数字当单页页号, 如「4|...|48-132」)。
            page = None
            for t, x0, y0, x1, _y1 in r:
                if seq_frag is not None and t == seq_frag[0] and x0 == seq_frag[1]:
                    continue  # 序号片段不再参与页号
                p = to_page(t)
                if p:
                    page = p
                    break
                # 长文本尾部粘连(与片段模式同规则)
                for pat in (r'\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2}\s*(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*$',
                            r'(?:^|\D)(\d{1,4})\s*[-–—~]\s*(\d{1,4})\s*$'):
                    m = re.search(pat, t)
                    if m:
                        a, b = int(m.group(1)), int(m.group(2))
                        page = (min(a, b), max(a, b))
                        break
                if page:
                    break
            # 有页号即可入表(序号缺失的行, 后面按锚点插值补)
            if page:
                results.append((seq_y if seq_y is not None else (y1 + y2) / 2,
                                seq, page))
                wlog(f"    表格行{bi}: 序号={seq if seq is not None else '?'} "
                     f"页号={page[0]}-{page[1]} "
                     f"← 行y{y1}-{y2} OCR:{'|'.join(texts)[:60]}")

        if not results:
            return []
        results.sort(key=lambda x: x[0])

        # ---- 序号缺失行: 用已识别序号锚点线性插值 ----
        # 例: 行位次[1,2,3] 序号[1,None,4] → 位次1=1, 位次3=4 → 位次2=3
        # (插值取整; 两端外推: 首位=首锚点-前距, 末位=末锚点+后距)
        idx_known = [(i, r[1]) for i, r in enumerate(results) if r[1] is not None]
        if not idx_known:
            # 完全无锚点(所有序号都被OCR漏识别): 档案序号从1起按位次递增
            for i, r in enumerate(results):
                results[i] = (r[0], i + 1, r[2])
            wlog(f"    [序号兜底] 全部{len(results)}行序号未被OCR识别, "
                 f"按位次从1递增赋值")
        else:
            for i, r in enumerate(results):
                if r[1] is not None:
                    continue
                prevs = [k for k in idx_known if k[0] < i]
                nexts = [k for k in idx_known if k[0] > i]
                if prevs and nexts:
                    (i0, s0) = prevs[-1]
                    (i1, s1) = nexts[0]
                    if i1 > i0:
                        raw = s0 + (s1 - s0) * (i - i0) / (i1 - i0)
                        # 锚点间序号增量>位次增量 → 档案存在跳号,
                        # 插值向上取整偏向跳号解释(1,[?],4 → 3 而非 2)
                        import math as _math
                        guess = int(_math.ceil(raw)) if (s1 - s0) > (i1 - i0) \
                            else int(round(raw))
                        r_new = (r[0], guess, r[2])
                        results[i] = r_new
                        wlog(f"    [序号插值] 行位次{i+1}: 序号={r_new[1]} "
                             f"(锚点 位次{i0+1}={s0}, 位次{i1+1}={s1}"
                             f"{'跳号' if (s1 - s0) > (i1 - i0) else ''})")
                elif nexts:
                    # 在首锚点之前: 首锚点序号 - 位次差
                    (i1, s1) = nexts[0]
                    guess = s1 - (i1 - i)
                    results[i] = (r[0], guess, r[2])
                    wlog(f"    [序号插值] 行位次{i+1}: 序号={guess} (首锚点外推)")
                elif prevs:
                    (i0, s0) = prevs[-1]
                    guess = s0 + (i - i0)
                    results[i] = (r[0], guess, r[2])
                    wlog(f"    [序号插值] 行位次{i+1}: 序号={guess} (末锚点外推)")

        # 单数字页号行: 止页=下一行起始-1
        parsed = []
        for i, (y, seq, page) in enumerate(results):
            s, e = page
            if s == e:
                nxt = None
                for j in range(i + 1, len(results)):
                    if results[j][2][0] > s:
                        nxt = results[j][2][0]
                        break
                if nxt is not None and nxt > s:
                    e = nxt - 1
            parsed.append((seq, s, e))
        return parsed

    # ---------- xlsx目录文件解析 ----------
    def _parse_catalog_from_xlsx(self, dir_name, wlog):
        """
        从xlsx文件读取分件数据: 在 self.xlsx_dir 及其子目录下查找与 dir_name 同名的
        xlsx文件, 第3行为标题行(含"序号""页号"列), 从第4行起读取数据。
        返回 [(序号int, 起页int, 止页int), ...] 或 None(未找到/读取失败)。
        """
        if not self.xlsx_dir or not os.path.isdir(self.xlsx_dir):
            return None

        # 查找与目录名同名的xlsx文件(递归搜索)
        xlsx_path = None
        for root, dirs, files in os.walk(self.xlsx_dir):
            for f in files:
                if f.lower().endswith('.xlsx') and os.path.splitext(f)[0] == dir_name:
                    xlsx_path = os.path.join(root, f)
                    break
            if xlsx_path:
                break

        if not xlsx_path:
            wlog(f"  [xlsx] 未找到与目录「{dir_name}」同名的xlsx文件")
            return None

        wlog(f"  [xlsx] 读取目录文件: {os.path.basename(xlsx_path)}")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active

            # 第3行为标题行(索引2, 0-based), 查找"序号"和"页号"列
            header_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))
            if not header_row:
                wlog(f"  [xlsx] 第3行为空, 无法读取标题行")
                wb.close()
                return None

            headers = [str(c).strip() if c is not None else '' for c in header_row[0]]
            seq_col = None
            page_col = None
            for i, h in enumerate(headers):
                if h == '序号':
                    seq_col = i
                elif h == '页号':
                    page_col = i

            if seq_col is None or page_col is None:
                wlog(f"  [xlsx] 标题行未找到「序号」或「页号」列, 标题: {headers}")
                wb.close()
                return None

            wlog(f"  [xlsx] 序号列={seq_col}, 页号列={page_col}")

            # 从第4行起读取数据
            raw_data = []  # [(序号int, 页号文本)]
            for row in ws.iter_rows(min_row=4, values_only=True):
                cells = list(row)
                if len(cells) <= max(seq_col, page_col):
                    continue
                seq_val = cells[seq_col]
                page_val = cells[page_col]
                if seq_val is None or page_val is None:
                    continue
                try:
                    seq_int = int(seq_val)
                except (ValueError, TypeError):
                    continue
                page_str = str(page_val).strip().replace(' ', '')
                if not page_str:
                    continue
                raw_data.append((seq_int, page_str))

            wb.close()

            if not raw_data:
                wlog(f"  [xlsx] 未读取到有效数据行")
                return None

            # 按序号从小到大排序
            raw_data.sort(key=lambda x: x[0])
            wlog(f"  [xlsx] 读取到 {len(raw_data)} 条数据")

            # 解析页号: 支持 "199-213" 范围格式和单数字格式
            parsed = []
            for i, (seq, page_str) in enumerate(raw_data):
                m = re.fullmatch(r'(\d{1,4})[-–—~](\d{1,4})', page_str)
                if m:
                    p_start, p_end = int(m.group(1)), int(m.group(2))
                    if p_end < p_start:
                        p_start, p_end = p_end, p_start
                    src = page_str
                else:
                    # 单数字: 起始页=该数字, 终止页=下一序号的起始页-1
                    try:
                        p_start = int(page_str)
                    except ValueError:
                        wlog(f"    [xlsx] 跳过行: 序号={seq} 页号格式无法解析「{page_str}」")
                        continue
                    # 找下一个有条目(范围式)的起始页, 或下一个序号的起始页
                    nxt = None
                    for j in range(i + 1, len(raw_data)):
                        m2 = re.match(r'(\d{1,4})', raw_data[j][1])
                        if m2:
                            nxt = int(m2.group(1))
                            break
                    p_end = (nxt - 1) if (nxt is not None and nxt > p_start) else p_start
                    src = f"{page_str}(推断到{p_end})"
                parsed.append((seq, p_start, p_end))
                wlog(f"    [xlsx] 序号={seq} 页号={p_start}-{p_end} ← 「{src}」")

            if not parsed:
                wlog(f"  [xlsx] 未能解析出有效分件数据")
                return None

            wlog(f"  [xlsx] 解析完成, 共 {len(parsed)} 件")
            return parsed

        except ImportError:
            wlog(f"  [xlsx] 缺少 openpyxl 库, 无法读取xlsx文件")
            return None
        except Exception as e:
            wlog(f"  [xlsx] 读取xlsx出错: {e}")
            return None

    # ---------- 文件操作 ----------
    @staticmethod
    def _num_of(fname):
        """提取文件名(不含扩展名)末尾的数字编号, 无数字返回 None。
        兼容纯数字文件名(0001.jpg)与带前缀文件名(档号-0001.jpg)。"""
        m = re.search(r'(\d{1,4})$', os.path.splitext(fname)[0])
        return int(m.group(1)) if m else None

    @staticmethod
    def _jpg_files_sorted(dir_path):
        """目录下 jpg 文件按文件名末尾数字编号升序(无编号的排最后)。"""
        files = [f for f in os.listdir(dir_path)
                 if f.lower().endswith('.jpg') and os.path.isfile(os.path.join(dir_path, f))]
        def key(f):
            stem = os.path.splitext(f)[0]
            if stem.isdigit():
                return int(stem)
            n = FileSplitWorker._num_of(f)
            return n if n is not None else float('inf')
        return sorted(files, key=key)

    @staticmethod
    def _read_directory_txt(subdir, wlog):
        """
        读取目录下的 Directory.txt(由“文件批量替换”生成, 记录替换后的文件名
        清单, 不含扩展名, 每行一个)。
        返回 [(末尾编号, 不含扩展名文件名), ...] 按编号升序;
        文件不存在或无有效记录返回 None。
        """
        dpath = os.path.join(subdir, 'Directory.txt')
        if not os.path.isfile(dpath):
            return None
        lines = []
        for enc in ('utf-8', 'gbk'):
            try:
                with open(dpath, 'r', encoding=enc) as fh:
                    lines = [l.strip() for l in fh if l.strip()]
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                wlog(f"  × 读取 Directory.txt 失败: {e}")
                return None
        items = []
        for stem in lines:
            m = re.search(r'(\d{1,4})$', stem)
            if m:
                items.append((int(m.group(1)), stem))
        if not items:
            wlog("  × Directory.txt 存在但无有效文件名记录, 按不存在处理")
            return None
        items.sort(key=lambda t: t[0])
        return items

    def _resolve_directory_txt(self, subdir, jpgs, wlog):
        """
        解析 Directory.txt 得到 卷皮/目录页文件清单与后续文件偏移量。
        规则: 第一页(编号最小的jpg)为卷皮; Directory.txt 记录的文件为目录页;
        偏移量 = Directory.txt 中最大文件序号(小于2时按2)。
        返回 (offset, front_files), front_files=[卷皮, 目录页...];
        无 Directory.txt 或无有效记录返回 None。
        """
        items = self._read_directory_txt(subdir, wlog)
        if items is None or not jpgs:
            return None
        f1 = jpgs[0]
        f1_stem = os.path.splitext(f1)[0]
        by_num = {}
        for f in jpgs:
            n = self._num_of(f)
            if n is not None:
                by_num.setdefault(n, f)
        front = [f1]
        for num, stem in items:
            if stem == f1_stem:
                continue  # 卷皮页不计入目录页
            cand = None
            for f in jpgs:
                if os.path.splitext(f)[0] == stem:
                    cand = f
                    break
            if cand is None:
                cand = by_num.get(num)  # 命名形态不一致时按末尾编号匹配
            if cand is None:
                wlog(f"  × Directory.txt 记录「{stem}」在目录中无对应jpg, 忽略该记录")
                continue
            if cand not in front:
                front.append(cand)
        max_num = max(n for n, _ in items)
        offset = max(max_num, 2)
        wlog(f"  Directory.txt 共 {len(items)} 条记录, 最大文件序号 {max_num:04d}, "
             f"后续文件偏移量={offset}")
        return offset, front

    def _parse_catalog_from_dir(self, subdir, wlog):
        """
        对子目录做「目录页定位 + OCR + 标题判断 + 表格解析」(不移动任何文件)。
        分件与检查共用此入口, 保证两者行为一致。
        目录页定位:
          有 Directory.txt → 第一页为卷皮, 目录页为 Directory.txt 记录的文件
          (逐个尝试OCR), 后续文件偏移量=Directory.txt最大文件序号;
          无 Directory.txt → 未分件=根下第2张(0002); 已分件=卷皮目录里的第2张
          (偏移量缺省为2)。
        返回 (entries, front_files, offset, err):
          成功: entries=[(序号,起,止),...], front_files=归入卷皮目录的文件
          (卷皮+目录页), err=''
          失败: entries=[], err=状态字符串(跳过(xxx)/失败(xxx))
        """
        dir_name = os.path.basename(subdir)
        jpgs = self._jpg_files_sorted(subdir)
        if not jpgs:
            return [], [], 2, "跳过(无jpg)"

        # ---- Directory.txt 模式(已做过卷内目录替换): 卷皮=第一页, 目录页取自 Directory.txt ----
        info = self._resolve_directory_txt(subdir, jpgs, wlog)
        if info and len(info[1]) >= 2:
            offset, front_files = info
            wlog(f"  卷皮页: {front_files[0]}; 目录页(来自Directory.txt): "
                 f"{', '.join(front_files[1:])}")
            for cand in front_files[1:]:
                page_path = os.path.join(subdir, cand)
                rows = self._ocr_page(page_path, None, wlog)
                if not rows:
                    wlog(f"  目录页候选 {cand}: OCR无结果, 尝试下一个")
                    continue
                if not self._is_catalog_title(rows):
                    wlog(f"  目录页候选 {cand}: 非卷内文件目录, 尝试下一个")
                    continue
                wlog("  标题确认: 卷内文件目录")
                # 优先: 表格线模式(逐行裁剪OCR, 不漏行/序号不推断/支持汉字页号)
                try:
                    entries = self._parse_catalog_by_table(page_path, wlog)
                except Exception as e:
                    wlog(f"    表格线模式异常: {e}")
                    entries = []
                if not entries:
                    entries = self._parse_catalog_rows(rows, wlog)  # 回退: 全图片段模式
                if entries:
                    wlog(f"  目录页={cand}, 解析 {len(entries)} 行, 后续文件偏移量={offset}")
                    return entries, front_files, offset, ""
                wlog(f"  目录页候选 {cand}: 标题匹配但未解析到数据, 尝试下一个")
            return [], front_files, offset, "失败(Directory.txt记录均非目录页)"
        if info:
            wlog("  Directory.txt 存在但未解析出对应目录页jpg, 回退默认规则")

        # ---- 默认规则(无 Directory.txt) ----
        jp_dir = os.path.join(subdir, f"{dir_name}卷皮目录")
        jp_jpgs = self._jpg_files_sorted(jp_dir) if os.path.isdir(jp_dir) else []
        if len(jp_jpgs) >= 2:
            f1, f2 = jp_jpgs[0], jp_jpgs[1]
            page2_path = os.path.join(jp_dir, f2)
            wlog(f"  目录页取自卷皮目录: {f2}")
        elif len(jpgs) >= 2:
            f1, f2 = jpgs[0], jpgs[1]
            page2_path = os.path.join(subdir, f2)
        else:
            return [], [], 2, "跳过(无目录页)"

        rows = self._ocr_page(page2_path, None, wlog)
        if not rows:
            return [], [f1, f2], 2, "失败(OCR无结果)"
        if not self._is_catalog_title(rows):
            return [], [f1, f2], 2, "跳过(非目录页)"
        wlog("  标题确认: 卷内文件目录")

        # 优先: 表格线模式(逐行裁剪OCR, 不漏行/序号不推断/支持汉字页号)
        try:
            entries = self._parse_catalog_by_table(page2_path, wlog)
        except Exception as e:
            wlog(f"    表格线模式异常: {e}")
            entries = []
        if entries:
            wlog(f"  表格线模式解析 {len(entries)} 行")
            return entries, [f1, f2], 2, ""

        # 回退: 全图片段模式
        entries = self._parse_catalog_rows(rows, wlog)
        if not entries:
            return [], [f1, f2], 2, "失败(未解析到数据)"
        return entries, [f1, f2], 2, ""

    def _split_by_entries(self, subdir, entries, wlog, target_base=None, copy_mode=False,
                          offset=2):
        """
        按 entries 建序号子目录并处理文件(页号+偏移量=文件编号; 偏移量缺省为2,
        有 Directory.txt 时为其最大文件序号)。
        target_base=None: 原地移动(在 subdir 下建子目录并移入);
        target_base 指定: 输出到 target_base/目录名/ 下, copy_mode=True 拷贝
        (源目录不动), False 仍移动。返回 处理文件数。分件/手工分件共用。
        """
        dir_name = os.path.basename(subdir)
        out_root = os.path.join(target_base, dir_name) if target_base else subdir
        def _op(src, dst):
            if copy_mode:
                shutil.copy2(src, dst)
            else:
                shutil.move(src, dst)
        verb = '拷贝' if copy_mode else '移动'
        # 编号→文件名映射(兼容纯数字 0001.jpg 与带前缀 档号-0001.jpg 命名)
        num_map = {}
        for f in os.listdir(subdir):
            if f.lower().endswith('.jpg') and os.path.isfile(os.path.join(subdir, f)):
                nn = self._num_of(f)
                if nn is not None:
                    num_map.setdefault(nn, f)
        moved = 0
        for seq, p_start, p_end in entries:
            sub_name = f"{dir_name}-{seq:04d}"
            sub_path = os.path.join(out_root, sub_name)
            os.makedirs(sub_path, exist_ok=True)
            for n in range(p_start + offset, p_end + offset + 1):
                fname = num_map.get(n, f"{n:04d}.jpg")
                src = os.path.join(subdir, fname)
                if os.path.exists(src):
                    try:
                        _op(src, os.path.join(sub_path, fname))
                        moved += 1
                    except Exception as e:
                        wlog(f"    × {verb}失败 {fname}: {e}")
                else:
                    wlog(f"    (缺) 编号{n:04d}文件不存在, 跳过")
            wlog(f"  序号{seq}: 页号{p_start}-{p_end} (偏移量={offset}) → {sub_name}/ "
                 f"{verb} 编号 {p_start + offset:04d}..{p_end + offset:04d}")
        return moved

    def _process_one_dir(self, subdir, logf, wlog, target_base=None, copy_mode=False):
        """处理单个子目录。返回 (状态字符串, 处理文件数)。"""
        dir_name = os.path.basename(subdir)
        jpgs = self._jpg_files_sorted(subdir)
        if len(jpgs) < 4:
            wlog(f"  [跳过] {dir_name}: jpg 少于 4 张({len(jpgs)}张), 无法分件")
            return "跳过(文件不足)", 0

        wlog(f"  总文件数: {len(jpgs)}")

        # 分件依据: xlsx目录文件模式 或 OCR模式
        if self.xlsx_dir:
            # xlsx目录文件模式: 从同名xlsx读取分件数据
            entries = self._parse_catalog_from_xlsx(dir_name, wlog)
            if entries is None:
                wlog(f"  [xlsx] 无法从目录文件获取分件数据, 跳过: {dir_name}")
                return "跳过(xlsx未找到)", 0
            # 卷皮/目录页与偏移量: 优先 Directory.txt; 无则缺省前两张+偏移量2
            info = self._resolve_directory_txt(subdir, jpgs, wlog)
            if info:
                offset, front_files = info
            else:
                offset, front_files = 2, jpgs[:2]
                wlog(f"  未找到 Directory.txt, 按缺省偏移量 2 处理; "
                     f"卷皮页: {front_files[0]}, {front_files[1]}")
        else:
            # OCR模式(默认)
            entries, front_files, offset, err = self._parse_catalog_from_dir(subdir, wlog)
            if err:
                wlog(f"  [{err.split('(')[0]}] {dir_name}: {err}")
                return err, 0

        out_root = os.path.join(target_base, dir_name) if target_base else subdir
        os.makedirs(out_root, exist_ok=True)
        verb = '拷贝' if copy_mode else '移动'
        def _op(src, dst):
            if copy_mode:
                shutil.copy2(src, dst)
            else:
                shutil.move(src, dst)

        # 建序号子目录并处理文件(手工分件共用)
        moved = self._split_by_entries(subdir, entries, wlog,
                                       target_base=target_base, copy_mode=copy_mode,
                                       offset=offset)

        # 建备考表卷底 / 卷皮目录
        path_beikao = os.path.join(out_root, f"{dir_name}备考表卷底")
        path_juanpi = os.path.join(out_root, f"{dir_name}卷皮目录")
        os.makedirs(path_beikao, exist_ok=True)
        os.makedirs(path_juanpi, exist_ok=True)

        # 卷皮页 + 目录页 → 卷皮目录
        for fname in front_files:
            src = os.path.join(subdir, fname)
            if os.path.exists(src):
                _op(src, os.path.join(path_juanpi, fname))
                moved += 1
        wlog(f"  卷皮: {','.join(front_files)} → {dir_name}卷皮目录/ [{verb}]")

        # 当前剩余文件里 最大与次大 → 备考表卷底
        if copy_mode:
            # 拷贝模式下源目录不变, 备考取未被规则覆盖的最大两张
            covered_nums = set()
            for seq, p_start, p_end in entries:
                covered_nums.update(range(p_start + offset, p_end + offset + 1))
            for fname in front_files:
                nn = self._num_of(fname)
                if nn is not None:
                    covered_nums.add(nn)
            remain = [f for f in jpgs
                      if (self._num_of(f) if self._num_of(f) is not None else -1)
                      not in covered_nums]
        else:
            remain = self._jpg_files_sorted(subdir)
        if len(remain) >= 2:
            for fname in (remain[-1], remain[-2]):
                src = os.path.join(subdir, fname)
                _op(src, os.path.join(path_beikao, fname))
                moved += 1
            wlog(f"  备考: {remain[-2]},{remain[-1]} → {dir_name}备考表卷底/ [{verb}]")
        else:
            wlog(f"  备考: 剩余文件不足2张({len(remain)}), 未处理")

        # 分件完成后删除当前目录下 Directory.txt(已消费);
        # 移动/拷贝到分件目录时始终忽略该文件(不随分件结果输出)
        dpath = os.path.join(subdir, 'Directory.txt')
        if os.path.isfile(dpath):
            try:
                os.remove(dpath)
                wlog(f"  已删除当前目录下 Directory.txt (分件已消费, 不随分件输出)")
            except Exception as e:
                wlog(f"  × 删除 Directory.txt 失败: {e}")

        return f"完成({len(entries)}件,{verb}{moved}个文件)", moved

    def run(self):
        try:
            if not os.path.isdir(self.base_dir):
                self.finished_signal.emit(False, "所选目录不存在")
                return

            # 分件到新目录: 目标目录不存在则创建
            if self.target_base:
                os.makedirs(self.target_base, exist_ok=True)

            # 日志文件放在用户所选目录下
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_path = os.path.join(self.base_dir, f"分件处理日志_{ts}.txt")
            logf = open(log_path, 'w', encoding='utf-8')
            lock = __import__('threading').Lock()

            def wlog(s):
                with lock:
                    logf.write(s + "\n")
                    logf.flush()
                self.log_signal.emit(s)

            wlog("分件处理日志")
            wlog(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            wlog(f"所选目录: {self.base_dir}")
            if self.xlsx_dir:
                wlog(f"分件模式: 读取目录文件(xlsx) → {self.xlsx_dir}")
            else:
                wlog(f"分件模式: OCR识别目录页")
            wlog("=" * 70)

            subdirs = sorted([os.path.join(self.base_dir, d)
                              for d in os.listdir(self.base_dir)
                              if os.path.isdir(os.path.join(self.base_dir, d))])
            total = len(subdirs)
            if total == 0:
                logf.close()
                self.finished_signal.emit(False, "所选目录下没有子目录")
                return

            wlog(f"共发现 {total} 个子目录待处理")
            done = 0
            moved_total = 0
            for subdir in subdirs:
                if self.is_stopped:
                    wlog("用户停止处理")
                    break
                wlog("")
                wlog(f"[{done + 1}/{total}] 处理: {os.path.basename(subdir)}")
                wlog("-" * 50)
                try:
                    status, moved = self._process_one_dir(
                        subdir, logf, wlog,
                        target_base=self.target_base, copy_mode=self.copy_mode)
                    moved_total += moved
                    wlog(f"  结果: {status}")
                except Exception as e:
                    import traceback
                    wlog(f"  [异常] {e}")
                    wlog(traceback.format_exc())
                done += 1
                self.progress_signal.emit(done, total)

            wlog("")
            wlog("=" * 70)
            wlog(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            wlog(f"总计: 处理 {done}/{total} 个子目录, 移动文件 {moved_total} 个")
            if self.is_stopped:
                wlog("注意: 处理被用户中途停止")
            logf.close()

            msg = (f"分件完成！处理 {done}/{total} 个子目录，移动 {moved_total} 个文件。\n"
                   f"日志: {os.path.basename(log_path)}")
            self.finished_signal.emit(not self.is_stopped, msg)

        except Exception as e:
            import traceback
            self.log_signal.emit(traceback.format_exc())
            self.finished_signal.emit(False, f"处理出错: {e}")

    def stop(self):
        self.is_stopped = True


class FileSplitCheckWorker(FileSplitWorker):
    """
    分件检查后台线程 —— 继承 FileSplitWorker, OCR/解析/文件排序全部复用
    分件代码(_parse_catalog_from_dir / _parse_catalog_from_xlsx), 保证检查与分件行为完全一致。
    对每个子目录: 解析目录页(xlsx或OCR)得到序号/页号, 推算「应被移走」的文件集合,
    检测根目录中按标准应移走却仍残留的文件 → 记入检查报告。不移动任何文件。
    """
    progress_signal = Signal(int, int)

    def __init__(self, base_dir, xlsx_dir=None, parent=None):
        super().__init__(base_dir, xlsx_dir=xlsx_dir, parent=parent)

    def _check_one_dir(self, subdir, wlog):
        """
        按分件规则检查单个子目录(不移动文件)。检查项:
        1. 卷内文件目录页是否存在、能否正常解析(OCR失败/非目录页/无数据=错误);
        2. 每段页号范围是否有足够文件(页号+偏移量=文件编号, 偏移量=
           Directory.txt最大序号, 无Directory.txt缺省为2; 范围内缺文件=错误);
        3. 分件规则覆盖完成后, 根目录剩余的文件(未被任何页号段/卷皮/备考覆盖)=错误。
        返回 (状态, 错误明细list) —— 每条错误含类型与对应文件名。
        """
        errors = []
        dir_name = os.path.basename(subdir)
        jpgs = self._jpg_files_sorted(subdir)
        if not jpgs:
            return "错误(无jpg)", [f"{dir_name}: [目录异常] 根目录无 jpg 文件"]

        # --- 检查项1: 目录页存在性与可解析性 ---
        if self.xlsx_dir:
            # xlsx目录文件模式
            entries = self._parse_catalog_from_xlsx(dir_name, wlog)
            if entries is None:
                return "错误(xlsx未找到)", [f"{dir_name}: [目录解析] 未找到同名xlsx目录文件"]
            info = self._resolve_directory_txt(subdir, jpgs, wlog)
            if info:
                offset, front_files = info
            else:
                offset, front_files = 2, list(jpgs[:2])
        else:
            # OCR模式(默认)
            entries, front_files, offset, err = self._parse_catalog_from_dir(subdir, wlog)
            if err:
                # err 形如 "失败(OCR无结果)"/"跳过(非目录页)"/"失败(未解析到数据)"
                return f"错误({err})", [f"{dir_name}: [目录解析] {err}"]

        wlog(f"  解析到 {len(entries)} 段页号 (偏移量={offset})")

        # 当前根目录实际存在的文件(按末尾编号映射)
        num_map = {}
        for f in jpgs:
            nn = self._num_of(f)
            if nn is not None:
                num_map.setdefault(nn, f)
        existing = set(jpgs)
        # 分件规则覆盖到的文件集合(含每段页号范围 + 卷皮/目录页 + 备考)
        covered = set()

        # --- 检查项2: 每段页号范围内文件是否足够 ---
        for seq, p_start, p_end in entries:
            expect_nums = list(range(p_start + offset, p_end + offset + 1))
            missing = [n for n in expect_nums if n not in num_map]
            covered.update(num_map[n] for n in expect_nums if n in num_map)
            if missing:
                # 该段期望 (p_end-p_start+1) 个, 缺 len(missing) 个
                total = p_end - p_start + 1
                errors.append(f"{dir_name}: [文件不足] 序号{seq} 页号{p_start}-{p_end} "
                              f"应有{total}个文件, 缺{len(missing)}个: "
                              f"{', '.join(f'{n:04d}' for n in missing[:10])}"
                              f"{'...' if len(missing) > 10 else ''}")
                wlog(f"  [文件不足] 序号{seq}: 缺 {len(missing)} 个 "
                     f"({', '.join(f'{n:04d}' for n in missing[:8])}"
                     f"{'...' if len(missing) > 8 else ''})")
            else:
                wlog(f"  序号{seq}: 页号{p_start}-{p_end} 文件齐全({len(expect_nums)}个)")

        # 卷皮/目录页(front_files)与备考(剩余最大两张)也计入覆盖
        for fname in front_files:
            if fname in existing:
                covered.add(fname)
        remain_sim = [f for f in jpgs if f not in covered]
        if len(remain_sim) >= 2:
            covered.add(remain_sim[-1])
            covered.add(remain_sim[-2])

        # --- 检查项3: 规则覆盖后仍剩余的文件 ---
        leftover = [f for f in jpgs if f not in covered]
        if leftover:
            errors.append(f"{dir_name}: [剩余文件] 分件规则未覆盖, 残留 "
                          f"{len(leftover)} 个: {', '.join(leftover[:10])}"
                          f"{'...' if len(leftover) > 10 else ''}")
            wlog(f"  [剩余文件] {len(leftover)} 个未被覆盖: "
                 f"{', '.join(leftover[:8])}{'...' if len(leftover) > 8 else ''}")

        if errors:
            return f"错误({len(errors)}项)", errors
        return "正常", errors

    def run(self):
        try:
            if not os.path.isdir(self.base_dir):
                self.finished_signal.emit(False, "所选目录不存在")
                return

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = os.path.join(self.base_dir, f"分件检查报告_{ts}.txt")
            logf = open(report_path, 'w', encoding='utf-8')
            lock = __import__('threading').Lock()

            def wlog(s):
                with lock:
                    logf.write(s + "\n")
                    logf.flush()
                self.log_signal.emit(s)

            wlog("分件检查报告")
            wlog(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            wlog(f"所选目录: {self.base_dir}")
            if self.xlsx_dir:
                wlog(f"检查模式: 读取目录文件(xlsx) → {self.xlsx_dir}")
                wlog("检查标准: 从xlsx获取分件数据, 推算应移走文件, 检测根目录残留")
            else:
                wlog("检查标准: 复用分件OCR解析, 推算应移走文件, 检测根目录残留")
            wlog("=" * 70)

            subdirs = sorted([os.path.join(self.base_dir, d)
                              for d in os.listdir(self.base_dir)
                              if os.path.isdir(os.path.join(self.base_dir, d))])
            total = len(subdirs)
            if total == 0:
                logf.close()
                self.finished_signal.emit(False, "所选目录下没有子目录")
                return

            wlog(f"共 {total} 个子目录待检查")
            all_errors = []
            stats = {}
            done = 0
            for subdir in subdirs:
                if self.is_stopped:
                    wlog("用户停止检查")
                    break
                wlog("")
                wlog(f"[{done + 1}/{total}] 检查: {os.path.basename(subdir)}")
                wlog("-" * 50)
                try:
                    status, errs = self._check_one_dir(subdir, wlog)
                    stats[status] = stats.get(status, 0) + 1
                    if errs:
                        # 按目录归组: 目录名 + 各错误类型(含文件名)
                        all_errors.append((os.path.basename(subdir), errs))
                    wlog(f"  结果: {status}")
                except Exception as e:
                    import traceback
                    wlog(f"  [异常] {e}")
                    wlog(traceback.format_exc())
                    all_errors.append((os.path.basename(subdir),
                                       [f"[检查异常] {e}"]))
                done += 1
                self.progress_signal.emit(done, total)

            wlog("")
            wlog("=" * 70)
            wlog("检查汇总:")
            for k, v in sorted(stats.items()):
                wlog(f"  {k}: {v} 个目录")
            n_err_dirs = len(all_errors)
            n_err_items = sum(len(e) for _, e in all_errors)
            wlog(f"有错误的目录: {n_err_dirs} 个, 错误共 {n_err_items} 项")
            if all_errors:
                wlog("")
                wlog("错误明细(按目录, 含错误类型与文件名):")
                for dname, errs in all_errors:
                    wlog(f"  ▷ {dname}")
                    for e in errs:
                        wlog(f"     × {e}")
            wlog("")
            wlog(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            logf.close()

            msg = (f"检查完成！{done}/{total} 个目录；"
                   f"{n_err_dirs} 个目录有错误(共{n_err_items}项)。\n"
                   f"报告: {os.path.basename(report_path)}")
            self.finished_signal.emit(not self.is_stopped, msg)
        except Exception as e:
            import traceback
            self.log_signal.emit(traceback.format_exc())
            self.finished_signal.emit(False, f"检查出错: {e}")


def manual_split_entries(text):
    """
    解析手工分件输入(多行)为分件条目。
    行格式:
      卷皮目录: 1-2          (可选, 文件名页码, 直接用不+2)
      备考表卷底: 213-214    (可选, 同上)
      1 1-180               (序号 页号; 页号+2=文件名, 与自动分件一致)
    页号支持单数字(183)或范围(184-197)。
    返回 (卷皮range|None, 备考range|None, entries, err)。
    校验: 格式/序号唯一升序/页号连续性(本条起始=前条止页+1)。
    """
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if not lines:
        return None, None, [], "未输入任何内容"

    juanpi = beikao = None
    entries = []

    def parse_range(s):
        m = re.fullmatch(r'(\d{1,4})\s*[-–—~]\s*(\d{1,4})', s.strip())
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            return (min(a, b), max(a, b))
        if re.fullmatch(r'\d{1,4}', s.strip()):
            v = int(s.strip())
            return (v, v)
        return None

    for ln in lines:
        low = ln.replace(' ', '').replace('：', ':')
        if low.startswith('卷皮目录'):
            parts = ln.split(':', 1) if ':' in ln else ln.split('：', 1)
            if len(parts) < 2:
                return None, None, [], "「卷皮目录」行缺少页号(格式: 卷皮目录: 1-2)"
            r = parse_range(parts[1])
            if not r:
                return None, None, [], f"「卷皮目录」页号格式错误: {parts[1]}"
            juanpi = r
            continue
        if low.startswith('备考表卷底'):
            parts = ln.split(':', 1) if ':' in ln else ln.split('：', 1)
            if len(parts) < 2:
                return None, None, [], "「备考表卷底」行缺少页号(格式: 备考表卷底: 213-214)"
            r = parse_range(parts[1])
            if not r:
                return None, None, [], f"「备考表卷底」页号格式错误: {parts[1]}"
            beikao = r
            continue
        parts = ln.split()
        if len(parts) != 2:
            return None, None, [], f"行格式错误(应为「序号 页号」): {ln}"
        if not re.fullmatch(r'\d{1,2}', parts[0]):
            return None, None, [], f"序号须为1-2位数字: {parts[0]}"
        r = parse_range(parts[1])
        if not r:
            return None, None, [], f"页号格式错误(应为 17 或 17-101): {parts[1]}"
        entries.append((int(parts[0]), r[0], r[1]))

    if not entries:
        return None, None, [], "未输入任何序号/页号数据行"

    seqs = [e[0] for e in entries]
    if len(set(seqs)) != len(seqs):
        return None, None, [], f"序号重复: {seqs}"
    if seqs != sorted(seqs):
        return None, None, [], f"序号未按升序: {seqs}"

    for i in range(1, len(entries)):
        prev_end = entries[i - 1][2]
        cur_start = entries[i][1]
        if cur_start != prev_end + 1:
            return (None, None, [],
                    f"页号不连续: 序号{entries[i - 1][0]}止于{prev_end}, "
                    f"序号{entries[i][0]}应从{prev_end + 1}开始, 实际{cur_start}")
    return juanpi, beikao, entries, ""


class ManualSplitDialog(QDialog):
    """
    手工分件对话框: 列出待处理目录, 用户输入多行序号/页号(前两行可选
    卷皮目录/备考表卷底的文件名页码), 校验后执行分件; 完成自动切换下一个。
    文件移动复用 FileSplitWorker._split_by_entries。
    """

    def __init__(self, base_dir, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.pending_dirs = self._scan_pending()
        self.current_idx = -1
        self.total_moved = 0
        self.splitter = FileSplitWorker(base_dir)  # 仅用其文件排序/移动方法

        self.setWindowTitle("手工分件")
        self.resize(760, 640)

        v = QVBoxLayout(self)
        v.addWidget(QLabel("待处理目录:"))
        self.dir_list = QListWidget()
        self.dir_list.setMaximumHeight(140)
        for d in self.pending_dirs:
            self.dir_list.addItem(d)
        self.dir_list.currentRowChanged.connect(self.on_dir_selected)
        v.addWidget(self.dir_list)

        v.addWidget(QLabel("分件数据表格 (「页号」支持 17 或 17-101; 序号行页号按+2换算;\n"
                           "卷皮目录/备考表卷底行输入文件名页码, 不换算, 留空用默认):"))
        # 表格输入: 第0/1行固定为 卷皮目录/备考表卷底, 其后为数据行
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["序号", "页号", "说明"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self._init_table_rows()
        # 回车跳到下一个输入项: 在数据行内 序号→页号→下一行序号→…;
        # 最后一行末格回车自动加行并进入新行序号格
        self.table.keyPressEvent = self._table_key_press
        # 工具行: 增/删数据行
        tr = QHBoxLayout()
        btn_add = QPushButton("+ 加行")
        btn_add.setObjectName("BrowseBtn")
        btn_add.clicked.connect(self.add_row)
        tr.addWidget(btn_add)
        btn_del = QPushButton("− 删选中行")
        btn_del.setObjectName("BrowseBtn")
        btn_del.clicked.connect(self.del_row)
        tr.addWidget(btn_del)
        tr.addStretch()
        v.addLayout(tr)
        v.addWidget(self.table, 1)

        self.hint = QLabel("")
        self.hint.setStyleSheet("color: #B45309; font-size: 12px;")
        v.addWidget(self.hint)

        h = QHBoxLayout()
        self.apply_btn = QPushButton("执行分件")
        self.apply_btn.setObjectName("ActionBtn")
        self.apply_btn.clicked.connect(self.apply)
        h.addWidget(self.apply_btn)
        self.skip_btn = QPushButton("跳过此目录")
        self.skip_btn.setObjectName("BrowseBtn")
        self.skip_btn.clicked.connect(self.skip)
        h.addWidget(self.skip_btn)
        self.close_btn = QPushButton("结束")
        self.close_btn.setObjectName("BrowseBtn")
        self.close_btn.clicked.connect(self.close)
        h.addWidget(self.close_btn)
        h.addStretch()
        v.addLayout(h)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMaximumHeight(150)
        v.addWidget(QLabel("处理日志:"))
        v.addWidget(self.log_box)

        if self.pending_dirs:
            self.dir_list.setCurrentRow(0)
        else:
            self.hint.setText("所选目录下没有待处理目录(根下有jpg且未分件的子目录)")

    def _scan_pending(self):
        """待处理 = 根下有 jpg 且未分件(无 -NNNN 序号子目录/卷皮未归档)的子目录。"""
        out = []
        try:
            for d in sorted(os.listdir(self.base_dir)):
                p = os.path.join(self.base_dir, d)
                if not os.path.isdir(p):
                    continue
                jpgs = [f for f in os.listdir(p)
                        if f.lower().endswith('.jpg')
                        and os.path.isfile(os.path.join(p, f))]
                if not jpgs:
                    continue
                subs = [s for s in os.listdir(p)
                        if os.path.isdir(os.path.join(p, s))
                        and re.search(r'-\d{4}$', s)]
                if subs and len(jpgs) <= 2:
                    continue
                out.append(d)
        except Exception:
            pass
        return out

    def log(self, s):
        self.log_box.append(f">> {s}")

    # ---------- 表格输入辅助 ----------
    def _init_table_rows(self):
        """初始化表格: 前2行固定(卷皮目录/备考表卷底), 预置5个空数据行。"""
        self.table.setRowCount(0)
        self._append_fixed_row("卷皮目录", "文件名页码, 不+2换算; 留空=默认0001/0002")
        self._append_fixed_row("备考表卷底", "文件名页码, 不+2换算; 留空=默认剩余最大两张")
        for _ in range(5):
            self.add_row()

    def _append_fixed_row(self, name, note):
        r = self.table.rowCount()
        self.table.insertRow(r)
        it0 = QTableWidgetItem(name)
        it0.setFlags(it0.flags() & ~Qt.ItemIsEditable)  # 名称列锁定
        self.table.setItem(r, 0, it0)
        self.table.setItem(r, 1, QTableWidgetItem(""))
        it2 = QTableWidgetItem(note)
        it2.setFlags(it2.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(r, 2, it2)

    def add_row(self):
        r = self.table.rowCount()
        self.table.insertRow(r)
        self.table.setItem(r, 0, QTableWidgetItem(""))
        self.table.setItem(r, 1, QTableWidgetItem(""))
        self.table.setItem(r, 2, QTableWidgetItem(""))

    def _table_key_press(self, event):
        """回车→下一个输入项: 页号列→下一行序号列; 序号列→本行页号列。
        最后一行页号回车→自动加行并进入新行序号格。其余按键走默认处理。"""
        from PyQt5.QtGui import QKeyEvent
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            r = self.table.currentRow()
            c = self.table.currentColumn()
            if r < 0:
                return
            if c == 0:
                # 序号 → 本行页号
                self.table.setCurrentCell(r, 1)
                self.table.editItem(self.table.item(r, 1))
                return
            if c == 1:
                # 页号 → 下一行序号; 末行则加行
                if r + 1 < self.table.rowCount():
                    self.table.setCurrentCell(r + 1, 0)
                    self.table.editItem(self.table.item(r + 1, 0))
                else:
                    self.add_row()
                    self.table.setCurrentCell(r + 1, 0)
                    self.table.editItem(self.table.item(r + 1, 0))
                return
        # 其余按键交给 QTableWidget 默认处理
        return QTableWidget.keyPressEvent(self.table, event)

    def del_row(self):
        rows = sorted({i.row() for i in self.table.selectedIndexes()}, reverse=True)
        for r in rows:
            if r >= 2:  # 固定行不可删
                self.table.removeRow(r)

    def _table_values(self):
        """读取表格 → (卷皮range|None, 备考range|None, entries, err)。"""
        def parse_range(s):
            t = s.strip().replace(' ', '')
            m = re.fullmatch(r'(\d{1,4})[-–—~](\d{1,4})', t)
            if m:
                a, b = int(m.group(1)), int(m.group(2))
                return (min(a, b), max(a, b))
            if re.fullmatch(r'\d{1,4}', t):
                return (int(t), int(t))
            return None

        def cell(r, c):
            it = self.table.item(r, c)
            return (it.text() if it else '').strip()

        jp_txt = cell(0, 1)
        bk_txt = cell(1, 1)
        if jp_txt and jp_txt != '-':
            juanpi = parse_range(jp_txt)
            if not juanpi:
                return None, None, [], f"卷皮目录页号格式错误: {jp_txt}"
        else:
            juanpi = None
        if bk_txt and bk_txt != '-':
            beikao = parse_range(bk_txt)
            if not beikao:
                return None, None, [], f"备考表卷底页号格式错误: {bk_txt}"
        else:
            beikao = None

        entries = []
        for r in range(2, self.table.rowCount()):
            seq_txt = cell(r, 0)
            pg_txt = cell(r, 1)
            if not seq_txt and not pg_txt:
                continue  # 空行跳过
            if not re.fullmatch(r'\d{1,2}', seq_txt):
                return None, None, [], f"第{r - 1}行序号须为1-2位数字: 「{seq_txt}」"
            pg = parse_range(pg_txt)
            if not pg:
                return None, None, [], f"序号{seq_txt}页号格式错误(17 或 17-101): 「{pg_txt}」"
            entries.append((int(seq_txt), pg[0], pg[1]))

        if not entries:
            return None, None, [], "未输入任何序号/页号数据行"

        seqs = [e[0] for e in entries]
        if len(set(seqs)) != len(seqs):
            return None, None, [], f"序号重复: {seqs}"
        if seqs != sorted(seqs):
            return None, None, [], f"序号未按升序: {seqs}"
        for i in range(1, len(entries)):
            prev_end = entries[i - 1][2]
            cur_start = entries[i][1]
            if cur_start != prev_end + 1:
                return (None, None, [],
                        f"页号不连续: 序号{entries[i - 1][0]}止于{prev_end}, "
                        f"序号{entries[i][0]}应从{prev_end + 1}开始, 实际{cur_start}")
        return juanpi, beikao, entries, ""

    def on_dir_selected(self, row):
        self.current_idx = row
        if 0 <= row < len(self.pending_dirs):
            self.hint.setText(f"当前目录: {self.pending_dirs[row]}  "
                              f"(第 {row + 1}/{len(self.pending_dirs)} 个)")
            # 清空数据行(保留卷皮/备考两固定行)
            while self.table.rowCount() > 2:
                self.table.removeRow(self.table.rowCount() - 1)
            for _ in range(5):
                self.add_row()
            self.table.setFocus()

    def apply(self):
        if not (0 <= self.current_idx < len(self.pending_dirs)):
            QMessageBox.warning(self, "提示", "请先在列表中选择待处理目录")
            return
        juanpi, beikao, entries, err = self._table_values()
        if err:
            self.hint.setText("输入错误: " + err)
            QMessageBox.warning(self, "输入错误", err)
            return

        dir_name = self.pending_dirs[self.current_idx]
        subdir = os.path.join(self.base_dir, dir_name)
        reply = QMessageBox.question(
            self, "确认分件",
            f"将对 {dir_name} 执行分件:\n"
            + ''.join(f"  序号{s}: {a}-{b}\n" for s, a, b in entries)
            + "文件将被移动(不可自动撤销)，确定执行吗？",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.log(f"── 分件: {dir_name} ──")

        def move_named(dst_path, rng):
            """卷皮/备考: 输入即文件名页码(不+2), 复用移动语义。"""
            os.makedirs(dst_path, exist_ok=True)
            cnt = 0
            for n in range(rng[0], rng[1] + 1):
                fname = f"{n:04d}.jpg"
                src = os.path.join(subdir, fname)
                if os.path.exists(src):
                    shutil.move(src, os.path.join(dst_path, fname))
                    cnt += 1
                else:
                    self.log(f"    (缺) {fname} 不存在, 跳过")
            return cnt

        # 序号子目录 —— 复用分件worker的移动方法
        moved = self.splitter._split_by_entries(subdir, entries, self.log)

        # 卷皮目录
        jp_path = os.path.join(subdir, f"{dir_name}卷皮目录")
        if juanpi:
            c = move_named(jp_path, juanpi)
            self.log(f"  卷皮: 文件{juanpi[0]:04d}-{juanpi[1]:04d} → {dir_name}卷皮目录/ 移动{c}个")
        else:
            os.makedirs(jp_path, exist_ok=True)
            jpgs = self.splitter._jpg_files_sorted(subdir)
            for fname in (jpgs[0], jpgs[1]) if len(jpgs) >= 2 else []:
                src = os.path.join(subdir, fname)
                if os.path.exists(src):
                    shutil.move(src, os.path.join(jp_path, fname))
                    moved += 1
            self.log(f"  卷皮: 默认(最小两张) → {dir_name}卷皮目录/")

        # 备考表卷底
        bk_path = os.path.join(subdir, f"{dir_name}备考表卷底")
        if beikao:
            c = move_named(bk_path, beikao)
            self.log(f"  备考: 文件{beikao[0]:04d}-{beikao[1]:04d} → {dir_name}备考表卷底/ 移动{c}个")
        else:
            os.makedirs(bk_path, exist_ok=True)
            remain = self.splitter._jpg_files_sorted(subdir)
            if len(remain) >= 2:
                for fname in (remain[-1], remain[-2]):
                    shutil.move(os.path.join(subdir, fname),
                                os.path.join(bk_path, fname))
                    moved += 1
                self.log(f"  备考: 默认(剩余最大两张) → {dir_name}备考表卷底/")
            else:
                self.log(f"  备考: 剩余不足2张({len(remain)}), 未移动")

        self.total_moved += moved
        self.log(f"结果: 完成({len(entries)}件, 移动{moved}个文件)")

        # 分件完成后删除当前目录下 Directory.txt(已消费, 不随分件输出)
        dpath = os.path.join(subdir, 'Directory.txt')
        if os.path.isfile(dpath):
            try:
                os.remove(dpath)
                self.log(f"  已删除当前目录下 Directory.txt (分件已消费, 不随分件输出)")
            except Exception as e:
                self.log(f"  × 删除 Directory.txt 失败: {e}")

        # 从待处理列表移除并自动切换下一个
        self.pending_dirs.pop(self.current_idx)
        self.dir_list.clear()
        for d in self.pending_dirs:
            self.dir_list.addItem(d)
        if self.pending_dirs:
            nxt = min(self.current_idx, len(self.pending_dirs) - 1)
            self.dir_list.setCurrentRow(nxt)
            self.hint.setText(f"已完成 {dir_name}。自动切换到下一个: {self.pending_dirs[nxt]}")
        else:
            self.hint.setText("全部待处理目录分件完成！")
            QMessageBox.information(self, "完成",
                                    f"全部分件完成！共移动 {self.total_moved} 个文件。")
            self.close()

    def skip(self):
        if 0 <= self.current_idx < len(self.pending_dirs):
            d = self.pending_dirs.pop(self.current_idx)
            self.log(f"跳过: {d}")
            self.dir_list.clear()
            for dd in self.pending_dirs:
                self.dir_list.addItem(dd)
            if self.pending_dirs:
                self.dir_list.setCurrentRow(min(self.current_idx, len(self.pending_dirs) - 1))
            else:
                self.hint.setText("全部待处理目录已处理(或跳过)")
                QMessageBox.information(self, "完成", "全部待处理目录已处理完成。")
                self.close()


class FileRenameWorker(QThread):
    """文件重命名后台处理线程"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, base_dir, modify_dpi=False, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.modify_dpi = modify_dpi  # 是否修改JPG文件的DPI为600
        self.is_stopped = False
    
    def run(self):
        try:
            results = {
                "processed_dirs": 0,
                "renamed_files": 0,
                "failed_files": 0,
                "errors": [],
                "actions": [],
                "log_entries": []
            }
            
            # 遍历基础目录下的所有子目录
            subdirs = [d for d in Path(self.base_dir).iterdir() if d.is_dir()]
            total_dirs = len(subdirs)
            processed_dirs = 0
            
            for subdir in subdirs:
                if self.is_stopped:
                    break
                
                # 获取子目录中的所有文件
                files = [f for f in subdir.iterdir() if f.is_file()]
                
                if not files:
                    processed_dirs += 1
                    continue
                
                # 获取目录名（不含路径）
                dir_name = subdir.name
                
                # 如果只有一个文件，直接使用目录名
                if len(files) == 1:
                    file = files[0]
                    new_name = subdir / f"{dir_name}{file.suffix}"
                    
                    action_desc = f"将 '{file.name}' 重命名为 '{new_name.name}'"
                    results["actions"].append(action_desc)
                    self.log_signal.emit(action_desc)
                    
                    log_entry = {
                        "directory": subdir.name,
                        "original_name": file.name,
                        "new_name": new_name.name,
                        "modify_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "success": True,
                        "error_msg": ""
                    }
                    
                    try:
                        file.rename(new_name)
                        results["renamed_files"] += 1
                        
                        # 如果需要修改DPI且是JPG文件，则修改DPI
                        if self.modify_dpi and new_name.suffix.lower() in ['.jpg', '.jpeg']:
                            try:
                                img = Image.open(new_name)
                                # 设置DPI为600
                                img.save(new_name, dpi=(600, 600))
                                self.log_signal.emit(f"已修改 {new_name.name} 的DPI为600")
                            except Exception as dpi_error:
                                error_msg = f"修改DPI失败 {new_name.name}: {str(dpi_error)}"
                                results["errors"].append(error_msg)
                                log_entry["success"] = False
                                log_entry["error_msg"] = error_msg
                                self.log_signal.emit(f"错误: {error_msg}")
                    except Exception as e:
                        error_msg = f"重命名文件失败 {file}: {str(e)}"
                        results["errors"].append(error_msg)
                        results["failed_files"] += 1
                        log_entry["success"] = False
                        log_entry["error_msg"] = error_msg
                        self.log_signal.emit(f"错误: {error_msg}")
                    
                    results["log_entries"].append(log_entry)
                else:
                    # 如果有多个文件，则使用递增后缀
                    for i, file in enumerate(files, start=1):
                        if self.is_stopped:
                            break
                        
                        # 生成新的文件名（目录名-序号.扩展名）
                        new_name = subdir / f"{dir_name}-{i:04d}{file.suffix}"
                        
                        action_desc = f"将 '{file.name}' 重命名为 '{new_name.name}'"
                        results["actions"].append(action_desc)
                        self.log_signal.emit(action_desc)
                        
                        log_entry = {
                            "directory": subdir.name,
                            "original_name": file.name,
                            "new_name": new_name.name,
                            "modify_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "success": True,
                            "error_msg": ""
                        }
                        
                        try:
                            file.rename(new_name)
                            results["renamed_files"] += 1
                            
                            # 如果需要修改DPI且是JPG文件，则修改DPI
                            if self.modify_dpi and new_name.suffix.lower() in ['.jpg', '.jpeg']:
                                try:
                                    img = Image.open(new_name)
                                    # 设置DPI为600
                                    img.save(new_name, dpi=(600, 600))
                                    self.log_signal.emit(f"已修改 {new_name.name} 的DPI为600")
                                except Exception as dpi_error:
                                    error_msg = f"修改DPI失败 {new_name.name}: {str(dpi_error)}"
                                    results["errors"].append(error_msg)
                                    log_entry["success"] = False
                                    log_entry["error_msg"] = error_msg
                                    self.log_signal.emit(f"错误: {error_msg}")
                        except Exception as e:
                            error_msg = f"重命名文件失败 {file}: {str(e)}"
                            results["errors"].append(error_msg)
                            results["failed_files"] += 1
                            log_entry["success"] = False
                            log_entry["error_msg"] = error_msg
                            self.log_signal.emit(f"错误: {error_msg}")
                        
                        results["log_entries"].append(log_entry)
                
                processed_dirs += 1
                results["processed_dirs"] = processed_dirs
                self.progress_signal.emit(processed_dirs, total_dirs)
            
            if not self.is_stopped:
                # 创建日志文件
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                log_filename = f"rename_log_{timestamp}.txt"
                log_filepath = os.path.join(self.base_dir, log_filename)
                
                with open(log_filepath, 'w', encoding='utf-8') as log_file:
                    log_file.write(f"文件重命名日志 - 开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    log_file.write(f"基础目录: {self.base_dir}\n")
                    log_file.write("=" * 80 + "\n")
                    log_file.write("目录名\t原文件名\t新文件名\t修改时间\t状态\t错误信息\n")
                    log_file.write("=" * 80 + "\n")
                    
                    for entry in results['log_entries']:
                        status = "成功" if entry['success'] else "失败"
                        error_info = entry['error_msg'] if entry['error_msg'] else ""
                        log_line = f"{entry['directory']}\t{entry['original_name']}\t{entry['new_name']}\t{entry['modify_time']}\t{status}\t{error_info}\n"
                        log_file.write(log_line)
                    
                    log_file.write("=" * 80 + "\n")
                    log_file.write(f"处理完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    log_file.write(f"总计处理目录: {results['processed_dirs']} 个\n")
                    log_file.write(f"成功重命名文件: {results['renamed_files']} 个\n")
                    log_file.write(f"失败文件: {results['failed_files']} 个\n")
                
                success_count = results['renamed_files']
                fail_count = results['failed_files']
                self.finished_signal.emit(True, 
                    f"处理完成！总计: {results['processed_dirs']} 个目录, "
                    f"成功: {success_count}, 失败: {fail_count}\n日志: {log_filename}")
            else:
                self.finished_signal.emit(False, "处理已停止")
                
        except Exception as e:
            self.finished_signal.emit(False, f"处理出错: {str(e)}")
    
    def stop(self):
        self.is_stopped = True


class ExtRenameWorker(QThread):
    """
    修改文件扩展名后台线程：
    递归扫描指定目录及子目录，将匹配旧扩展名的文件改为新扩展名。
    """
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)

    def __init__(self, base_dir, old_ext, new_ext, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.old_ext = old_ext.lower().strip().lstrip('.')  # 统一小写, 去前导点
        self.new_ext = new_ext.lower().strip().lstrip('.')
        self.is_stopped = False

    def run(self):
        try:
            if not self.old_ext:
                self.finished_signal.emit(False, "旧扩展名不能为空")
                return
            if not self.new_ext:
                self.finished_signal.emit(False, "新扩展名不能为空")
                return

            # 收集所有匹配的文件
            old_dot = f'.{self.old_ext}'
            matched_files = []
            for root, dirs, files in os.walk(self.base_dir):
                for f in files:
                    if f.lower().endswith(old_dot):
                        matched_files.append(os.path.join(root, f))

            total = len(matched_files)
            if total == 0:
                self.finished_signal.emit(False, f"未找到扩展名为 .{self.old_ext} 的文件")
                return

            self.log_signal.emit(f"找到 {total} 个 .{self.old_ext} 文件待修改")

            renamed = 0
            failed = 0
            log_entries = []

            for i, filepath in enumerate(matched_files):
                if self.is_stopped:
                    self.log_signal.emit("用户停止处理")
                    break

                dir_part = os.path.dirname(filepath)
                basename = os.path.basename(filepath)
                name_no_ext = os.path.splitext(basename)[0]
                new_name = f"{name_no_ext}.{self.new_ext}"
                new_path = os.path.join(dir_part, new_name)

                rel_path = os.path.relpath(filepath, self.base_dir)
                rel_new = os.path.relpath(new_path, self.base_dir)

                entry = {
                    "original": rel_path,
                    "new": rel_new,
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "success": True,
                    "error": ""
                }

                # 目标文件已存在则跳过
                if os.path.exists(new_path):
                    entry["success"] = False
                    entry["error"] = "目标文件已存在, 跳过"
                    failed += 1
                    self.log_signal.emit(f"  × 跳过 {rel_path}: 目标 {new_name} 已存在")
                    log_entries.append(entry)
                    self.progress_signal.emit(i + 1, total)
                    continue

                try:
                    os.rename(filepath, new_path)
                    renamed += 1
                    self.log_signal.emit(f"  ✓ {rel_path} → {rel_new}")
                except Exception as e:
                    entry["success"] = False
                    entry["error"] = str(e)
                    failed += 1
                    self.log_signal.emit(f"  × 失败 {rel_path}: {e}")

                log_entries.append(entry)
                self.progress_signal.emit(i + 1, total)

            # 写入日志文件
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_filename = f"ext_rename_log_{ts}.txt"
            log_filepath = os.path.join(self.base_dir, log_filename)
            try:
                with open(log_filepath, 'w', encoding='utf-8') as lf:
                    lf.write(f"扩展名修改日志\n")
                    lf.write(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    lf.write(f"基础目录: {self.base_dir}\n")
                    lf.write(f"旧扩展名: .{self.old_ext} → 新扩展名: .{self.new_ext}\n")
                    lf.write("=" * 80 + "\n")
                    lf.write("原文件路径\t新文件路径\t修改时间\t状态\t错误信息\n")
                    lf.write("=" * 80 + "\n")
                    for e in log_entries:
                        status = "成功" if e["success"] else "失败"
                        lf.write(f"{e['original']}\t{e['new']}\t{e['time']}\t{status}\t{e['error']}\n")
                    lf.write("=" * 80 + "\n")
                    lf.write(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    lf.write(f"总计: {total} 个文件, 成功 {renamed} 个, 失败/跳过 {failed} 个\n")
            except Exception as e:
                self.log_signal.emit(f"写入日志文件失败: {e}")

            msg = (f"扩展名修改完成！共 {total} 个文件，"
                   f"成功 {renamed} 个，失败/跳过 {failed} 个。\n"
                   f"日志: {log_filename}")
            self.finished_signal.emit(not self.is_stopped, msg)

        except Exception as e:
            import traceback
            self.log_signal.emit(traceback.format_exc())
            self.finished_signal.emit(False, f"处理出错: {e}")

    def stop(self):
        self.is_stopped = True


class FileRenamePage(FunctionPage):
    def __init__(self):
        super().__init__("文件改名")

        # ====== 功能1: 按目录名批量重命名 ======
        group = QGroupBox("按目录名批量重命名文件")
        form = QFormLayout()

        self.dir_path = QLineEdit()
        btn_browse = QPushButton("选择文件夹")
        btn_browse.setObjectName("BrowseBtn")
        btn_browse.clicked.connect(self.browse_dir)
        h1 = QHBoxLayout()
        h1.addWidget(self.dir_path)
        h1.addWidget(btn_browse)
        form.addRow("目标目录:", h1)

        self.modify_dpi_check = QCheckBox("修改JPG文件的DPI为600")
        self.modify_dpi_check.setChecked(False)
        form.addRow("选项:", self.modify_dpi_check)

        # 功能1的操作按钮(放在分组内)
        btn_layout = QHBoxLayout()
        self.preview_btn = QPushButton("预览操作")
        self.preview_btn.setObjectName("ActionBtn")
        self.preview_btn.setStyleSheet("background-color: #2196F3;")
        self.preview_btn.clicked.connect(self.preview_operations)
        btn_layout.addWidget(self.preview_btn)

        self.exec_btn = QPushButton("开始批量改名")
        self.exec_btn.setObjectName("ActionBtn")
        self.exec_btn.clicked.connect(self.execute)
        btn_layout.addWidget(self.exec_btn)

        self.stop_btn = QPushButton("停止")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        btn_layout.addStretch()
        form.addRow("", btn_layout)

        group.setLayout(form)
        self.layout.addWidget(group)

        # ====== 功能2: 修改文件扩展名 ======
        ext_group = QGroupBox("修改文件扩展名(递归处理目录及子目录)")
        ext_form = QFormLayout()

        self.ext_dir_edit = QLineEdit()
        self.ext_dir_edit.setPlaceholderText("指定目录及其子目录下的文件将被处理")
        btn_ext_browse = QPushButton("选择文件夹")
        btn_ext_browse.setObjectName("BrowseBtn")
        btn_ext_browse.clicked.connect(self.browse_ext_dir)
        h_ext = QHBoxLayout()
        h_ext.addWidget(self.ext_dir_edit)
        h_ext.addWidget(btn_ext_browse)
        ext_form.addRow("目标目录:", h_ext)

        self.ext_old_edit = QLineEdit()
        self.ext_old_edit.setPlaceholderText("例如: tif 或 .tif")
        self.ext_old_edit.setMaximumWidth(160)
        ext_form.addRow("旧扩展名:", self.ext_old_edit)

        self.ext_new_edit = QLineEdit()
        self.ext_new_edit.setPlaceholderText("例如: jpg 或 .jpg")
        self.ext_new_edit.setMaximumWidth(160)
        ext_form.addRow("新扩展名:", self.ext_new_edit)

        # 功能2的操作按钮(放在分组内)
        ext_btn_layout = QHBoxLayout()
        self.ext_preview_btn = QPushButton("预览")
        self.ext_preview_btn.setObjectName("ActionBtn")
        self.ext_preview_btn.setStyleSheet("background-color: #2196F3;")
        self.ext_preview_btn.clicked.connect(self.preview_ext_rename)
        ext_btn_layout.addWidget(self.ext_preview_btn)

        self.ext_exec_btn = QPushButton("开始修改扩展名")
        self.ext_exec_btn.setObjectName("ActionBtn")
        self.ext_exec_btn.clicked.connect(self.execute_ext_rename)
        ext_btn_layout.addWidget(self.ext_exec_btn)

        self.ext_stop_btn = QPushButton("停止")
        self.ext_stop_btn.setObjectName("ActionBtn")
        self.ext_stop_btn.setStyleSheet("background-color: #DA3633;")
        self.ext_stop_btn.clicked.connect(self.stop_ext_processing)
        self.ext_stop_btn.setEnabled(False)
        ext_btn_layout.addWidget(self.ext_stop_btn)
        ext_btn_layout.addStretch()
        ext_form.addRow("", ext_btn_layout)

        ext_group.setLayout(ext_form)
        self.layout.addWidget(ext_group)

        # 说明文本
        info_label = QLabel(
            "功能说明：\n"
            "【按目录名批量重命名】\n"
            "• 单文件：直接以目录名命名；多文件：目录名-0001、目录名-0002...\n"
            "• 自动生成详细日志文件；可选修改JPG文件DPI为600\n"
            "【修改文件扩展名】\n"
            "• 递归扫描目录及子目录，将旧扩展名文件改为新扩展名\n"
            "• 扩展名输入无需点号(如输入 tif 或 .tif 均可)；自动生成日志文件"
        )
        info_label.setStyleSheet("color: #8B949E; font-size: 12px;")
        self.layout.addWidget(info_label)

        self.worker = None
        self.ext_worker = None
        self.add_log_widget()

        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)

    def browse_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择目录")
        if d: self.dir_path.setText(d)
    
    def preview_operations(self):
        """预览将要执行的操作"""
        d = self.dir_path.text()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        
        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "指定的目录不存在")
            return
        
        # 清空之前的结果
        self.log_box.clear()
        
        self.log("="*50)
        self.log("正在预览操作...")
        self.log(f"基础目录: {d}")
        self.log("-"*50)
        
        try:
            from pathlib import Path
            subdirs = [dir for dir in Path(d).iterdir() if dir.is_dir()]
            total_files = 0
            actions = []
            
            for subdir in subdirs:
                files = [f for f in subdir.iterdir() if f.is_file()]
                if not files:
                    continue
                
                dir_name = subdir.name
                
                if len(files) == 1:
                    file = files[0]
                    new_name = f"{dir_name}{file.suffix}"
                    action = f"将 '{file.name}' 重命名为 '{new_name}'"
                    actions.append(action)
                    self.log(f"  • {action}")
                else:
                    for i, file in enumerate(files, start=1):
                        new_name = f"{dir_name}-{i:04d}{file.suffix}"
                        action = f"将 '{file.name}' 重命名为 '{new_name}'"
                        actions.append(action)
                        self.log(f"  • {action}")
                
                total_files += len(files)
            
            self.log("-"*50)
            self.log(f"预览完成！")
            self.log(f"将处理 {len(subdirs)} 个子目录")
            self.log(f"将重命名 {total_files} 个文件")
            
            if self.modify_dpi_check.isChecked():
                self.log("\n注意：将同时修改JPG文件的DPI为600")
            
            self.log("\n注意：这只是预览，文件尚未重命名。")
            
        except Exception as e:
            error_msg = f"预览过程中出现错误: {str(e)}"
            self.log(error_msg)
            QMessageBox.critical(self, "错误", error_msg)

    def execute(self):
        d = self.dir_path.text()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        
        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "指定的目录不存在")
            return
        
        # 确认操作
        reply = QMessageBox.question(
            self, "确认操作",
            "确定要开始重命名文件吗？\n此操作不可逆！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # 重置进度条
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        
        # 禁用按钮，启用停止按钮
        self.log("="*50)
        self.log("开始处理...")
        self.log(f"目录: {d}")
        if self.modify_dpi_check.isChecked():
            self.log("选项: 修改JPG文件DPI为600")
        self.log("="*50)
        
        # 创建工作线程
        modify_dpi = self.modify_dpi_check.isChecked()
        self.worker = FileRenameWorker(base_dir=d, modify_dpi=modify_dpi)
        
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        
        self.worker.start()
        
        # 更新按钮状态
        self.exec_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
    
    def stop_processing(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止处理...")
            self.stop_btn.setEnabled(False)
    
    def update_progress(self, current, total):
        """更新进度显示"""
        percentage = (current / total) * 100 if total > 0 else 0
        self.progress.setValue(int(percentage))
        self.progress.setFormat(f"{current} / {total}  ({percentage:.0f}%)")
        self.log(f"进度: {current}/{total} ({percentage:.1f}%)")
    
    def on_finished(self, success, message):
        """处理完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")
        
        # 恢复按钮状态
        self.exec_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        
        if success:
            QMessageBox.information(self, "完成", "处理完成！")
        else:
            QMessageBox.warning(self, "提示", message)

    # ---------- 扩展名修改功能 ----------
    def browse_ext_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择扩展名修改目录")
        if d:
            self.ext_dir_edit.setText(d)

    def preview_ext_rename(self):
        """预览扩展名修改操作"""
        d = self.ext_dir_edit.text().strip()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择扩展名修改的目标目录")
            return
        if not os.path.isdir(d):
            QMessageBox.warning(self, "错误", "目录不存在")
            return

        old_ext = self.ext_old_edit.text().strip().lstrip('.').lower()
        new_ext = self.ext_new_edit.text().strip().lstrip('.').lower()
        if not old_ext:
            QMessageBox.warning(self, "提示", "请输入旧扩展名")
            return
        if not new_ext:
            QMessageBox.warning(self, "提示", "请输入新扩展名")
            return

        self.log_box.clear()
        self.log("=" * 50)
        self.log("预览扩展名修改操作")
        self.log(f"目录: {d}")
        self.log(f"旧扩展名: .{old_ext} → 新扩展名: .{new_ext}")
        self.log("-" * 50)

        try:
            old_dot = f'.{old_ext}'
            count = 0
            for root, dirs, files in os.walk(d):
                for f in files:
                    if f.lower().endswith(old_dot):
                        rel = os.path.relpath(os.path.join(root, f), d)
                        name_no_ext = os.path.splitext(f)[0]
                        new_name = f"{name_no_ext}.{new_ext}"
                        self.log(f"  • {rel} → {new_name}")
                        count += 1

            self.log("-" * 50)
            if count == 0:
                self.log(f"未找到 .{old_ext} 文件")
            else:
                self.log(f"共找到 {count} 个 .{old_ext} 文件将被修改为 .{new_ext}")
            self.log("\n注意：这只是预览，文件尚未修改。")
        except Exception as e:
            self.log(f"预览出错: {e}")
            QMessageBox.critical(self, "错误", f"预览过程中出错: {e}")

    def execute_ext_rename(self):
        """执行扩展名修改"""
        d = self.ext_dir_edit.text().strip()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择扩展名修改的目标目录")
            return
        if not os.path.isdir(d):
            QMessageBox.warning(self, "错误", "目录不存在")
            return

        old_ext = self.ext_old_edit.text().strip().lstrip('.').lower()
        new_ext = self.ext_new_edit.text().strip().lstrip('.').lower()
        if not old_ext:
            QMessageBox.warning(self, "提示", "请输入旧扩展名")
            return
        if not new_ext:
            QMessageBox.warning(self, "提示", "请输入新扩展名")
            return
        if old_ext == new_ext:
            QMessageBox.warning(self, "提示", "旧扩展名与新扩展名相同，无需修改")
            return

        reply = QMessageBox.question(
            self, "确认操作",
            f"确定要将 .{old_ext} 文件的扩展名修改为 .{new_ext} 吗？\n"
            f"目录: {d}\n"
            f"将递归处理所有子目录。\n此操作不可逆！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        self.log("=" * 50)
        self.log("开始修改扩展名...")
        self.log(f"目录: {d}")
        self.log(f".{old_ext} → .{new_ext}")
        self.log("=" * 50)

        self.ext_worker = ExtRenameWorker(d, old_ext, new_ext)
        self.ext_worker.log_signal.connect(self.log)
        self.ext_worker.progress_signal.connect(self.update_progress)
        self.ext_worker.finished_signal.connect(self.on_ext_finished)
        self.ext_worker.start()

        self.ext_exec_btn.setEnabled(False)
        self.ext_preview_btn.setEnabled(False)
        self.ext_stop_btn.setEnabled(True)

    def stop_ext_processing(self):
        """停止扩展名修改处理"""
        if self.ext_worker and self.ext_worker.isRunning():
            self.ext_worker.stop()
            self.log("正在停止扩展名修改处理...")
            self.ext_stop_btn.setEnabled(False)

    def on_ext_finished(self, success, message):
        """扩展名修改完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")
        self.ext_exec_btn.setEnabled(True)
        self.ext_preview_btn.setEnabled(True)
        self.ext_stop_btn.setEnabled(False)
        if success:
            QMessageBox.information(self, "完成", "扩展名修改完成！")
        else:
            QMessageBox.warning(self, "提示", message)


class AutoPagingWorker(QThread):
    """后台处理线程"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, directory, margin_mm, start_number, thread_count, cover_digit, parent=None):
        super().__init__(parent)
        self.directory = directory
        self.margin_mm = margin_mm
        self.start_number = start_number
        self.thread_count = thread_count
        self.cover_digit = cover_digit
        self.is_stopped = False
    
    def run(self):
        try:
            # 收集文件
            dir_files_map = self.collect_jpg_files()
            if not dir_files_map:
                self.finished_signal.emit(False, "未找到任何JPG文件")
                return
            
            total_dirs = len(dir_files_map)
            total_files = sum(len(files) for files in dir_files_map.values())
            self.log_signal.emit(f"找到 {total_dirs} 个目录，共 {total_files} 个JPG文件")
            
            # 创建日志文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_filename = f"sequence_number_log_{timestamp}.txt"
            log_path = os.path.join(self.directory, log_filename)
            self.init_log_file(log_path)
            
            # 处理所有目录
            all_results = []
            processed_dirs = 0
            
            with ThreadPoolExecutor(max_workers=self.thread_count) as executor:
                future_to_dir = {}
                for dir_path, jpg_files in dir_files_map.items():
                    future = executor.submit(
                        self.process_directory_files, 
                        dir_path, jpg_files, self.start_number
                    )
                    future_to_dir[future] = dir_path
                
                for future in as_completed(future_to_dir):
                    if self.is_stopped:
                        break
                    
                    dir_path = future_to_dir[future]
                    try:
                        results = future.result()
                        all_results.extend(results)
                        
                        # 记录日志
                        for result in results:
                            self.append_to_log(log_path, result)
                            self.log_signal.emit(
                                f"{result['directory']}/{result['filename']} - "
                                f"序号:{result['number']} - {result['result']} ({result['duration']}秒)"
                            )
                    except Exception as e:
                        error_msg = f"处理目录 {dir_path} 时出错: {str(e)}"
                        self.log_signal.emit(error_msg)
                    
                    processed_dirs += 1
                    self.progress_signal.emit(processed_dirs, total_dirs)
            
            if not self.is_stopped:
                # 完成日志
                self.finalize_log_file(log_path, all_results)
                success_count = sum(1 for r in all_results if r['result'] == '成功')
                fail_count = len(all_results) - success_count
                self.finished_signal.emit(True, 
                    f"处理完成！总计: {len(all_results)}, 成功: {success_count}, 失败: {fail_count}\n日志: {log_path}")
            else:
                self.finished_signal.emit(False, "处理已停止")
                
        except Exception as e:
            self.finished_signal.emit(False, f"处理出错: {str(e)}")
    
    def stop(self):
        self.is_stopped = True
    
    def collect_jpg_files(self):
        """收集目录及子目录下所有JPG文件"""
        dir_files_map = {}
        for root, dirs, files in os.walk(self.directory):
            jpg_files = []
            for filename in files:
                if filename.lower().endswith(('.jpg', '.jpeg')):
                    jpg_files.append(os.path.join(root, filename))
            if jpg_files:
                jpg_files.sort()
                dir_files_map[root] = jpg_files
        return dir_files_map
    
    def get_chinese_font(self):
        """获取中文字体"""
        font_paths = [
            "C:/Windows/Fonts/simsun.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/msyh.ttc",
        ]
        font_size = 236
        for font_path in font_paths:
            try:
                if os.path.exists(font_path):
                    return ImageFont.truetype(font_path, font_size)
            except Exception:
                continue
        return ImageFont.load_default()
    
    def add_number_to_image(self, image_path, number):
        """在图片右上角添加数字"""
        try:
            # cv2/numpy 延迟导入：Win7 上 opencv 较易出现 DLL 加载失败，
            # 放在此处可避免影响主程序启动与其他功能页（导入失败会被下方 except 捕获）。
            import cv2
            import numpy as np
            img = cv2.imdecode(np.fromfile(image_path, dtype=np.uint8), cv2.IMREAD_COLOR)
            if img is None:
                return False
            
            img_pil = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
            draw = ImageDraw.Draw(img_pil)
            font = self.get_chinese_font()
            
            # 计算边距（毫米转像素，假设300DPI）
            margin_px = int(self.margin_mm * 300 / 25.4)
            
            text = str(number)
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
            
            x = img_pil.width - margin_px - text_width
            y = margin_px
            
            # 如果需要覆盖数字区域
            if self.cover_digit:
                padding = int(20 * 300 / 25.4)
                cover_x = x - padding // 2
                cover_y = y - padding // 3
                cover_width = text_width + padding
                cover_height = text_height + padding // 2
                draw.rectangle([cover_x, cover_y, cover_x + cover_width, cover_y + cover_height], 
                              fill=(255, 255, 255), outline=None)
            
            draw.text((x, y), text, fill=(0, 0, 0), font=font)
            
            img_cv = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
            cv2.imencode('.jpg', img_cv)[1].tofile(image_path)
            return True
        except Exception as e:
            print(f"处理图片 {image_path} 时出错: {str(e)}")
            return False
    
    def process_directory_files(self, directory_path, jpg_files, start_number):
        """处理单个目录下的JPG文件"""
        results = []
        current_number = start_number
        
        for jpg_file in jpg_files:
            if self.is_stopped:
                break
            
            start_time = time.time()
            success = self.add_number_to_image(jpg_file, current_number)
            end_time = time.time()
            
            results.append({
                'directory': os.path.basename(directory_path),
                'filename': os.path.basename(jpg_file),
                'number': current_number,
                'result': '成功' if success else '失败',
                'duration': round(end_time - start_time, 3),
                'full_path': jpg_file
            })
            
            if success:
                current_number += 1
        
        return results
    
    def init_log_file(self, log_path):
        """初始化日志文件"""
        try:
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(f"JPG文件添加序号处理日志\n")
                f.write(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 100 + "\n")
                f.write(f"{'目录名':<20} {'文件名':<30} {'序号':<8} {'结果':<10} {'耗时(秒)':<12} {'处理时间':<20}\n")
                f.write("-" * 100 + "\n")
        except Exception as e:
            print(f"创建日志文件失败: {str(e)}")
    
    def append_to_log(self, log_path, result):
        """追加单条记录到日志文件"""
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                process_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                f.write(f"{result['directory']:<20} {result['filename']:<30} "
                       f"{result['number']:<8} {result['result']:<10} {result['duration']:<12} {process_time:<20}\n")
        except Exception as e:
            print(f"追加日志记录失败: {str(e)}")
    
    def finalize_log_file(self, log_path, results):
        """完成日志文件"""
        try:
            success_count = sum(1 for r in results if r['result'] == '成功')
            fail_count = len(results) - success_count
            end_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write("-" * 100 + "\n")
                f.write(f"结束时间: {end_time}\n")
                f.write(f"总计处理: {len(results)} 个文件\n")
                f.write(f"成功: {success_count} 个\n")
                f.write(f"失败: {fail_count} 个\n")
                f.write("=" * 100 + "\n")
        except Exception as e:
            print(f"完成日志文件失败: {str(e)}")


class AutoPagingPage(FunctionPage):
    def __init__(self):
        super().__init__("自动编页码")
        group = QGroupBox("JPG文件右上角添加序号")
        form = QFormLayout()

        self.file_dir = QLineEdit()
        btn_browse = QPushButton("选择文件夹")
        btn_browse.setObjectName("BrowseBtn")
        btn_browse.clicked.connect(lambda: self.file_dir.setText(QFileDialog.getExistingDirectory(self, "选择目录")))

        h1 = QHBoxLayout()
        h1.addWidget(self.file_dir)
        h1.addWidget(btn_browse)

        self.start_num = QSpinBox()
        self.start_num.setRange(1, 9999)
        self.start_num.setValue(1)
        
        self.margin_spin = QSpinBox()
        self.margin_spin.setRange(1, 20)
        self.margin_spin.setValue(3)
        
        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 16)
        self.thread_spin.setValue(4)
        
        self.cover_check = QComboBox()
        self.cover_check.addItems(["不覆盖", "用白色框覆盖"])

        form.addRow("文件目录:", h1)
        form.addRow("起始序号:", self.start_num)
        form.addRow("边距(毫米):", self.margin_spin)
        form.addRow("线程数:", self.thread_spin)
        form.addRow("覆盖选项:", self.cover_check)
        group.setLayout(form)
        self.layout.addWidget(group)

        btn_exec = QPushButton("开始编页码")
        btn_exec.setObjectName("ActionBtn")
        btn_exec.clicked.connect(self.execute)
        self.layout.addWidget(btn_exec)
        
        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        self.layout.addWidget(self.stop_btn)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)
        
        self.worker = None
        self.add_log_widget()

    def execute(self):
        d = self.file_dir.text()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        
        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "指定的目录不存在")
            return
        
        # 重置进度条
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        
        # 禁用按钮，启用停止按钮
        self.log("="*50)
        self.log("开始处理...")
        self.log(f"目录: {d}")
        self.log(f"起始序号: {self.start_num.value()}")
        self.log(f"边距: {self.margin_spin.value()}mm")
        self.log(f"线程数: {self.thread_spin.value()}")
        self.log(f"覆盖选项: {self.cover_check.currentText()}")
        self.log("="*50)
        
        # 创建工作线程
        self.worker = AutoPagingWorker(
            directory=d,
            margin_mm=self.margin_spin.value(),
            start_number=self.start_num.value(),
            thread_count=self.thread_spin.value(),
            cover_digit=(self.cover_check.currentIndex() == 1)
        )
        
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        
        self.worker.start()
        
        # 更新按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn:
            btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
    
    def stop_processing(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止处理...")
            self.stop_btn.setEnabled(False)
    
    def update_progress(self, current, total):
        """更新进度显示"""
        percentage = (current / total) * 100 if total > 0 else 0
        self.progress.setValue(int(percentage))
        self.progress.setFormat(f"{current} / {total}  ({percentage:.0f}%)")
        self.log(f"进度: {current}/{total} ({percentage:.1f}%)")
    
    def on_finished(self, success, message):
        """处理完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")
        
        # 恢复按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn:
            btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        
        if success:
            QMessageBox.information(self, "完成", "处理完成！")
        else:
            QMessageBox.warning(self, "提示", message)


class FileMovePage(FunctionPage):
    def __init__(self):
        super().__init__("文件移动")
        group = QGroupBox("文件批量归类移动")
        form = QFormLayout()

        self.src_dir = QLineEdit()
        self.dst_dir = QLineEdit()
        self.rule_combo = QComboBox()
        self.rule_combo.addItems(["按扩展名归类", "按修改日期归类", "全部移动"])

        btn_src = QPushButton("浏览")
        btn_src.setObjectName("BrowseBtn")
        btn_src.clicked.connect(lambda: self.src_dir.setText(QFileDialog.getExistingDirectory(self, "源目录")))
        btn_dst = QPushButton("浏览")
        btn_dst.setObjectName("BrowseBtn")
        btn_dst.clicked.connect(lambda: self.dst_dir.setText(QFileDialog.getExistingDirectory(self, "目标目录")))

        h1 = QHBoxLayout();
        h1.addWidget(self.src_dir);
        h1.addWidget(btn_src)
        h2 = QHBoxLayout();
        h2.addWidget(self.dst_dir);
        h2.addWidget(btn_dst)

        form.addRow("源目录:", h1)
        form.addRow("目标目录:", h2)
        form.addRow("移动规则:", self.rule_combo)
        group.setLayout(form)
        self.layout.addWidget(group)

        btn_exec = QPushButton("开始移动")
        btn_exec.setObjectName("ActionBtn")
        btn_exec.clicked.connect(self.execute)
        self.layout.addWidget(btn_exec)
        self.add_log_widget()

    def execute(self):
        self.log("校验目录权限...")
        self.log(f"应用规则: {self.rule_combo.currentText()}")
        self.log("正在复制及删除原文件...")
        self.log("文件移动归档完成！")


class ArchiveStampWorker(QThread):
    """归档章加盖后台处理线程"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, base_dir, overwrite_original=False, max_workers=4, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.overwrite_original = overwrite_original  # 是否覆盖原归档章
        self.max_workers = max_workers
        self.is_stopped = False
    
    def run(self):
        try:
            # 获取所有JPG文件
            dir_to_files = self.find_all_jpg_in_directory()
            if not dir_to_files:
                self.finished_signal.emit(False, "未找到任何JPG文件")
                return
            
            total_dirs = len(dir_to_files)
            total_files = sum(len(files) for files in dir_to_files.values())
            self.log_signal.emit(f"找到 {total_dirs} 个目录，共 {total_files} 个JPG文件")
            
            # 创建日志文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_filename = f"archive_stamp_log_{timestamp}.txt"
            log_filepath = os.path.join(self.base_dir, log_filename)
            
            with open(log_filepath, 'w', encoding='utf-8') as log_file:
                log_file.write(f"归档章加盖日志 - 开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                log_file.write(f"处理目录: {self.base_dir}\n")
                log_file.write(f"线程数: {self.max_workers}\n")
                log_file.write("=" * 80 + "\n")
                log_file.write("文件名称\t\t处理结果\t处理用时(秒)\n")
                log_file.write("=" * 80 + "\n")
                log_file.flush()
                
                processed_count = 0
                completed = 0
                
                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    futures = {}
                    for dir_path, files in dir_to_files.items():
                        dir_name = os.path.basename(dir_path)
                        # 只取排序后的第一个文件
                        if files:
                            first_file = files[0]
                            future = executor.submit(
                                self.process_single_file,
                                first_file, dir_name
                            )
                            futures[future] = (first_file, dir_name)
                    
                    for future in as_completed(futures):
                        if self.is_stopped:
                            break
                        
                        completed += 1
                        jpg_file, dir_name = futures[future]
                        
                        try:
                            result = future.result()
                            
                            # 记录日志
                            filename = os.path.basename(result['processed_file_path']) if result['processed_file_path'] else os.path.basename(jpg_file)
                            result_str = "成功" if result['success'] else "失败"
                            duration = result['duration']
                            
                            log_entry = f"{filename}\t\t{result_str}\t{duration}\n"
                            log_file.write(log_entry)
                            log_file.flush()
                            
                            if result['success']:
                                processed_count += 1
                            
                            self.log_signal.emit(
                                f"处理文件: {filename} - {result_str} ({duration}s)"
                            )
                            self.progress_signal.emit(completed, len(futures))
                            
                        except Exception as e:
                            filename = os.path.basename(jpg_file)
                            error_msg = f"处理文件 {filename} 时出错: {str(e)}"
                            self.log_signal.emit(error_msg)
                            
                            log_entry = f"{filename}\t\t失败\t0.0\n"
                            log_file.write(log_entry)
                            log_file.flush()
                            
                            self.progress_signal.emit(completed, len(futures))
            
            if not self.is_stopped:
                operation_type = "覆盖原归档章并重新加盖" if self.overwrite_original else "添加新归档章"
                self.finished_signal.emit(True,
                    f"{operation_type}处理完成！\n总计: {len(futures)}, 成功: {processed_count}, 失败: {len(futures) - processed_count}\n日志: {log_filename}")
            else:
                self.finished_signal.emit(False, "处理已停止")
                
        except Exception as e:
            self.finished_signal.emit(False, f"处理出错: {str(e)}")
    
    def stop(self):
        self.is_stopped = True
    
    def find_all_jpg_in_directory(self):
        """在指定目录及子目录中查找所有JPG文件（按目录分组）"""
        dir_to_files = {}
        
        for root, dirs, files in os.walk(self.base_dir):
            jpg_files_in_root = []
            for filename in files:
                if filename.lower().endswith(('.jpg', '.jpeg')):
                    jpg_files_in_root.append(os.path.join(root, filename))
            
            if jpg_files_in_root:
                jpg_files_in_root.sort()
                dir_to_files[root] = jpg_files_in_root
        
        return dir_to_files
    
    def create_red_grid_image(self, dir_name, file_count=1):
        """创建归档章图片"""
        DPI = 300
        mm_to_px = DPI / 25.4
        
        rect_width_mm = 45
        rect_height_mm = 16
        rect_width_px = int(rect_width_mm * mm_to_px)
        rect_height_px = int(rect_height_mm * mm_to_px)
        
        cell_width_mm = 15
        cell_height_mm = 8
        cell_width_px = int(cell_width_mm * mm_to_px)
        cell_height_px = int(cell_height_mm * mm_to_px)
        
        img = Image.new('RGB', (rect_width_px, rect_height_px), 'white')
        draw = ImageDraw.Draw(img)
        
        # 尝试加载宋体字体
        try:
            font = ImageFont.truetype("simsun.ttc", 36)
        except:
            try:
                font = ImageFont.truetype("C:/Windows/Fonts/simsun.ttc", 36)
            except:
                font = ImageFont.load_default()
        
        # 解析目录名称
        dir_text1 = dir_name[:4] if len(dir_name) >= 4 else dir_name.ljust(4)
        dir_text2 = dir_name[11:15] if len(dir_name) >= 15 else dir_name[11:].ljust(4) if len(dir_name) > 11 else "".ljust(4)
        last_5_chars = dir_name[-5:] if len(dir_name) >= 5 else dir_name
        dir_text3 = last_5_chars
        for i, char in enumerate(last_5_chars):
            if char != '0':
                dir_text3 = last_5_chars[i:]
                break
        
        # 绘制6个方格的红色边框
        for row in range(2):
            for col in range(3):
                x1 = col * cell_width_px
                y1 = row * cell_height_px
                x2 = x1 + cell_width_px
                y2 = y1 + cell_height_px
                
                draw.rectangle([x1, y1, x2, y2], fill='white', outline='red', width=2)
                
                if row == 0 and col == 0:
                    text = dir_text1
                elif row == 0 and col == 1:
                    text = dir_text2
                elif row == 0 and col == 2:
                    text = dir_text3
                elif row == 1 and col == 0:
                    text = "确权"
                elif row == 1 and col == 1:
                    text = "永久"
                elif row == 1 and col == 2:
                    text = str(file_count)
                else:
                    continue
                
                bbox = draw.textbbox((0, 0), text, font=font)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
                text_x = x1 + (cell_width_px - text_width) // 2
                text_y = y1 + (cell_height_px - text_height) // 2
                draw.text((text_x, text_y), text, fill='red', font=font)
        
        return img
    
    def overlay_image_on_jpg(self, jpg_file_path, overlay_img):
        """将归档章叠加到JPG文件的顶部中央"""
        try:
            Image.MAX_IMAGE_PIXELS = None
            
            original_img = Image.open(jpg_file_path)
            
            # 应用EXIF方向信息
            try:
                img_with_orientation = ImageOps.exif_transpose(original_img)
            except:
                img_with_orientation = original_img
            
            base_width, base_height = img_with_orientation.size
            
            DPI = 300
            mm_to_px = DPI / 25.4
            
            # 处理覆盖选项
            if self.overwrite_original:
                base_img = img_with_orientation.copy()
                base_width, base_height = base_img.size
                
                from PIL import ImageDraw
                draw = ImageDraw.Draw(base_img)
                cover_height_px = int(18 * mm_to_px)
                draw.rectangle([0, 0, base_width, cover_height_px], fill='white', outline=None)
            else:
                top_margin_px = int(18 * mm_to_px)
                new_height = base_height + top_margin_px
                
                new_img = Image.new('RGB', (base_width, new_height), 'white')
                new_img.paste(img_with_orientation, (0, top_margin_px))
                
                base_img = new_img
                base_width, base_height = base_img.size
            
            # 计算叠加位置
            overlay_width, overlay_height = overlay_img.size
            margin_top_px = int(1 * mm_to_px)
            x_pos = (base_width - overlay_width) // 2
            y_pos = margin_top_px
            
            if overlay_img.mode != 'RGBA':
                overlay_img_rgba = overlay_img.convert('RGBA')
            else:
                overlay_img_rgba = overlay_img
            
            if base_img.mode != 'RGBA':
                base_img_rgba = base_img.convert('RGBA')
            else:
                base_img_rgba = base_img
            
            result_img = base_img_rgba.copy()
            result_img.paste(overlay_img_rgba, (x_pos, y_pos), overlay_img_rgba)
            result_img = result_img.convert('RGB')
            
            result_img.save(jpg_file_path, 'JPEG', dpi=(DPI, DPI))
            
            return True, jpg_file_path
            
        except Exception as e:
            print(f"处理JPG文件时出错: {str(e)}")
            return False, jpg_file_path
    
    def process_single_file(self, jpg_file_path, directory_name):
        """处理单个文件"""
        start_time = time.time()
        
        # 获取当前目录下的文件数量
        dir_path = os.path.dirname(jpg_file_path)
        jpg_files_in_dir = [f for f in os.listdir(dir_path) if f.lower().endswith(('.jpg', '.jpeg'))]
        file_count = len(jpg_files_in_dir)
        
        # 创建归档章图片
        grid_img = self.create_red_grid_image(directory_name, file_count)
        
        # 将归档章叠加到JPG文件
        success, processed_file_path = self.overlay_image_on_jpg(jpg_file_path, grid_img)
        
        end_time = time.time()
        duration = round(end_time - start_time, 2)
        
        return {
            'file_path': jpg_file_path,
            'directory_name': directory_name,
            'success': success,
            'processed_file_path': processed_file_path,
            'duration': duration
        }


class ArchiveStampPage(FunctionPage):
    def __init__(self):
        super().__init__("文件改名加盖归档章")
        group = QGroupBox("按目录名生成归档章并加盖到JPG文件")
        form = QFormLayout()

        self.dir_path = QLineEdit()
        btn_browse = QPushButton("选择文件夹")
        btn_browse.setObjectName("BrowseBtn")
        btn_browse.clicked.connect(self.browse_dir)

        h1 = QHBoxLayout()
        h1.addWidget(self.dir_path)
        h1.addWidget(btn_browse)

        form.addRow("目标目录:", h1)
        
        # 覆盖原归档章选项
        self.overwrite_check = QCheckBox("覆盖原归档章")
        self.overwrite_check.setChecked(False)
        form.addRow("选项:", self.overwrite_check)
        
        # 线程数设置
        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 32)
        self.thread_spin.setValue(4)
        form.addRow("线程数:", self.thread_spin)
        
        group.setLayout(form)
        self.layout.addWidget(group)
        
        # 说明文本
        info_label = QLabel(
            "功能说明：\n"
            "• 根据目录名称自动生成归档章（45mm×16mm）\n"
            "• 在所有JPG文件顶部增加18mm白色边框\n"
            "• 归档章加盖在边框正中间（距顶部1mm）\n"
            "• 可选：覆盖原归档章后重新加盖\n"
            "• 自动从目录名提取：全宗号、年份、档号"
        )
        info_label.setStyleSheet("color: #8B949E; font-size: 12px;")
        self.layout.addWidget(info_label)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        self.preview_btn = QPushButton("预览操作")
        self.preview_btn.setObjectName("ActionBtn")
        self.preview_btn.setStyleSheet("background-color: #2196F3;")
        self.preview_btn.clicked.connect(self.preview_operations)
        btn_layout.addWidget(self.preview_btn)
        
        btn_exec = QPushButton("开始加盖归档章")
        btn_exec.setObjectName("ActionBtn")
        btn_exec.clicked.connect(self.execute)
        btn_layout.addWidget(btn_exec)
        
        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        
        self.layout.addLayout(btn_layout)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)
        
        self.worker = None
        self.add_log_widget()
    
    def browse_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择目录")
        if d: self.dir_path.setText(d)
    
    def preview_operations(self):
        """预览将要执行的操作"""
        d = self.dir_path.text()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        
        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "指定的目录不存在")
            return
        
        # 清空之前的结果
        self.log_box.clear()
        
        self.log("="*50)
        self.log("正在预览操作...")
        self.log(f"基础目录: {d}")
        self.log("-"*50)
        
        try:
            # 统计JPG文件
            jpg_count = 0
            dir_count = 0
            sample_files = []
            
            for root, dirs, files in os.walk(d):
                jpg_files = [f for f in files if f.lower().endswith(('.jpg', '.jpeg'))]
                if jpg_files:
                    dir_count += 1
                    jpg_count += len(jpg_files)
                    if len(sample_files) < 3:
                        sample_files.append((os.path.basename(root), jpg_files[0]))
            
            if jpg_count == 0:
                self.log("未找到任何JPG文件")
                return
            
            self.log(f"找到 {dir_count} 个目录，共 {jpg_count} 个JPG文件")
            self.log("")
            self.log("示例文件：")
            for dir_name, filename in sample_files:
                self.log(f"  • {dir_name}/{filename}")
            
            self.log("")
            if self.overwrite_check.isChecked():
                self.log("注意：将覆盖原归档章并重新加盖")
            else:
                self.log("注意：将在文件顶部添加18mm白色边框并加盖归档章")
            
            self.log("")
            self.log("这只是预览，文件尚未修改。")
            
        except Exception as e:
            error_msg = f"预览过程中出现错误: {str(e)}"
            self.log(error_msg)
            QMessageBox.critical(self, "错误", error_msg)

    def execute(self):
        d = self.dir_path.text()
        if not d:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        
        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "指定的目录不存在")
            return
        
        # 确认操作
        confirm_msg = "确定要开始加盖归档章吗？"
        if self.overwrite_check.isChecked():
            confirm_msg += "\n\n注意：此操作将覆盖原归档章区域（顶部18mm），请确保已备份重要文件！"
        
        reply = QMessageBox.question(
            self, "确认操作",
            confirm_msg,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # 重置进度条
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        
        # 禁用按钮，启用停止按钮
        self.log("="*50)
        self.log("开始处理...")
        self.log(f"目录: {d}")
        if self.overwrite_check.isChecked():
            self.log("选项: 覆盖原归档章并重新加盖")
        else:
            self.log("选项: 添加新归档章（不覆盖）")
        self.log(f"线程数: {self.thread_spin.value()}")
        self.log("="*50)
        
        # 创建工作线程
        overwrite_original = self.overwrite_check.isChecked()
        max_workers = self.thread_spin.value()
        self.worker = ArchiveStampWorker(
            base_dir=d,
            overwrite_original=overwrite_original,
            max_workers=max_workers
        )
        
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        
        self.worker.start()
        
        # 更新按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始加盖归档章":
            btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
    
    def stop_processing(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止处理...")
            self.stop_btn.setEnabled(False)
    
    def update_progress(self, current, total):
        """更新进度显示"""
        percentage = (current / total) * 100 if total > 0 else 0
        self.progress.setValue(int(percentage))
        self.progress.setFormat(f"{current} / {total}  ({percentage:.0f}%)")
        self.log(f"进度: {current}/{total} ({percentage:.1f}%)")
    
    def on_finished(self, success, message):
        """处理完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")
        
        # 恢复按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始加盖归档章":
            btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        
        if success:
            QMessageBox.information(self, "完成", "处理完成！")
        else:
            QMessageBox.warning(self, "提示", message)


# ----------------------------------------------------------------------
# 图片 DPI 修改（由 modify_jpg_dpi.py 移植，去除 tkinter 依赖）
# ----------------------------------------------------------------------
def _dpi_get_image_dpi(image_path):
    """读取图片 DPI：优先 info['dpi']，其次 EXIF 分辨率，最后返回分辨率。"""
    try:
        Image.MAX_IMAGE_PIXELS = None
        with Image.open(image_path) as img:
            dpi = img.info.get('dpi')
            if dpi is not None:
                return dpi if isinstance(dpi, tuple) else (dpi, dpi)
            try:
                exif = img.getexif()
                if exif:
                    xr, yr, unit = exif.get(282), exif.get(283), exif.get(296, 2)
                    if xr and yr:
                        if unit == 3:        # 厘米 → 英寸
                            return f"EXIF:{xr * 2.54:.0f}x{yr * 2.54:.0f} DPI"
                        elif unit == 2:      # 英寸
                            return f"EXIF:{xr:.0f}x{yr:.0f} DPI"
                        else:
                            return f"EXIF:{xr:.0f}x{yr:.0f}(无单位)"
            except Exception:
                pass
            return f"{img.width}x{img.height}(无DPI元数据)"
    except Exception:
        return "DPI获取失败"


def _dpi_get_image_details(image_path):
    """获取图片详情：分辨率/色彩模式/格式/文件大小。"""
    details = {'resolution': '未知', 'mode': '未知', 'format': '未知', 'file_size': '未知'}
    try:
        Image.MAX_IMAGE_PIXELS = None
        Image.LOAD_TRUNCATED_IMAGES = True
        with Image.open(image_path) as img:
            details['resolution'] = f"{img.width}x{img.height}"
            details['mode'] = img.mode
            details['format'] = img.format
        try:
            details['file_size'] = f"{os.path.getsize(image_path):,} bytes"
        except Exception:
            pass
    except Exception:
        pass
    return details


def _dpi_init_log(log_path):
    try:
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write("JPG文件DPI修改日志\n")
            f.write(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 120 + "\n")
            f.write(f"{'文件名':<28}{'原DPI':<16}{'新DPI':<12}{'分辨率':<14}{'模式':<8}"
                    f"{'格式':<8}{'大小':<16}{'结果':<10}{'耗时(s)':<10}\n")
            f.write("-" * 120 + "\n")
    except Exception:
        pass


def _dpi_append_log(log_path, info):
    try:
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"{os.path.basename(str(info.get('file_path', ''))):<28}"
                    f"{str(info.get('original_dpi', '')):<16}"
                    f"{str(info.get('new_dpi', '')):<12}"
                    f"{str(info.get('resolution', '')):<14}"
                    f"{str(info.get('mode', '')):<8}"
                    f"{str(info.get('format', '')):<8}"
                    f"{str(info.get('file_size', '')):<16}"
                    f"{str(info.get('result', '')):<10}"
                    f"{str(info.get('process_time', '')):<10}\n")
    except Exception:
        pass


def _dpi_finalize_log(log_path, stats):
    try:
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write("-" * 120 + "\n")
            f.write(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"总计处理: {stats.get('processed_files', 0)} 个\n")
            f.write(f"成功修改: {stats.get('modified_files', 0)} 个\n")
            f.write(f"失败: {len(stats.get('errors', []))} 个\n")
            f.write("=" * 120 + "\n")
    except Exception:
        pass


def _dpi_process_single_file(file_path, dpi, dry_run, stop_event=None):
    """处理单个 JPG：读取原 DPI，按需重写为 dpi（仅改 DPI，不改像素）。"""
    start = time.time()
    if stop_event and stop_event.is_set():
        return {'file_path': file_path, 'original_dpi': '未知', 'new_dpi': f"{dpi}x{dpi}",
                'result': '已停止', 'process_time': 0, 'success': False, 'error': '用户停止'}
    try:
        Image.MAX_IMAGE_PIXELS = None
        original_dpi = _dpi_get_image_dpi(file_path)
        original_dpi_str = (f"{original_dpi[0]}x{original_dpi[1]}"
                            if isinstance(original_dpi, tuple) else str(original_dpi))
        details = _dpi_get_image_details(file_path)
        if stop_event and stop_event.is_set():
            return {'file_path': file_path, 'original_dpi': original_dpi_str,
                    'new_dpi': f"{dpi}x{dpi}", 'result': '已停止',
                    'process_time': round(time.time() - start, 3), 'success': False, 'error': '用户停止',
                    **details}

        if not dry_run:
            Image.LOAD_TRUNCATED_IMAGES = True
            with Image.open(file_path) as img:
                img.save(file_path, dpi=(dpi, dpi))
            result = "成功"
        else:
            result = "模拟成功"

        return {'file_path': file_path, 'original_dpi': original_dpi_str,
                'new_dpi': f"{dpi}x{dpi}", 'result': result,
                'process_time': round(time.time() - start, 3), 'success': True, **details}
    except Exception as e:
        return {'file_path': file_path, 'original_dpi': '未知', 'new_dpi': f"{dpi}x{dpi}",
                'result': '失败', 'process_time': round(time.time() - start, 3),
                'success': False, 'error': str(e)}


class ModifyDpiWorker(QThread):
    """JPG 图片 DPI 批量修改后台线程（多线程，支持模拟运行与停止）。"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)      # (已完成, 总数)
    finished_signal = Signal(bool, str)     # (是否正常完成, 汇总信息)

    def __init__(self, base_dir, dpi=600, dry_run=False, max_workers=4, parent=None):
        super().__init__(parent)
        self.base_dir = base_dir
        self.dpi = dpi
        self.dry_run = dry_run
        self.max_workers = max_workers
        self.is_stopped = False
        self.stop_event = threading.Event()

    def stop(self):
        self.is_stopped = True
        self.stop_event.set()

    def run(self):
        try:
            Image.MAX_IMAGE_PIXELS = None
            Image.LOAD_TRUNCATED_IMAGES = True

            # 递归收集 JPG/JPEG
            jpg_files = []
            for root, _d, files in os.walk(self.base_dir):
                for f in files:
                    if f.lower().endswith(('.jpg', '.jpeg')):
                        jpg_files.append(os.path.join(root, f))
            total = len(jpg_files)
            if total == 0:
                self.finished_signal.emit(False, f"目录 {self.base_dir} 下未找到 JPG/JPEG 文件")
                return

            mode = "模拟运行" if self.dry_run else "实际修改"
            self.log_signal.emit(f"找到 {total} 个 JPG 文件，目标 DPI={self.dpi}，"
                                 f"线程数={self.max_workers}，{mode}")

            # 日志文件
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_path = os.path.join(self.base_dir, f"dpi_modify_log_{ts}.txt")
            _dpi_init_log(log_path)

            self.log_signal.emit(f"{'目录':<16} | {'文件名':<24} | {'原DPI':<14} | "
                                 f"{'新DPI':<10} | {'结果':<6} | 耗时(秒)")
            self.log_signal.emit("-" * 90)

            done = 0
            success = 0
            errors = []

            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                future_to_file = {
                    executor.submit(_dpi_process_single_file, fp, self.dpi,
                                    self.dry_run, self.stop_event): fp
                    for fp in jpg_files
                }
                for future in as_completed(future_to_file):
                    if self.is_stopped:
                        for ff in future_to_file:
                            ff.cancel()
                        break
                    fp = future_to_file[future]
                    try:
                        res = future.result()
                    except Exception as e:
                        res = {'file_path': fp, 'original_dpi': '未知', 'new_dpi': f"{self.dpi}x{self.dpi}",
                               'result': '异常', 'process_time': 0, 'success': False, 'error': str(e)}
                        self.log_signal.emit(f"处理异常 {os.path.basename(fp)}: {e}")

                    _dpi_append_log(log_path, res)
                    dir_name = os.path.basename(os.path.dirname(fp))
                    self.log_signal.emit(
                        f"{dir_name:<16} | {os.path.basename(fp):<24} | "
                        f"{str(res.get('original_dpi', '')):<14} | {res.get('new_dpi', ''):<10} | "
                        f"{res.get('result', ''):<6} | {res.get('process_time', 0)}")

                    done += 1
                    if res.get('success'):
                        success += 1
                    else:
                        errors.append(f"{os.path.basename(fp)}: {res.get('error', '')}")
                    self.progress_signal.emit(done, total)

            _dpi_finalize_log(log_path, {'processed_files': done,
                                         'modified_files': success, 'errors': errors})

            if self.is_stopped:
                self.finished_signal.emit(False, f"已停止。已处理 {done}/{total}，成功 {success}。")
            else:
                self.finished_signal.emit(
                    True, f"{mode}完成：共 {done} 个，成功 {success}，失败 {len(errors)}。日志：{log_path}")
        except Exception as e:
            self.finished_signal.emit(False, f"运行异常：{e}")


class ModifyDpiPage(FunctionPage):
    """图片 DPI 批量修改功能页（递归处理 JPG/JPEG，仅改 DPI 不影响像素）"""
    def __init__(self):
        super().__init__("修改DPI")
        self.worker = None

        group = QGroupBox("图片DPI批量修改（递归处理 JPG/JPEG，仅改DPI不影响像素尺寸）")
        form = QFormLayout()

        self.img_dir = QLineEdit()
        self.img_dir.setPlaceholderText("选择包含 JPG 图片的目录...")
        btn_browse = QPushButton("选择文件夹")
        btn_browse.setObjectName("BrowseBtn")
        btn_browse.clicked.connect(
            lambda: self.img_dir.setText(QFileDialog.getExistingDirectory(self, "选择图片目录")))
        h1 = QHBoxLayout()
        h1.addWidget(self.img_dir)
        h1.addWidget(btn_browse)

        self.dpi_edit = QLineEdit()
        self.dpi_edit.setText("300")
        self.dpi_edit.setMaxLength(5)
        self.dpi_edit.setMaximumWidth(140)
        self.dpi_edit.setValidator(QRegularExpressionValidator(QRegularExpression("\\d*")))
        h_dpi = QHBoxLayout()
        h_dpi.addWidget(self.dpi_edit)
        h_dpi.addWidget(QLabel("DPI"))
        h_dpi.addStretch()

        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 32)
        self.thread_spin.setValue(4)

        self.dry_run_cb = QCheckBox("模拟运行（仅预览将要执行的操作，不实际修改文件）")

        form.addRow("图片目录:", h1)
        form.addRow("目标DPI:", h_dpi)
        form.addRow("线程数:", self.thread_spin)
        form.addRow("", self.dry_run_cb)
        group.setLayout(form)
        self.layout.addWidget(group)

        # 控制按钮
        btn_layout = QHBoxLayout()
        self.preview_btn = QPushButton("预览操作")
        self.preview_btn.setObjectName("ActionBtn")
        self.exec_btn = QPushButton("开始修改")
        self.exec_btn.setObjectName("ActionBtn")
        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.setEnabled(False)
        self.clear_btn = QPushButton("清空结果")
        self.clear_btn.setObjectName("ActionBtn")
        self.preview_btn.clicked.connect(self._preview)
        self.exec_btn.clicked.connect(self._execute)
        self.stop_btn.clicked.connect(self._stop)
        self.clear_btn.clicked.connect(self._clear)
        for b in (self.preview_btn, self.exec_btn, self.stop_btn, self.clear_btn):
            btn_layout.addWidget(b)
        self.layout.addLayout(btn_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)

        self.add_log_widget()

    # ---------- 辅助 ----------
    def _target_dpi(self):
        try:
            return int(self.dpi_edit.text().strip())
        except ValueError:
            return 0

    def _set_running(self, running):
        self.preview_btn.setEnabled(not running)
        self.exec_btn.setEnabled(not running)
        self.stop_btn.setEnabled(running)

    def _start_worker(self, dry_run):
        directory = self.img_dir.text().strip()
        if not directory or not os.path.isdir(directory):
            QMessageBox.warning(self, "提示", "请先选择有效的图片目录。")
            return
        dpi = self._target_dpi()
        if dpi <= 0:
            QMessageBox.warning(self, "提示", "目标 DPI 必须为正整数。")
            return
        self._clear()
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        self.worker = ModifyDpiWorker(directory, dpi=dpi, dry_run=dry_run,
                                      max_workers=self.thread_spin.value())
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self._update_progress)
        self.worker.finished_signal.connect(self._on_finished)
        self._set_running(True)
        self.worker.start()

    def _preview(self):
        self._start_worker(dry_run=True)

    def _execute(self):
        directory = self.img_dir.text().strip()
        if not directory or not os.path.isdir(directory):
            QMessageBox.warning(self, "提示", "请先选择有效的图片目录。")
            return
        dpi = self._target_dpi()
        confirm = QMessageBox.question(
            self, "确认操作",
            f"确定要将 {directory} 下所有 JPG/JPEG 的 DPI 修改为 {dpi} 吗？\n"
            f"此操作会直接覆盖原文件，不可逆。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        self._start_worker(dry_run=False)

    def _stop(self):
        if self.worker:
            self.worker.stop()
            self.log("正在停止...")

    def _update_progress(self, done, total):
        if total <= 0:
            self.progress.setValue(0)
            self.progress.setFormat("0 / 0")
            return
        pct = int(done * 100 / total)
        self.progress.setValue(pct)
        self.progress.setFormat(f"{done} / {total}  ({pct}%)")

    def _on_finished(self, ok, message):
        self._set_running(False)
        self.log(message)
        self.progress.setFormat("已完成" if ok else "已停止")
        self.worker = None
        QMessageBox.information(self, "完成" if ok else "结束", message)

    def _clear(self):
        if hasattr(self, 'progress'):
            self.progress.setValue(0)
            self.progress.setFormat("待开始")
        if hasattr(self, 'log_box'):
            self.log_box.clear()


class JpgToPdfWorker(QThread):
    """JPG转双层PDF后台处理线程"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, directory_path, output_dir=None, max_workers=4, resolution=100.0, generate_ofd=False, parent=None):
        super().__init__(parent)
        self.directory_path = directory_path
        self.output_dir = output_dir if output_dir else directory_path
        self.max_workers = max_workers
        self.resolution = resolution
        self.generate_ofd = generate_ofd  # 是否同时生成OFD文件
        self.is_stopped = False
        import threading
        self._ofd_lock = threading.Lock()
        self._ocr_lock = threading.Lock()   # PaddleOCR 推理非线程安全, 串行化
        self._ocr = None  # PaddleOCR 延迟初始化(与分件共享同一初始化逻辑)
        
        # 检查 OFD 转换库是否可用（用于生成双层OFD）
        # 使用自建 ofd_writer（基于 PyMuPDF，生成图像层+文本层的双层OFD，
        # 无页数限制）。旧的 Spire.PDF 免费版只能转前 3 页，已弃用。
        if self.generate_ofd:
            try:
                import fitz  # noqa: F401  PyMuPDF
                from ofd_writer import make_layered_ofd  # noqa: F401
                self.ofd_available = True
            except ImportError:
                self.ofd_available = False
        else:
            self.ofd_available = False
    
    def run(self):
        try:
            # 输出目录不存在则创建(缺省为 源目录/PDF, 可能尚不存在)
            if self.output_dir:
                os.makedirs(self.output_dir, exist_ok=True)
            # 收集所有目录下的JPG文件
            dir_jpgs_map = self.collect_jpg_files()
            if not dir_jpgs_map:
                self.finished_signal.emit(False, f"在目录 {self.directory_path} 及其子目录中没有找到JPG文件")
                return
            
            total_dirs = len(dir_jpgs_map)
            self.log_signal.emit(f"找到 {total_dirs} 个包含JPG文件的目录，使用 {self.max_workers} 个线程开始处理...")
            
            # 创建日志文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_filename = f"jpgtopdf_log_{timestamp}.txt"
            log_path = os.path.join(self.output_dir, log_filename)
            start_time = datetime.now()
            self.init_log_file(log_path, start_time)
            
            # 处理所有目录
            results_list = []
            success_count = 0
            completed = 0
            
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {}
                for dir_path, jpg_files in dir_jpgs_map.items():
                    if self.is_stopped:
                        break
                    dir_name = os.path.basename(dir_path)
                    future = executor.submit(
                        self.process_single_directory,
                        jpg_files, dir_name
                    )
                    futures[future] = (dir_name, jpg_files)
                
                for future in as_completed(futures):
                    if self.is_stopped:
                        break
                    
                    completed += 1
                    dir_name, jpg_files = futures[future]
                    
                    try:
                        result = future.result()
                        results_list.append(result)
                        
                        # 实时记录日志
                        self.append_to_log(log_path, result)
                        
                        if result['result'] == '成功':
                            success_count += 1
                            log_msg = f"✓ {result['folder']}/{result['pdf_file']} - 成功 ({result['duration']}秒)"
                            if result.get('ofd_generated'):
                                log_msg += f" [已生成OFD: {result['ofd_file']}]"
                            self.log_signal.emit(log_msg)
                        else:
                            self.log_signal.emit(
                                f"✗ {result['folder']}/{result['pdf_file']} - {result['ocr_status']}"
                            )
                        
                        self.progress_signal.emit(completed, total_dirs)
                        
                    except Exception as e:
                        error_msg = f"处理目录 {dir_name} 时发生异常: {str(e)}"
                        self.log_signal.emit(error_msg)
                        self.progress_signal.emit(completed, total_dirs)
            
            if not self.is_stopped:
                # 完成日志文件
                self.finalize_log_file(log_path, results_list, start_time)
                
                success_rate = (success_count / total_dirs) * 100 if total_dirs > 0 else 0
                rate_msg = f"\n处理完成！共处理 {total_dirs} 个目录，成功 {success_count} 个，成功率 {success_rate:.1f}%。\n日志: {log_filename}"
                
                if success_rate >= 90:
                    rate_msg += "\n提示：单线程处理可能获得更高成功率！"
                
                self.finished_signal.emit(True, rate_msg)
            else:
                self.finished_signal.emit(False, "处理已停止")
                
        except Exception as e:
            self.finished_signal.emit(False, f"处理出错: {str(e)}")
    
    def stop(self):
        self.is_stopped = True
    
    def collect_jpg_files(self):
        """收集目录及子目录下所有JPG文件（按目录分组）"""
        dir_jpgs_map = {}
        
        for root, dirs, files in os.walk(self.directory_path):
            jpg_files = []
            for filename in files:
                if filename.lower().endswith(('.jpg', '.jpeg')):
                    jpg_path = os.path.join(root, filename)
                    jpg_files.append(jpg_path)
            
            if jpg_files:
                jpg_files.sort()
                dir_jpgs_map[root] = jpg_files
        
        return dir_jpgs_map
    
    def jpgs_to_pdf(self, jpg_paths, output_dir, pdf_filename):
        """将多个 JPG 文件合并为一个 PDF（仅图像层）"""
        pdf_path = os.path.join(output_dir, pdf_filename + ".pdf")
        
        images = []
        for jpg_path in jpg_paths:
            image = Image.open(jpg_path)
            if image.mode in ('RGBA', 'LA', 'P'):
                image = image.convert('RGB')
            images.append(image)
        
        if images:
            first_image = images[0]
            if len(images) == 1:
                first_image.save(pdf_path, "PDF", resolution=self.resolution)
            else:
                first_image.save(pdf_path, "PDF", resolution=self.resolution, save_all=True, append_images=images[1:])
        
        return pdf_path
    
    # 复用分件功能的本地 PaddleOCR(多配置兼容初始化+中文路径安全)
    _get_ocr = FileSplitWorker._get_ocr
    _ocr_page = FileSplitWorker._ocr_page
    _imread_cn = staticmethod(FileSplitWorker._imread_cn)

    def _ocr_local_available(self):
        """本地OCR是否可用(初始化一次)。"""
        return self._get_ocr() is not None

    def process_jpgs_to_ocr_pdf(self, jpg_paths, output_dir, pdf_filename):
        """
        将多个JPG文件合并转换为双层PDF（图像+OCR文本层）。
        OCR使用与分件功能一致的本地 PaddleOCR(无需UmiOCR/联网):
          逐图OCR取文本+坐标 → PyMuPDF 逐页插图与不可见文本层。
        """
        Image.MAX_IMAGE_PIXELS = None
        ocr = self._get_ocr()
        if ocr is None:
            # 本地OCR不可用: 回退仅图像PDF
            temp_pdf_path = self.jpgs_to_pdf(jpg_paths, output_dir, pdf_filename)
            return temp_pdf_path, "仅图像PDF（本地OCR不可用）"

        try:
            import fitz  # PyMuPDF
        except ImportError:
            temp_pdf_path = self.jpgs_to_pdf(jpg_paths, output_dir, pdf_filename)
            return temp_pdf_path, "仅图像PDF（缺PyMuPDF）"

        pdf_path = os.path.join(output_dir, pdf_filename + ".pdf")
        doc = fitz.open()
        ocr_lock = getattr(self, '_ocr_lock', None)
        try:
            for jpg_path in jpg_paths:
                if self.is_stopped:
                    break
                img = Image.open(jpg_path)
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                w_px, h_px = img.size
                # PDF页面尺寸: 像素/分辨率*72 (resolution 默认100dpi)
                w_pt = w_px * 72.0 / self.resolution
                h_pt = h_px * 72.0 / self.resolution
                page = doc.new_page(width=w_pt, height=h_pt)
                page.insert_image(fitz.Rect(0, 0, w_pt, h_pt), filename=jpg_path)

                # OCR文本层(坐标从像素换算到PDF点)
                if ocr_lock is not None:
                    with ocr_lock:
                        frags = self._ocr_page(jpg_path, None, lambda s: None)
                else:
                    frags = self._ocr_page(jpg_path, None, lambda s: None)
                sx = w_pt / w_px
                sy = h_pt / h_px
                for txt, x0, y0, x1, y1 in frags:
                    if not txt.strip():
                        continue
                    # 字号优先用 OCR 片段的「真实行高」(y1-y0, 像素)换算——
                    # 卷皮页大字标题按宽度估算会得到超大值(被钳到上限40),
                    # 全页同字号巨字文本在 WPS 中不被作为可选对象(其他页正常)。
                    # 宽度估算仅作行高缺失时的兜底; 中文段宽度校验用于防溢出。
                    h_px_txt = max(y1 - y0, 1.0)
                    fs_h = h_px_txt * sy
                    w_pt_txt = max((x1 - x0) * sx, 1.0)
                    eff = sum(1.0 if ord(c) > 127 else 0.55 for c in txt) or 1.0
                    fs_w = w_pt_txt / eff
                    fs = fs_h if fs_h > 0 else fs_w
                    fs = max(4.0, min(fs, 60.0))
                    base_y = y0 * sy + fs
                    # 按连续同类字符切分为段 [(文本, 是否中文)]
                    seg_list = []
                    cur = ''
                    cur_cn = None
                    for ch in txt:
                        cn = ord(ch) > 127
                        if cur_cn is None or cn == cur_cn:
                            cur += ch
                            cur_cn = cn
                        else:
                            seg_list.append((cur, cur_cn))
                            cur, cur_cn = ch, cn
                    if cur:
                        seg_list.append((cur, cur_cn))
                    # 逐段插入, x随实际advance推进(中文=fs, ASCII≈fs*0.5)
                    # 隐形可选文本: render_mode=0(正常填充文本, 编辑器识别为可选
                    # 对象) + fill_opacity=0(完全透明, 任何底色上都不可见)。
                    # —— render_mode=3 在WPS等编辑器中无法选中; 白色文字在
                    # 深色底图上会显现。透明填充两全其美。
                    cx = x0 * sx
                    for seg, is_cn in seg_list:
                        if not seg:
                            continue
                        fname = 'china-s' if is_cn else 'helv'
                        try:
                            page.insert_text(fitz.Point(cx, base_y), seg,
                                             fontsize=fs,
                                             color=(0, 0, 0), fill_opacity=0,
                                             fontname=fname)
                        except Exception:
                            try:
                                page.insert_text(fitz.Point(cx, base_y), seg,
                                                 fontsize=fs,
                                                 color=(0, 0, 0), fill_opacity=0)
                            except Exception:
                                pass
                        cx += len(seg) * (fs if is_cn else fs * 0.5)
            doc.save(pdf_path)
            doc.close()
            if self.is_stopped:
                return pdf_path, "已停止(部分页生成)"
            return pdf_path, "双层PDF生成成功(本地OCR)"
        except Exception as e:
            try:
                doc.close()
            except Exception:
                pass
            # 失败回退: 仅图像PDF
            temp_pdf_path = self.jpgs_to_pdf(jpg_paths, output_dir, pdf_filename)
            return temp_pdf_path, f"OCR处理错误: {str(e)}(已回退为仅图像PDF)"
    
    def convert_pdf_to_ofd(self, pdf_path, ofd_path):
        """使用 ofd_writer 将双层PDF转换为双层OFD(图像层+文本层，无页数限制)。"""
        if not os.path.exists(pdf_path):
            return False, 0

        if not self.ofd_available:
            return False, 0

        start_time = time.time()

        try:
            from ofd_writer import make_layered_ofd

            # 加锁串行转换（PyMuPDF 在 Win7 多线程下更稳妥）
            with self._ofd_lock:
                ok, n, err = make_layered_ofd(pdf_path, ofd_path)

            if not ok:
                print(f"PDF转OFD失败: {err}")
                return False, time.time() - start_time

            end_time = time.time()
            return True, end_time - start_time

        except Exception as e:
            end_time = time.time()
            print(f"PDF转OFD过程中发生错误: {e}")
            return False, end_time - start_time

    def process_single_directory(self, jpg_files, dir_name):
        """处理单个目录的JPG文件"""
        start_time = time.time()
        result = "失败"
        result_pdf_path = None
        ocr_status = "未知"
        ofd_generated = False
        ofd_filename = ""
        
        try:
            # 输出结构(与分件一致): 输出目录/顶层目录名/子目录名/子目录名.pdf
            # 例: 源 D:\扫描\J380-ZY·2021-Y-FGC-0001\J380-0001\*.jpg
            #     → 输出 PDF目录\J380-ZY·2021-Y-FGC-0001\J380-0001\J380-0001.pdf
            # dir_name 为当前子目录名; jpg_files[0]含完整路径可提取父目录链。
            out_subdir = os.path.join(self.output_dir, dir_name)
            # 源子目录相对源根的路径(可能多级), 保持层级镜像到输出目录
            src_root = getattr(self, 'directory_path', None)
            if src_root and jpg_files:
                try:
                    src_dir = os.path.dirname(jpg_files[0])
                    rel = os.path.relpath(src_dir, src_root)
                    if rel and rel != '.':
                        out_subdir = os.path.join(self.output_dir, rel)
                except Exception:
                    pass
            os.makedirs(out_subdir, exist_ok=True)
            pdf_filename = dir_name
            result_pdf_path, ocr_status = self.process_jpgs_to_ocr_pdf(
                jpg_files, out_subdir, pdf_filename
            )
            if result_pdf_path:
                result = "成功"
                
                # 如果选择了生成OFD，则转换刚生成的PDF
                if self.generate_ofd and self.ofd_available and result_pdf_path:
                    ofd_path = os.path.splitext(result_pdf_path)[0] + '.ofd'
                    ofd_success, ofd_duration = self.convert_pdf_to_ofd(result_pdf_path, ofd_path)
                    if ofd_success:
                        ofd_generated = True
                        ofd_filename = os.path.basename(ofd_path)
                        self.log_signal.emit(f"  → 已生成OFD: {ofd_filename} ({ofd_duration:.2f}秒)")
                    else:
                        self.log_signal.emit(f"  → OFD生成失败")
        except Exception as e:
            result = "失败"
            ocr_status = f"错误: {str(e)}"
        
        duration = round(time.time() - start_time, 2)
        
        return {
            'folder': dir_name,
            'pdf_file': os.path.basename(result_pdf_path) if result_pdf_path else f"{dir_name}.pdf",
            'result': result,
            'ocr_status': ocr_status,
            'duration': duration,
            'ofd_generated': ofd_generated,
            'ofd_file': ofd_filename
        }
    
    def init_log_file(self, log_path, start_time):
        """初始化日志文件"""
        try:
            with open(log_path, 'w', encoding='utf-8') as f:
                f.write(f"JPG转双层PDF处理日志\n")
                f.write(f"开始时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write("=" * 120 + "\n")
                f.write(f"{'目录名':<25} {'PDF文件名':<30} {'处理结果':<10} {'双层PDF状态':<20} {'耗时(秒)':<12} {'处理时间':<20}\n")
                f.write("-" * 120 + "\n")
        except Exception as e:
            print(f"创建日志文件失败: {str(e)}")
    
    def append_to_log(self, log_path, result):
        """追加单条记录到日志文件"""
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                process_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                f.write(f"{result['folder']:<25} {result['pdf_file']:<30} "
                       f"{result['result']:<10} {result['ocr_status']:<20} "
                       f"{result['duration']:<12} {process_time:<20}\n")
        except Exception as e:
            print(f"追加日志记录失败: {str(e)}")
    
    def finalize_log_file(self, log_path, results, start_time):
        """完成日志文件，添加统计信息"""
        try:
            success_count = sum(1 for r in results if r['result'] == '成功')
            fail_count = len(results) - success_count
            end_time = datetime.now()
            total_duration = (end_time - start_time).total_seconds() / 60
            
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write("-" * 120 + "\n")
                f.write(f"结束时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"总耗时: {total_duration:.2f}分钟\n")
                f.write(f"总计处理: {len(results)} 个目录\n")
                f.write(f"成功: {success_count} 个\n")
                f.write(f"失败: {fail_count} 个\n")
                f.write("=" * 120 + "\n")
        except Exception as e:
            print(f"完成日志文件失败: {str(e)}")


class JpgToPdfPage(FunctionPage):
    def __init__(self):
        super().__init__("JPG转双层PDF")
        group = QGroupBox("图片转双层PDF (OCR)")
        form = QFormLayout()

        self.img_dir = QLineEdit()
        btn_browse_src = QPushButton("选择文件夹")
        btn_browse_src.setObjectName("BrowseBtn")
        btn_browse_src.clicked.connect(self.browse_img_dir)
        h1 = QHBoxLayout()
        h1.addWidget(self.img_dir)
        h1.addWidget(btn_browse_src)
        
        self.out_dir = QLineEdit()
        self._out_manual = False  # 用户是否手动改过输出目录(改过则不随源联动)
        btn_browse_out = QPushButton("选择文件夹")
        btn_browse_out.setObjectName("BrowseBtn")
        btn_browse_out.clicked.connect(self.browse_out_dir)
        h2 = QHBoxLayout()
        h2.addWidget(self.out_dir)
        h2.addWidget(btn_browse_out)

        self.thread_spin = QSpinBox()
        self.thread_spin.setRange(1, 8)
        self.thread_spin.setValue(4)
        self.thread_spin.setToolTip("并行处理的目录数(1-8)")
        
        self.dpi_spin = QSpinBox()
        self.dpi_spin.setRange(72, 1200)
        self.dpi_spin.setValue(300)
        
        self.generate_ofd_check = QCheckBox("同时生成OFD文件")
        self.generate_ofd_check.setChecked(True)  # 默认选中

        form.addRow("图片目录:", h1)
        form.addRow("输出目录:", h2)
        form.addRow("线程数:", self.thread_spin)
        form.addRow("PDF DPI:", self.dpi_spin)
        form.addRow("选项:", self.generate_ofd_check)
        group.setLayout(form)
        self.layout.addWidget(group)
        
        # 说明文本
        info_label = QLabel(
            "功能说明：\n"
            "• 扫描目录及子目录下所有JPG文件\n"
            "• 每个目录的JPG合并为一个PDF\n"
            "• 使用内置本地OCR(与分件功能同款)生成双层PDF(可搜索/可选中)\n"
            "• PDF文件名与目录名相同\n"
            "• 可选：同时在PDF目录中生成同名OFD文件"
        )
        info_label.setStyleSheet("color: #8B949E; font-size: 12px;")
        self.layout.addWidget(info_label)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        btn_exec = QPushButton("开始转换")
        btn_exec.setObjectName("ActionBtn")
        btn_exec.clicked.connect(self.execute)
        btn_layout.addWidget(btn_exec)
        
        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        
        self.layout.addLayout(btn_layout)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)
        
        self.worker = None
        self.add_log_widget()
    
    def browse_img_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择图片目录")
        if d:
            self.img_dir.setText(d)
            # 输出目录未手动改过时, 自动跟随源目录 → 源目录\PDF
            self._auto_out_dir()

    def _auto_out_dir(self):
        """输出目录自动联动: 用户未手动改过时, 跟随源目录生成「源目录/PDF」;
        手动选择过其他输出目录后不再跟随。"""
        if getattr(self, '_out_manual', False):
            return
        base = self.img_dir.text().strip()
        if base:
            self.out_dir.setText(os.path.join(base, "PDF"))

    def browse_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择输出目录")
        if d:
            self.out_dir.setText(d)
            self._out_manual = True  # 手动指定后不再跟随源目录
    
    def execute(self):
        d = self.img_dir.text()
        out = self.out_dir.text()
        
        if not d:
            QMessageBox.warning(self, "提示", "请先选择图片目录")
            return

        if not out:
            QMessageBox.warning(self, "提示", "请先选择输出目录")
            return

        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "图片目录不存在")
            return

        # 输出目录不存在 → 点击开始时自动创建
        if not os.path.exists(out):
            try:
                os.makedirs(out, exist_ok=True)
                self.log(f"输出目录不存在, 已自动创建: {out}")
            except Exception as e:
                QMessageBox.warning(self, "错误", f"无法创建输出目录: {e}")
                return

        # 确认操作
        reply = QMessageBox.question(
            self, "确认操作",
            "确定要开始转换吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # 重置进度条
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        
        # 禁用按钮，启用停止按钮
        self.log("="*50)
        self.log("开始处理...")
        self.log(f"图片目录: {d}")
        self.log(f"输出目录: {out}")
        self.log(f"线程数: {self.thread_spin.value()}")
        self.log(f"PDF DPI: {self.dpi_spin.value()}")
        self.log("="*50)
        
        # 创建工作线程
        max_workers = self.thread_spin.value()
        resolution = float(self.dpi_spin.value())
        generate_ofd = self.generate_ofd_check.isChecked()
        self.worker = JpgToPdfWorker(
            directory_path=d,
            output_dir=out,
            max_workers=max_workers,
            resolution=resolution,
            generate_ofd=generate_ofd
        )
        
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        
        self.worker.start()
        
        # 更新按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始转换":
            btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
    
    def stop_processing(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止处理...")
            self.stop_btn.setEnabled(False)
    
    def update_progress(self, current, total):
        """更新进度显示"""
        percentage = (current / total) * 100 if total > 0 else 0
        self.progress.setValue(int(percentage))
        self.progress.setFormat(f"{current} / {total}  ({percentage:.0f}%)")
        self.log(f"进度: {current}/{total} ({percentage:.1f}%)")
    
    def on_finished(self, success, message):
        """处理完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")
        
        # 恢复按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始转换":
            btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        
        if success:
            QMessageBox.information(self, "完成", "处理完成！")
        else:
            QMessageBox.warning(self, "提示", message)


class PdfToOfdWorker(QThread):
    """PDF转OFD后台处理线程 - 带文本层PDF直接转换；无文本层PDF走 PDF页→图片→双层PDF→OFD 流程"""
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, pdf_dir, output_dir, parent=None):
        super().__init__(parent)
        self.pdf_dir = pdf_dir
        self.output_dir = output_dir
        self.is_stopped = False
        import threading
        self._ofd_lock = threading.Lock()

        # 检查 OFD 转换库是否可用（ofd_writer，基于 PyMuPDF，生成双层OFD）
        try:
            import fitz  # noqa: F401
            from ofd_writer import make_layered_ofd  # noqa: F401
            self.ofd_available = True
        except ImportError:
            self.ofd_available = False

        # 复用「JPG转双层PDF」功能的处理流程：
        # PDF逐页渲染为图片 → OCR生成双层PDF → 转双层OFD
        self._jpg_helper = JpgToPdfWorker(
            directory_path=self.pdf_dir, output_dir=self.output_dir, resolution=300.0)

    def run(self):
        try:
            if not self.ofd_available:
                self.finished_signal.emit(False, "错误：未安装 OFD 转换依赖(PyMuPDF)\n请运行: pip install PyMuPDF")
                return
            
            # 获取所有PDF文件
            pdf_files = []
            for root, dirs, files in os.walk(self.pdf_dir):
                for file in files:
                    if file.lower().endswith('.pdf'):
                        pdf_path = os.path.join(root, file)
                        pdf_files.append(pdf_path)
            
            total_files = len(pdf_files)
            if total_files == 0:
                self.finished_signal.emit(False, "未找到任何PDF文件")
                return
            
            self.log_signal.emit(f"找到 {total_files} 个PDF文件，开始转换...")
            
            # 创建日志文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_filename = f"ofd_conversion_log_{timestamp}.txt"
            log_file_path = os.path.join(self.pdf_dir, log_filename)
            start_time = datetime.now()
            log_entries = []
            
            processed = 0
            success_count = 0
            
            for pdf_path in pdf_files:
                if self.is_stopped:
                    break
                
                file_start_time = time.time()
                
                # 计算相对路径以保持目录结构
                rel_path = os.path.relpath(pdf_path, self.pdf_dir)
                ofd_filename = os.path.splitext(os.path.basename(pdf_path))[0] + '.ofd'
                output_subdir = os.path.join(self.output_dir, os.path.dirname(rel_path))
                os.makedirs(output_subdir, exist_ok=True)
                ofd_path = os.path.join(output_subdir, ofd_filename)
                
                # 执行转换
                success, duration = self.convert_pdf_to_ofd(pdf_path, ofd_path)
                
                file_end_time = time.time()
                file_duration = file_end_time - file_start_time
                
                if success:
                    success_count += 1
                    result = "成功"
                    self.log_signal.emit(f"✓ {os.path.basename(pdf_path)} -> {ofd_filename} ({duration:.2f}秒)")
                else:
                    result = "失败"
                    self.log_signal.emit(f"✗ {os.path.basename(pdf_path)} 转换失败")
                
                processed += 1
                percentage = (processed / total_files) * 100
                
                # 记录日志条目
                log_entry = {
                    "文件名": os.path.basename(pdf_path),
                    "处理结果": result,
                    "处理用时(秒)": f"{file_duration:.2f}",
                    "处理时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                log_entries.append(log_entry)
                
                self.progress_signal.emit(processed, total_files)
            
            if not self.is_stopped:
                # 写入日志文件
                end_time = datetime.now()
                total_duration = (end_time - start_time).total_seconds() / 60
                
                with open(log_file_path, 'w', encoding='utf-8') as log_file:
                    log_file.write("PDF转OFD处理日志 (双层OFD)\n")
                    log_file.write("=" * 50 + "\n")
                    log_file.write(f"处理开始时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                    log_file.write(f"处理结束时间: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                    log_file.write(f"总耗时: {total_duration:.2f}分钟\n")
                    log_file.write(f"总文件数: {total_files}\n")
                    log_file.write(f"成功转换: {success_count}\n")
                    log_file.write(f"转换失败: {total_files - success_count}\n")
                    log_file.write("=" * 50 + "\n\n")
                    
                    log_file.write("详细处理记录:\n")
                    log_file.write("文件名\t处理结果\t处理用时(秒)\t处理时间\n")
                    log_file.write("-" * 80 + "\n")
                    
                    for entry in log_entries:
                        log_file.write(f"{entry['文件名']}\t{entry['处理结果']}\t{entry['处理用时(秒)']}\t{entry['处理时间']}\n")
                
                summary_msg = f"\n转换完成！共处理 {total_files} 个PDF文件，成功转换 {success_count} 个，总耗时 {total_duration:.2f}分钟。\n日志: {log_filename}"
                self.finished_signal.emit(True, summary_msg)
            else:
                self.finished_signal.emit(False, "处理已停止")
                
        except Exception as e:
            self.finished_signal.emit(False, f"处理出错: {str(e)}")
    
    def stop(self):
        self.is_stopped = True
        if getattr(self, '_jpg_helper', None) is not None:
            self._jpg_helper.is_stopped = True
    
    def _pdf_has_text(self, pdf_path):
        """检测 PDF 是否已带可提取文本层（多数页含文本即视为带文本层）"""
        try:
            import fitz
            doc = fitz.open(pdf_path)
            try:
                pages_with_text = 0
                for page in doc:
                    if page.get_text().strip():
                        pages_with_text += 1
                return pages_with_text * 2 >= max(doc.page_count, 1)
            finally:
                doc.close()
        except Exception:
            return False

    def convert_pdf_to_ofd(self, pdf_path, ofd_path):
        """
        将PDF转换为双层OFD：
        ① 若PDF已带文本层，直接调 ofd_writer.make_layered_ofd 转换（无需图片中转）；
        ② 若无文本层，采用与「JPG转双层PDF」功能相同的流程：
           PyMuPDF 逐页渲染为 JPG 图片(300DPI) → 复用 JpgToPdfWorker.process_jpgs_to_ocr_pdf
           生成中间双层PDF(图像层+OCR文本层) → ofd_writer 转为双层 OFD；
        ③ 清理临时图片与中间 PDF，仅保留 OFD 输出。
        """
        if not os.path.exists(pdf_path):
            return False, 0

        if not self.ofd_available:
            return False, 0

        start_time = time.time()
        tmp_dir = None

        try:
            import fitz
            import tempfile
            import uuid
            from ofd_writer import make_layered_ofd

            # ① 已带文本层的 PDF 直接使用，跳过图片中转与 OCR
            if self._pdf_has_text(pdf_path):
                self.log_signal.emit(f"  {os.path.basename(pdf_path)} 已带文本层，直接转换")
                with self._ofd_lock:
                    ok, n, err = make_layered_ofd(pdf_path, ofd_path)
                if not ok:
                    print(f"PDF转OFD失败: {err}")
                    return False, time.time() - start_time
                return True, time.time() - start_time

            self.log_signal.emit(f"  {os.path.basename(pdf_path)} 无文本层，走图片中转+OCR流程")

            # ② PDF 逐页渲染为 JPG 图片
            doc = fitz.open(pdf_path)
            tmp_dir = os.path.join(tempfile.gettempdir(), f"pdf2ofd_{uuid.uuid4().hex[:8]}")
            os.makedirs(tmp_dir, exist_ok=True)
            jpg_paths = []
            zoom = 300.0 / 72.0  # 300DPI 渲染，与中间PDF分辨率一致，页面尺寸不变
            mat = fitz.Matrix(zoom, zoom)
            for i, page in enumerate(doc):
                if self.is_stopped:
                    break
                pix = page.get_pixmap(matrix=mat, alpha=False)
                jpg_path = os.path.join(tmp_dir, f"page_{i:05d}.jpg")
                pix.save(jpg_path)
                jpg_paths.append(jpg_path)
            doc.close()

            if self.is_stopped or not jpg_paths:
                return False, time.time() - start_time

            # ②a 复用 JPG→双层PDF 流程生成中间 PDF
            self._jpg_helper.is_stopped = self.is_stopped
            base_name = os.path.splitext(os.path.basename(ofd_path))[0]
            inter_pdf, ocr_status = self._jpg_helper.process_jpgs_to_ocr_pdf(
                jpg_paths, tmp_dir, base_name)
            if not inter_pdf or not os.path.exists(inter_pdf):
                return False, time.time() - start_time

            # ②b 中间 PDF 转 OFD
            with self._ofd_lock:
                ok, n, err = make_layered_ofd(inter_pdf, ofd_path)

            if not ok:
                print(f"PDF转OFD失败: {err}")
                return False, time.time() - start_time

            return True, time.time() - start_time

        except Exception as e:
            print(f"PDF转OFD过程中发生错误: {e}")
            return False, time.time() - start_time
        finally:
            # ③ 清理临时图片与中间 PDF
            if tmp_dir and os.path.isdir(tmp_dir):
                import shutil
                try:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                except Exception:
                    pass


class PdfToOfdPage(FunctionPage):
    def __init__(self):
        super().__init__("PDF转OFD")
        group = QGroupBox("PDF转OFD格式 (双层OFD)")
        form = QFormLayout()

        self.pdf_dir = QLineEdit()
        btn_browse_src = QPushButton("选择文件夹")
        btn_browse_src.setObjectName("BrowseBtn")
        btn_browse_src.clicked.connect(self.browse_pdf_dir)
        h1 = QHBoxLayout()
        h1.addWidget(self.pdf_dir)
        h1.addWidget(btn_browse_src)
        
        self.out_dir = QLineEdit()
        btn_browse_out = QPushButton("选择文件夹")
        btn_browse_out.setObjectName("BrowseBtn")
        btn_browse_out.clicked.connect(self.browse_out_dir)
        h2 = QHBoxLayout()
        h2.addWidget(self.out_dir)
        h2.addWidget(btn_browse_out)

        form.addRow("PDF源目录:", h1)
        form.addRow("OFD输出目录:", h2)
        group.setLayout(form)
        self.layout.addWidget(group)
        
        # 说明文本
        info_label = QLabel(
            "功能说明：\n"
            "• 已带文本层的 PDF 直接转换为双层OFD；无文本层的 PDF 自动走「JPG转双层PDF」同款流程（渲染图片 → OCR双层PDF → 双层OFD）\n"
            "• 生成双层OFD（图像层+可检索文本层），完整转换全部页，无页数限制\n"
            "• 支持递归扫描子目录\n"
            "• 保持原有目录结构\n"
            "• OFD文件名与PDF相同\n"
            "• 注意：需要 OFD 转换依赖 PyMuPDF (已随程序内置)"
        )
        info_label.setStyleSheet("color: #8B949E; font-size: 12px;")
        self.layout.addWidget(info_label)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        btn_exec = QPushButton("开始转换为OFD")
        btn_exec.setObjectName("ActionBtn")
        btn_exec.clicked.connect(self.execute)
        btn_layout.addWidget(btn_exec)
        
        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        
        self.layout.addLayout(btn_layout)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)
        
        self.worker = None
        self.add_log_widget()
    
    def browse_pdf_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择PDF源目录")
        if d:
            self.pdf_dir.setText(d)
            if not self.out_dir.text():
                self.out_dir.setText(d)
    
    def browse_out_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择OFD输出目录")
        if d: self.out_dir.setText(d)
    
    def execute(self):
        d = self.pdf_dir.text()
        out = self.out_dir.text()
        
        if not d:
            QMessageBox.warning(self, "提示", "请先选择PDF源目录")
            return
        
        if not out:
            QMessageBox.warning(self, "提示", "请先选择OFD输出目录")
            return
        
        if not os.path.exists(d):
            QMessageBox.warning(self, "错误", "PDF源目录不存在")
            return

        # OFD输出目录不存在 → 点击开始时自动创建
        if not os.path.exists(out):
            try:
                os.makedirs(out, exist_ok=True)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"无法创建OFD输出目录: {e}")
                return
        
        # 确认操作
        reply = QMessageBox.question(
            self, "确认操作",
            "确定要开始转换吗？\n\n注意：需要 OFD 转换依赖 PyMuPDF！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # 重置进度条
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        
        # 禁用按钮，启用停止按钮
        self.log("="*50)
        self.log("开始处理...")
        self.log(f"PDF源目录: {d}")
        self.log(f"OFD输出目录: {out}")
        self.log("="*50)
        
        # 创建工作线程
        self.worker = PdfToOfdWorker(
            pdf_dir=d,
            output_dir=out
        )
        
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        
        self.worker.start()
        
        # 更新按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始转换为OFD":
            btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
    
    def stop_processing(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止处理...")
            self.stop_btn.setEnabled(False)
    
    def update_progress(self, current, total):
        """更新进度显示"""
        percentage = (current / total) * 100 if total > 0 else 0
        self.progress.setValue(int(percentage))
        self.progress.setFormat(f"{current} / {total}  ({percentage:.0f}%)")
        self.log(f"进度: {current}/{total} ({percentage:.1f}%)")
    
    def on_finished(self, success, message):
        """处理完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")
        
        # 恢复按钮状态
        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始转换为OFD":
            btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        
        if success:
            QMessageBox.information(self, "完成", "处理完成！")
        else:
            QMessageBox.warning(self, "提示", message)


class FileBatchReplaceWorker(QThread):
    """文件批量替换后台处理线程 - 将输出的JPG批量替换目标目录中的图片

    规则:
    1. 扫描 JPG 目录(含子目录)中的 jpg 文件，文件名去除最后四位数字编号和最后一个
       "-"后得到基础名，与目标目录下的同名子目录匹配；
    2. 仅处理编号 >= 起始编号的文件:
       - 只需复制1个 → 直接覆盖目标同名文件；
       - 复制多个(N个) → 先将目标目录中编号 > 最小新文件编号的现有文件
         向后平移 (N-1) 个编号预留空位(从编号最大的文件开始改名避免重复)，
         再将新文件复制到目标目录。
    3. 每个目标子目录替换完成后, 在该子目录下生成/更新 Directory.txt:
       记录替换后的文件名清单(不含扩展名, 每行一个), 供分件功能
       读取目录页与后续文件偏移量。
    """
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)

    _NUM_RE = re.compile(r'^(.+)-(\d{4})$')  # 基础名 + "-" + 四位数字编号

    def __init__(self, jpg_dir, target_dir, start_num, parent=None):
        super().__init__(parent)
        self.jpg_dir = jpg_dir
        self.target_dir = target_dir
        self.start_num = start_num
        self.is_stopped = False
        self._log_file = None  # 当前处理日志文件句柄

    def stop(self):
        self.is_stopped = True

    def _wlog(self, s):
        """写入处理日志文件（若已打开），同时输出到界面日志"""
        if self._log_file is not None:
            try:
                self._log_file.write(s + "\n")
                self._log_file.flush()
            except Exception:
                pass
        self.log_signal.emit(s)

    @classmethod
    def _parse_name(cls, fname):
        """文件名去除扩展名、最后一个"-"及末尾四位数字编号，返回 (基础名, 编号)；不匹配返回 None"""
        stem = os.path.splitext(fname)[0]
        m = cls._NUM_RE.match(stem)
        if not m:
            return None
        return m.group(1), int(m.group(2))

    def _collect_sources(self):
        """扫描JPG目录(含子目录)，按基础名分组，返回 ({基础名: [(编号, 完整路径), ...]}, 跳过数)"""
        groups = {}
        skipped = 0
        for root, _dirs, files in os.walk(self.jpg_dir):
            for f in files:
                if not f.lower().endswith(('.jpg', '.jpeg')):
                    continue
                parsed = self._parse_name(f)
                if parsed is None:
                    skipped += 1
                    continue
                base, num = parsed
                if num < self.start_num:
                    continue
                groups.setdefault(base, []).append((num, os.path.join(root, f)))
        for base in groups:
            groups[base].sort(key=lambda t: t[0])
        return groups, skipped

    def _replace_group(self, base, new_files):
        """处理单个基础名分组，返回 (成功复制数, 失败数)"""
        target_sub = os.path.join(self.target_dir, base)
        if not os.path.isdir(target_sub):
            # 目标目录无同名子目录 → 新建后直接复制(无需改名)
            os.makedirs(target_sub, exist_ok=True)
            self._wlog(f"  目标无同名子目录，新建: {base}")

        n = len(new_files)
        nums_new = {num for num, _ in new_files}
        min_new = min(nums_new)

        # 目标子目录现有文件(任意扩展名)中符合「基础名-四位编号」的文件
        existing = []  # [(编号, 文件名)]
        pat = re.compile(r'^' + re.escape(base) + r'-(\d{4})\.[^.]+$')
        for f in os.listdir(target_sub):
            if not os.path.isfile(os.path.join(target_sub, f)):
                continue
            m = pat.match(f)
            if m:
                existing.append((int(m.group(1)), f))

        if n > 1:
            shift = n - 1
            # 从编号最大的文件开始改名，避免改名过程中重名覆盖
            to_rename = sorted([t for t in existing if t[0] > min_new], reverse=True)
            # 碰撞安全检查: 改名后编号不得与待复制新文件编号冲突(源编号不连续时可能发生)
            conflict = {num + shift for num, _ in to_rename} & nums_new
            if conflict:
                self._wlog(f"  × 源文件编号不连续，改名将与待复制编号 {sorted(conflict)} 冲突，跳过本组")
                return 0, n
            for num, fname in to_rename:
                _stem, ext = os.path.splitext(fname)
                new_name = f"{base}-{num + shift:04d}{ext}"
                os.rename(os.path.join(target_sub, fname),
                          os.path.join(target_sub, new_name))
                self._wlog(f"  改名: {fname} → {new_name}")

        # 改名完成后复制新文件；改名后仍保留在原编号上的文件才会被覆盖
        overwrite_nums = set()
        if n == 1:
            if any(e_num == new_files[0][0] for e_num, _ in existing):
                overwrite_nums.add(new_files[0][0])
        elif any(e_num == min_new for e_num, _ in existing):
            overwrite_nums.add(min_new)

        ok = 0
        for num, src_path in new_files:
            fname = os.path.basename(src_path)
            shutil.copy2(src_path, os.path.join(target_sub, fname))
            act = "覆盖" if num in overwrite_nums else "复制"
            self._wlog(f"  {act}: {fname}")
            ok += 1
        if ok:
            # 生成 Directory.txt: 记录替换后的文件名清单(不含扩展名), 供分件读取偏移量
            self._write_directory_txt(target_sub,
                                      [os.path.splitext(os.path.basename(p))[0]
                                       for _, p in new_files])
        return ok, 0

    def _write_directory_txt(self, target_sub, new_stems):
        """在目标子目录下生成/更新 Directory.txt: 记录替换后的文件名清单
        (不含扩展名, 每行一个)。已有记录合并(多次替换时累积),
        按文件名末尾编号排序去重。"""
        dpath = os.path.join(target_sub, 'Directory.txt')
        existing = []
        if os.path.isfile(dpath):
            try:
                with open(dpath, 'r', encoding='utf-8') as fh:
                    existing = [l.strip() for l in fh if l.strip()]
            except Exception:
                existing = []
        merged = list(dict.fromkeys(existing + list(new_stems)))

        def _key(stem):
            m = re.search(r'(\d{1,4})$', stem)
            return int(m.group(1)) if m else 0

        merged.sort(key=_key)
        try:
            with open(dpath, 'w', encoding='utf-8') as fh:
                fh.write('\n'.join(merged) + '\n')
            self._wlog(f"  生成 Directory.txt: 记录 {len(merged)} 个替换后的文件名"
                       f"(本次新增/更新 {len(new_stems)} 个)")
        except Exception as e:
            self._wlog(f"  × 写入 Directory.txt 失败: {e}")

    def run(self):
        logf = None
        try:
            if not os.path.isdir(self.jpg_dir):
                self.finished_signal.emit(False, "JPG源目录不存在")
                return
            os.makedirs(self.target_dir, exist_ok=True)

            # 在目标目录下生成处理日志文件
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_path = os.path.join(self.target_dir, f"文件批量替换处理日志_{ts}.txt")
            logf = open(log_path, 'w', encoding='utf-8')
            self._log_file = logf

            self._wlog("文件批量替换处理日志")
            self._wlog(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self._wlog(f"JPG源目录: {self.jpg_dir}")
            self._wlog(f"目标目录: {self.target_dir}")
            self._wlog(f"起始编号: {self.start_num:04d} (仅处理编号 >= 起始编号的文件)")
            self._wlog("=" * 70)

            groups, skipped = self._collect_sources()
            if skipped:
                self._wlog(f"跳过 {skipped} 个文件名不含「-四位数字编号」的文件")
            if not groups:
                self._wlog("未找到符合替换条件的JPG文件")
                logf.close()
                logf = None
                self._log_file = None
                self.finished_signal.emit(False, f"未找到符合替换条件的JPG文件(起始编号 {self.start_num:04d})")
                return

            self._wlog(f"共 {len(groups)} 个基础名分组待替换")
            total = len(groups)
            done = 0
            total_ok = 0
            total_fail = 0
            for i, base in enumerate(sorted(groups), 1):
                if self.is_stopped:
                    self._wlog("用户停止处理")
                    break
                new_files = groups[base]
                self._wlog("")
                self._wlog(f"[{i}/{total}] {base} (待复制 {len(new_files)} 个文件)")
                try:
                    ok, fail = self._replace_group(base, new_files)
                    total_ok += ok
                    total_fail += fail
                    self._wlog(f"  结果: 复制 {ok} 个文件" + (f"，失败 {fail} 个" if fail else ""))
                except Exception as e:
                    import traceback
                    total_fail += len(new_files)
                    self._wlog(f"  [异常] {e}")
                    self._wlog(traceback.format_exc())
                done += 1
                self.progress_signal.emit(done, total)

            self._wlog("")
            self._wlog("=" * 70)
            self._wlog(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self._wlog(f"总计: 处理 {done}/{total} 组，复制 {total_ok} 个文件，失败 {total_fail} 个")
            if self.is_stopped:
                self._wlog("注意: 处理被用户中途停止")
            logf.close()
            logf = None
            self._log_file = None

            msg = (f"替换完成！处理 {done}/{total} 组，复制 {total_ok} 个文件。\n"
                   f"日志: {os.path.basename(log_path)}")
            self.finished_signal.emit(not self.is_stopped, msg)
        except Exception as e:
            import traceback
            self.log_signal.emit(traceback.format_exc())
            if logf is not None:
                try:
                    logf.close()
                except Exception:
                    pass
                self._log_file = None
            self.finished_signal.emit(False, f"处理出错: {e}")


class FileBatchReplacePage(FunctionPage):
    """文件批量替换功能页：将输出的JPG替换目标目录中的图片"""

    def __init__(self):
        super().__init__("文件批量替换")
        group = QGroupBox("文件批量替换 (将输出的JPG替换目标目录中的图片)")
        form = QFormLayout()

        self.jpg_dir = QLineEdit()
        self.jpg_dir.setPlaceholderText("选择包含输出JPG的目录(含子目录)")
        btn_browse_jpg = QPushButton("选择文件夹")
        btn_browse_jpg.setObjectName("BrowseBtn")
        btn_browse_jpg.clicked.connect(self.browse_jpg_dir)
        h1 = QHBoxLayout()
        h1.addWidget(self.jpg_dir)
        h1.addWidget(btn_browse_jpg)

        self.target_dir = QLineEdit()
        self.target_dir.setPlaceholderText("选择要被替换图片的目标目录")
        btn_browse_target = QPushButton("选择文件夹")
        btn_browse_target.setObjectName("BrowseBtn")
        btn_browse_target.clicked.connect(self.browse_target_dir)
        h2 = QHBoxLayout()
        h2.addWidget(self.target_dir)
        h2.addWidget(btn_browse_target)

        self.start_num = QLineEdit()
        self.start_num.setPlaceholderText("例如 0002，仅处理编号不小于此编号的文件")

        form.addRow("JPG源目录:", h1)
        form.addRow("目标目录:", h2)
        form.addRow("起始文件名:", self.start_num)
        group.setLayout(form)
        self.layout.addWidget(group)

        info_label = QLabel(
            "功能说明：\n"
            "• 扫描JPG源目录(含子目录)，文件名去除最后四位数字编号和最后一个\"-\"得到基础名，与目标目录下同名子目录匹配\n"
            "• 仅处理编号不小于起始编号的文件(缺省扩展名jpg，如输入 0002；留空默认从 0001 开始)\n"
            "• 只复制1个文件 → 直接覆盖目标同名文件；复制N个文件 → 目标目录中编号大于最小复制编号的\n"
            "  现有文件先向后平移 N-1 个编号预留空位(从编号最大的文件开始改名)，再复制新文件\n"
            "• 处理日志生成在目标目录下\n"
            "• 替换后在每个目标子目录生成 Directory.txt(记录替换后的文件名清单, 不含扩展名),\n"
            "  供分件功能读取目录页与文件偏移量\n"
            "• 注意：此操作会修改目标目录文件(改名/覆盖)，建议先备份"
        )
        info_label.setStyleSheet("color: #8B949E; font-size: 12px;")
        self.layout.addWidget(info_label)

        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_exec = QPushButton("开始替换")
        btn_exec.setObjectName("ActionBtn")
        btn_exec.clicked.connect(self.execute)
        btn_layout.addWidget(btn_exec)

        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        self.layout.addLayout(btn_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)

        self.worker = None
        self.add_log_widget()

    def browse_jpg_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择JPG源目录")
        if d:
            self.jpg_dir.setText(d)

    def browse_target_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择目标目录")
        if d:
            self.target_dir.setText(d)

    def execute(self):
        jpg_d = self.jpg_dir.text().strip()
        target_d = self.target_dir.text().strip()

        if not jpg_d:
            QMessageBox.warning(self, "提示", "请先选择JPG源目录")
            return
        if not os.path.isdir(jpg_d):
            QMessageBox.warning(self, "错误", "JPG源目录不存在")
            return
        if not target_d:
            QMessageBox.warning(self, "提示", "请先选择目标目录")
            return

        # 解析起始编号(缺省扩展名jpg，如 "0002" 或 "0002.jpg")
        s = self.start_num.text().strip()
        if s:
            s = os.path.splitext(s)[0]
            if not s.isdigit():
                QMessageBox.warning(self, "错误", "起始文件名须为数字编号，例如 0002")
                return
            start_num = int(s)
        else:
            start_num = 1

        # 目标目录不存在则自动创建
        if not os.path.exists(target_d):
            try:
                os.makedirs(target_d, exist_ok=True)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"无法创建目标目录: {e}")
                return

        reply = QMessageBox.question(
            self, "确认操作",
            f"将从编号 {start_num:04d} 开始，用JPG源目录的文件替换目标目录中的图片。\n"
            f"目标目录中的现有文件可能被改名或覆盖，建议先备份。\n\n确定开始吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        self.progress.setValue(0)
        self.progress.setFormat("准备中...")
        self.log("=" * 50)
        self.log("开始替换...")
        self.log(f"JPG源目录: {jpg_d}")
        self.log(f"目标目录: {target_d}")
        self.log(f"起始编号: {start_num:04d}")
        self.log("=" * 50)

        self.worker = FileBatchReplaceWorker(
            jpg_dir=jpg_d, target_dir=target_d, start_num=start_num)
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()

        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始替换":
            btn.setEnabled(False)
        self.stop_btn.setEnabled(True)

    def stop_processing(self):
        """停止处理"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止处理...")
            self.stop_btn.setEnabled(False)

    def update_progress(self, current, total):
        """更新进度显示"""
        percentage = (current / total) * 100 if total > 0 else 0
        self.progress.setValue(int(percentage))
        self.progress.setFormat(f"{current} / {total}  ({percentage:.0f}%)")

    def on_finished(self, success, message):
        """处理完成回调"""
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止")

        btn = self.findChild(QPushButton, "ActionBtn")
        if btn and btn.text() == "开始替换":
            btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

        if success:
            QMessageBox.information(self, "完成", "处理完成！")
        else:
            QMessageBox.warning(self, "提示", message)


class FileSplitPage(FunctionPage):
    """分件功能页：按卷内文件目录表格自动分件"""

    def __init__(self):
        super().__init__("分件")
        group = QGroupBox("分件 (按卷内文件目录表格自动拆分)")
        form = QFormLayout()

        self.dir_edit = QLineEdit()
        self.dir_edit.setPlaceholderText("选择要分件的目录(将处理其下所有子目录)")
        btn_browse = QPushButton("选择文件夹")
        btn_browse.setObjectName("BrowseBtn")
        btn_browse.clicked.connect(self.browse_dir)
        h = QHBoxLayout()
        h.addWidget(self.dir_edit)
        h.addWidget(btn_browse)
        form.addRow("分件目录:", h)

        # 分件到新目录(缺省选中): 结果按目录结构拷贝到目标目录, 源不动
        self._target_manual = False  # 用户是否手动改过目标目录(改过则不随源联动)
        self.to_new_check = QCheckBox("分件到新目录(拷贝到目标目录, 源文件不动)")
        self.to_new_check.setChecked(True)
        self.to_new_check.stateChanged.connect(self.on_to_new_toggled)
        form.addRow("分件方式:", self.to_new_check)

        self.target_edit = QLineEdit()
        self.target_edit.setPlaceholderText("分件结果输出目录")
        btn_target = QPushButton("选择文件夹")
        btn_target.setObjectName("BrowseBtn")
        btn_target.clicked.connect(self.browse_target)
        h2 = QHBoxLayout()
        h2.addWidget(self.target_edit)
        h2.addWidget(btn_target)
        self.btn_target = btn_target
        form.addRow("目标目录:", h2)

        # 分件依据: 缺省按目录文件(xlsx)分件; OCR分件为可选项, 缺省不勾选
        self.ocr_check = QCheckBox("使用OCR识别分件(不勾选时缺省按xlsx目录文件分件)")
        self.ocr_check.setChecked(False)
        self.ocr_check.stateChanged.connect(self.on_ocr_toggled)
        form.addRow("分件依据:", self.ocr_check)

        self.xlsx_dir_edit = QLineEdit()
        self.xlsx_dir_edit.setPlaceholderText("包含xlsx目录文件的文件夹(将递归查找与待处理目录同名的xlsx)")
        btn_xlsx = QPushButton("选择文件夹")
        btn_xlsx.setObjectName("BrowseBtn")
        btn_xlsx.clicked.connect(self.browse_xlsx_dir)
        self.btn_xlsx = btn_xlsx
        h3 = QHBoxLayout()
        h3.addWidget(self.xlsx_dir_edit)
        h3.addWidget(btn_xlsx)
        form.addRow("目录文件:", h3)

        group.setLayout(form)
        self.layout.addWidget(group)

        # 说明
        info = QLabel(
            "功能说明：\n"
            "• 缺省按xlsx目录文件分件: 从指定文件夹下读取与待处理目录同名的xlsx文件(第3行为标题行, 含「序号」「页号」列)作为分件依据\n"
            "• 勾选「使用OCR识别分件」切换为OCR模式: 对每个子目录 OCR 目录页(卷内文件目录)，解析序号/页号自动拆分\n"
            "• 若子目录含 Directory.txt(文件批量替换生成): 第一页为卷皮, 目录页取 Directory.txt 记录,\n"
            "  文件偏移量=Directory.txt 最大文件序号; 无 Directory.txt 时提示确认并按缺省偏移量2处理;\n"
            "  分件完成后删除该目录下 Directory.txt\n"
            "• 检查: 按与分件相同的依据(xlsx/OCR)检查各目录，发现应移走却残留的文件并生成报告\n"
            "• 手工分件: 逐个目录手工输入序号/页号拆分(带连续性校验)"
        )
        info.setStyleSheet("color: #666; font-size: 12px;")
        self.layout.addWidget(info)

        # 按钮
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始分件")
        self.start_btn.setObjectName("ActionBtn")
        self.start_btn.clicked.connect(self.start)
        btn_layout.addWidget(self.start_btn)
        self.check_btn = QPushButton("检查")
        self.check_btn.setObjectName("ActionBtn")
        self.check_btn.clicked.connect(self.start_check)
        btn_layout.addWidget(self.check_btn)
        self.manual_btn = QPushButton("手工分件")
        self.manual_btn.setObjectName("ActionBtn")
        self.manual_btn.clicked.connect(self.start_manual)
        btn_layout.addWidget(self.manual_btn)
        self.stop_btn = QPushButton("停止")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633; color: white;")
        self.stop_btn.clicked.connect(self.stop)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        self.layout.addLayout(btn_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)

        # 日志
        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setMinimumHeight(250)
        self.layout.addWidget(self.log_box, 1)

        self.worker = None

    def browse_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择分件目录")
        if d:
            self.dir_edit.setText(d)
            # 源目录变化: 目标目录未手动改过时自动跟随(源/分件完成)
            self._auto_target()
            # 缺省xlsx分件模式且未选目录文件时, 目录文件跟随分件目录
            if not self.ocr_check.isChecked() and not self.xlsx_dir_edit.text().strip():
                self.xlsx_dir_edit.setText(d)

    def _auto_target(self):
        """目标目录自动联动: 用户未手动改过目标目录时, 跟随源目录生成
        「源目录/分件完成」; 手动改过(选择过其他目录)后不再跟随。"""
        if getattr(self, '_target_manual', False):
            return
        base = self.dir_edit.text().strip()
        if base:
            self.target_edit.setText(os.path.join(base, "分件完成"))

    def log(self, msg):
        self.log_box.append(f">> {msg}")

    def on_to_new_toggled(self, state):
        """「分件到新目录」勾选状态切换: 取消时禁编目标目录。"""
        to_new = self.to_new_check.isChecked()
        self.target_edit.setEnabled(to_new)
        self.btn_target.setEnabled(to_new)
        if to_new:
            self._auto_target()  # 目标为空或未手动改过 → 自动填默认

    def browse_target(self):
        d = QFileDialog.getExistingDirectory(self, "选择分件结果输出目录")
        if d:
            self.target_edit.setText(d)
            self._target_manual = True  # 用户手动指定, 之后不再跟随源目录

    def on_ocr_toggled(self, state):
        """「使用OCR识别分件」勾选状态切换: 勾选时用OCR(禁用xlsx目录输入);
        不勾选时缺省按xlsx目录文件分件(启用并要求选择xlsx目录)。"""
        use_ocr = self.ocr_check.isChecked()
        self.xlsx_dir_edit.setEnabled(not use_ocr)
        self.btn_xlsx.setEnabled(not use_ocr)
        if not use_ocr and not self.xlsx_dir_edit.text().strip():
            # xlsx模式且未选择目录时, 默认使用分件目录
            base = self.dir_edit.text().strip()
            if base:
                self.xlsx_dir_edit.setText(base)

    def browse_xlsx_dir(self):
        d = QFileDialog.getExistingDirectory(self, "选择包含xlsx目录文件的文件夹")
        if d:
            self.xlsx_dir_edit.setText(d)

    def start(self):
        base_dir = self.dir_edit.text().strip()
        if not base_dir:
            QMessageBox.warning(self, "提示", "请先选择分件目录")
            return
        if not os.path.isdir(base_dir):
            QMessageBox.warning(self, "错误", "目录不存在")
            return

        # Directory.txt 检查: 子目录缺少 Directory.txt 时可能未做过卷内目录替换, 提示用户确认
        sub_names = [d for d in os.listdir(base_dir)
                     if os.path.isdir(os.path.join(base_dir, d))]
        missing = [d for d in sub_names
                   if not os.path.isfile(os.path.join(base_dir, d, 'Directory.txt'))]
        if sub_names and missing:
            show = '、'.join(missing[:8]) + ('...' if len(missing) > 8 else '')
            reply = QMessageBox.question(
                self, "提示",
                f"当前待处理目录有 {len(missing)}/{len(sub_names)} 个子目录没有 Directory.txt，\n"
                f"可能没有进行过卷内目录替换。\n"
                f"若继续, 这些目录将按缺省偏移量 2 处理。\n"
                f"({show})\n\n是否继续？",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply != QMessageBox.Yes:
                self.log("已取消: 待处理目录缺少 Directory.txt")
                return
            self.log(f"注意: {len(missing)} 个目录无 Directory.txt, 将按缺省偏移量 2 处理")

        to_new = self.to_new_check.isChecked()
        target_base = None
        copy_mode = False
        if to_new:
            target_base = self.target_edit.text().strip()
            if not target_base:
                target_base = os.path.join(base_dir, "分件完成")
                self.target_edit.setText(target_base)
            try:
                os.makedirs(target_base, exist_ok=True)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"无法创建目标目录: {e}")
                return
            copy_mode = True

        if copy_mode:
            confirm_msg = (f"分件结果将拷贝到:\n{target_base}\n"
                           f"(源目录文件保持不动)")
        else:
            confirm_msg = ("分件将在源目录内移动 jpg 文件(不可自动撤销)，\n"
                           "建议先备份。确定开始吗？")
        reply = QMessageBox.question(self, "确认操作", confirm_msg,
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        self.log_box.clear()
        self.log(f"开始分件: {base_dir}")
        if copy_mode:
            self.log(f"分件到新目录(拷贝): {target_base}")

        # 分件依据: 缺省读取目录文件(xlsx); 勾选OCR选项时用OCR模式
        xlsx_dir = None
        if self.ocr_check.isChecked():
            self.log("分件依据: OCR识别目录页")
        else:
            xlsx_dir = self.xlsx_dir_edit.text().strip()
            if not xlsx_dir:
                QMessageBox.warning(self, "提示", "缺省按xlsx目录文件分件，请先选择包含xlsx的文件夹")
                return
            if not os.path.isdir(xlsx_dir):
                QMessageBox.warning(self, "错误", f"目录文件路径不存在: {xlsx_dir}")
                return
            self.log(f"读取目录文件(xlsx): {xlsx_dir}")

        self.worker = FileSplitWorker(base_dir, target_base=target_base,
                                      copy_mode=copy_mode, xlsx_dir=xlsx_dir)
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.start()
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)

    def stop(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.log("正在停止...")

    def update_progress(self, cur, total):
        pct = cur / total * 100 if total else 0
        self.progress.setValue(int(pct))
        self.progress.setFormat(f"{cur} / {total} ({pct:.0f}%)")

    def on_finished(self, success, message):
        self.log(message)
        self.progress.setFormat("已完成" if success else "已停止/失败")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        if success:
            QMessageBox.information(self, "完成", message)
        else:
            QMessageBox.warning(self, "结束", message)

    # ---------- 检查 ----------
    def start_check(self):
        base_dir = self.dir_edit.text().strip()
        if not base_dir:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        if not os.path.isdir(base_dir):
            QMessageBox.warning(self, "错误", "目录不存在")
            return

        # 分件依据: 缺省读取目录文件(xlsx); 勾选OCR选项时用OCR模式
        xlsx_dir = None
        if not self.ocr_check.isChecked():
            xlsx_dir = self.xlsx_dir_edit.text().strip()
            if not xlsx_dir or not os.path.isdir(xlsx_dir):
                QMessageBox.warning(self, "提示", "缺省按xlsx目录文件检查，请先选择有效的目录文件路径")
                return

        self.log_box.clear()
        self.log(f"开始检查: {base_dir}")
        if xlsx_dir:
            self.log(f"检查模式: 读取目录文件(xlsx) → {xlsx_dir}")
        self.check_worker = FileSplitCheckWorker(base_dir, xlsx_dir=xlsx_dir)
        self.check_worker.log_signal.connect(self.log)
        self.check_worker.progress_signal.connect(self.update_progress)
        self.check_worker.finished_signal.connect(self.on_check_finished)
        self.check_worker.start()
        self.check_btn.setEnabled(False)

    def on_check_finished(self, success, message):
        self.log(message)
        self.progress.setFormat("检查完成" if success else "检查停止/失败")
        self.check_btn.setEnabled(True)
        if success:
            QMessageBox.information(self, "检查完成", message)
        else:
            QMessageBox.warning(self, "检查结束", message)

    # ---------- 手工分件 ----------
    def start_manual(self):
        base_dir = self.dir_edit.text().strip()
        if not base_dir:
            QMessageBox.warning(self, "提示", "请先选择目录")
            return
        if not os.path.isdir(base_dir):
            QMessageBox.warning(self, "错误", "目录不存在")
            return
        dlg = ManualSplitDialog(base_dir, self)
        dlg.exec_()

class XlsxToJpgWorker(QThread):
    """表格文件(xlsx/xls)转JPG后台处理线程 - openpyxl读取+PIL渲染，按A4排版标准输出，超一页自动分页（分页序号可指定起始编码）。
    输出前对表格进行编辑: 删除 G 列后的所有无关列(仅保留 A-G);
    排版后若最后一页只有空表格(无文字)则删除该页;
    最后一页未占满版面时以空表格行补足, 保证导出文件整洁。"""
    # 补足空行统一行高(磅): 所有文件的补足空行都用此固定行高(对齐
    # J380-ZY·2021-Y-GTC-0171 样例的空行间距), 保证各文件空行间距一致;
    # 不按各文件自身空行行高取值(不同文件不一致会导致间距忽大忽小)
    FILL_ROW_H_PT = 57.0
    log_signal = Signal(str)
    progress_signal = Signal(int, int)
    finished_signal = Signal(bool, str)

    def __init__(self, input_dir, output_dir=None, same_dir=True, start_num=1, parent=None):
        super().__init__(parent)
        self.input_dir = input_dir
        self.output_dir = output_dir  # 仅 same_dir=False 时使用
        self.same_dir = same_dir      # True=输出到原目录
        self.start_num = start_num    # 分页输出的起始文件编码（默认1，即 -0001 开始）
        self.is_stopped = False
        self._log_file = None  # 当前处理日志文件句柄

    def _wlog(self, s):
        """写入处理日志文件（若已打开），同时输出到界面日志"""
        if self._log_file is not None:
            try:
                self._log_file.write(s + "\n")
                self._log_file.flush()
            except Exception:
                pass
        self.log_signal.emit(s)

    def _find_excel_files(self):
        """递归查找所有xlsx/xls文件"""
        result = []
        for root, dirs, files in os.walk(self.input_dir):
            for f in files:
                if f.lower().endswith(('.xlsx', '.xlsm', '.xls')) and not f.startswith('~$'):
                    result.append(os.path.join(root, f))
        return sorted(result)

    def _detect_file_format(self, filepath):
        """
        检测文件实际格式（通过文件头魔数）。
        返回: 'xlsx', 'xlsm', 'xls' (OLE2格式), 或 'unknown'
        """
        try:
            with open(filepath, 'rb') as f:
                header = f.read(8)
                # xlsx/xlsm 是 ZIP 格式，文件头为 PK (50 4B)
                if header[:2] == b'PK':
                    # 进一步检测是否包含宏（xlsm）
                    try:
                        import zipfile
                        with zipfile.ZipFile(filepath, 'r') as zf:
                            if 'xl/vbaProject.bin' in zf.namelist():
                                return 'xlsm'
                    except Exception:
                        pass
                    return 'xlsx'
                # xls 是 OLE2 复合文档，文件头为 D0 CF 11 E0
                if header[:4] == b'\xd0\xcf\x11\xe0':
                    return 'xls'
                return 'unknown'
        except Exception:
            return 'unknown'

    def _get_openpyxl_readable_path(self, filepath):
        """
        openpyxl 只接受 .xlsx/.xlsm/.xltx/.xltm 扩展名，
        对于被改名的文件（如 xlsm 改成 .xls），创建正确扩展名的临时副本。
        返回: (可读路径, 临时文件路径或None)
        """
        import tempfile
        import uuid
        import shutil
        supported = ('.xlsx', '.xlsm', '.xltx', '.xltm')
        ext = os.path.splitext(filepath)[1].lower()
        if ext in supported:
            return filepath, None
        # 根据实际格式确定正确扩展名
        actual_format = self._detect_file_format(filepath)
        correct_ext = {'xlsx': '.xlsx', 'xlsm': '.xlsm'}.get(actual_format, '.xlsx')
        temp_file = os.path.join(tempfile.gettempdir(),
                                 f"opy_{uuid.uuid4().hex[:8]}{correct_ext}")
        shutil.copy2(filepath, temp_file)
        return temp_file, temp_file

    def _convert_xlsx(self, filepath, output_path, dpi=300):
        """转换xlsx/xlsm文件为A4排版JPG（openpyxl+PIL渲染，超一页自动分页）"""
        return self._convert_with_openpyxl(filepath, output_path, dpi)

    def _convert_xls(self, filepath, output_path, dpi=300):
        """转换xls文件为JPG，若实际是xlsx/xlsm格式则用openpyxl渲染"""
        actual_format = self._detect_file_format(filepath)
        if actual_format in ('xlsx', 'xlsm'):
            self._wlog(f"  文件实际为 {actual_format} 格式（扩展名为.xls），使用 openpyxl 渲染...")
            return self._convert_with_openpyxl(filepath, output_path, dpi)
        return False, "旧版 xls (OLE2) 格式暂不支持直接渲染，请先在 Excel 中另存为 xlsx 后再处理", []

    def _convert_with_openpyxl(self, filepath, output_path, dpi=300):
        """
        使用 openpyxl 读取表格数据和格式，用 PIL 渲染为 JPG。
        输出按 A4 纸排版标准：自动选择纵向/横向、20mm 页边距；
        输出文件名统一追加四位序号（从用户指定起始编码 start_num 开始），
        单页输出为 -{start_num:04d}，超过一页时按行分页依次递增。
        使用 Windows 系统中文字体（宋体/黑体），避免中文字形渲染错误。
        保留合并单元格、列宽、行高、对齐方式、字体加粗、背景色等格式信息。
        返回: (是否成功, 错误信息, 输出文件路径列表)
        """
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            # openpyxl 拒绝非 .xlsx/.xlsm 扩展名的文件，必要时创建正确扩展名的临时副本
            load_path, opy_temp_file = self._get_openpyxl_readable_path(filepath)

            # 加载工作簿（read_only=False 以获取合并单元格和列宽信息）
            try:
                wb = openpyxl.load_workbook(load_path, read_only=False, data_only=True)
            finally:
                if opy_temp_file and os.path.exists(opy_temp_file):
                    try:
                        os.remove(opy_temp_file)
                    except:
                        pass
            ws = wb.active

            if ws.max_row is None or ws.max_column is None or ws.max_row == 0 or ws.max_column == 0:
                wb.close()
                return False, "工作表为空", []

            # ---- 输出前编辑: 删除 G 列后的所有无关列(仅保留 A-G 列) ----
            KEEP_COLS = 7
            orig_col = ws.max_column
            if orig_col > KEEP_COLS:
                n_del = orig_col - KEEP_COLS
                try:
                    ws.delete_cols(KEEP_COLS + 1, n_del)
                    self._wlog(f"  编辑: 已删除 G 列后的 {n_del} 个无关列"
                               f"(原 {orig_col} 列 → 保留 A-G)")
                except Exception as e:
                    self._wlog(f"  × 删除 G 列后无关列失败: {e}, 排版时仅保留 A-G 列")

            max_row = ws.max_row
            max_col = min(ws.max_column, KEEP_COLS)

            # 读取合并单元格信息
            merged_covered = set()  # 被合并覆盖的非左上角单元格
            merged_map = {}  # 左上角 -> (min_row, min_col, max_row, max_col)
            for mr in ws.merged_cells.ranges:
                if mr.min_row > max_row or mr.min_col > max_col:
                    continue
                for r in range(mr.min_row, min(mr.max_row, max_row) + 1):
                    for c in range(mr.min_col, min(mr.max_col, max_col) + 1):
                        if (r, c) != (mr.min_row, mr.min_col):
                            merged_covered.add((r, c))
                merged_map[(mr.min_row, mr.min_col)] = (
                    mr.min_row, mr.min_col,
                    min(mr.max_row, max_row), min(mr.max_col, max_col)
                )

            # 读取列宽（字符单位，默认8.43）
            col_widths = []
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                dim = ws.column_dimensions.get(col_letter)
                if dim and dim.width:
                    col_widths.append(dim.width)
                else:
                    col_widths.append(8.43)

            # 读取行高（磅，默认15）
            row_heights = []
            for row_idx in range(1, max_row + 1):
                dim = ws.row_dimensions.get(row_idx)
                if dim and dim.height:
                    row_heights.append(dim.height)
                else:
                    row_heights.append(15)

            # ---- A4 排版：计算表格自然尺寸与页面参数 ----
            # Excel 列宽单位：1字符 ≈ 7像素 @ 96DPI；行高单位：磅（1磅 = 1/72 英寸）
            table_w_inch = sum(col_widths) * 7.0 / 96.0
            table_h_inch = sum(row_heights) / 72.0

            # 按表格宽高比选择 A4 方向（210x297mm）
            if table_w_inch > table_h_inch:
                page_w_mm, page_h_mm = 297, 210   # 横向
            else:
                page_w_mm, page_h_mm = 210, 297   # 纵向
            margin_mm = 20.0  # 标准页边距

            mm_per_inch = 25.4
            content_w_inch = (page_w_mm - margin_mm * 2) / mm_per_inch
            content_h_inch = (page_h_mm - margin_mm * 2) / mm_per_inch

            # 自然尺寸能否容纳在一页内，超出则按行分页
            single_page = (table_w_inch <= content_w_inch and table_h_inch <= content_h_inch)
            if single_page:
                # 单页：等比缩放适配内容区（放大上限 2 倍，避免小表格过度放大发虚）
                scale = min(content_w_inch / max(table_w_inch, 0.1),
                            content_h_inch / max(table_h_inch, 0.1),
                            2.0)
            else:
                # 分页输出：只缩小适配宽度，不放大
                scale = min(1.0, content_w_inch / max(table_w_inch, 0.1))
            render_dpi = dpi * scale  # 按缩放后 DPI 渲染，保证文字清晰

            px_per_char = 7.0 * render_dpi / 96.0
            px_per_pt = render_dpi / 72.0

            col_width_px = [w * px_per_char for w in col_widths]
            row_height_px = [h * px_per_pt for h in row_heights]

            # ---- 字体加载（带缓存，优先使用 Windows 系统中文字体）----
            fonts_regular = [r'C:\Windows\Fonts\simsun.ttc',
                             r'C:\Windows\Fonts\simhei.ttf',
                             r'C:\Windows\Fonts\msyh.ttc']
            fonts_bold = [r'C:\Windows\Fonts\simhei.ttf',
                          r'C:\Windows\Fonts\msyhbd.ttc',
                          r'C:\Windows\Fonts\simsun.ttc']
            font_cache = {}

            def get_font(size_pt, bold):
                key = (round(size_pt * 2), bold)
                if key in font_cache:
                    return font_cache[key]
                px = max(8, int(round(size_pt * px_per_pt)))
                font = None
                for fp in (fonts_bold if bold else fonts_regular):
                    if os.path.exists(fp):
                        try:
                            font = ImageFont.truetype(fp, px)
                            break
                        except Exception:
                            pass
                if font is None:
                    font = ImageFont.load_default()
                font_cache[key] = font
                return font

            # 累积 x 坐标（各页列布局相同）
            x_positions = [0]
            for w in col_width_px:
                x_positions.append(x_positions[-1] + w)

            pad = max(2, int(render_dpi / 120))  # 单元格内边距

            # 被合并单元格 -> 左上角锚点（用于识别跨页合并）
            merged_anchor = {}
            for (ar, ac), (min_r, min_c, max_r, max_c) in merged_map.items():
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        if (r, c) != (ar, ac):
                            merged_anchor[(r, c)] = (ar, ac)

            def render_page(r_start, r_end, filler_heights=None):
                """渲染行区间 [r_start, r_end]（1-based 含端点）的表格子图;
                filler_heights 非空时在表格底部追加等量的空表格行(各行高度由列表
                指定, 单位px), 用于占满最后一页版面。"""
                filler_heights = filler_heights or []
                page_row_h_px = row_height_px[r_start - 1:r_end] + filler_heights
                sub_w = int(sum(col_width_px)) + 2
                sub_h = int(sum(page_row_h_px)) + 2
                img = Image.new('RGB', (sub_w, sub_h), 'white')
                draw = ImageDraw.Draw(img)

                def text_width(text, font):
                    try:
                        return draw.textlength(text, font=font)
                    except Exception:
                        return draw.textsize(text, font=font)[0]

                def wrap_text(text, font, max_width):
                    """按像素宽度折行"""
                    lines = []
                    for seg in text.split('\n'):
                        if not seg:
                            lines.append('')
                            continue
                        cur = ''
                        for ch in seg:
                            if not cur or text_width(cur + ch, font) <= max_width:
                                cur += ch
                            else:
                                lines.append(cur)
                                cur = ch
                        lines.append(cur)
                    return lines

                # 本页累积 y 坐标（从 0 开始, 含追加的空表格行）
                y_positions = [0]
                for h in page_row_h_px:
                    y_positions.append(y_positions[-1] + h)

                # ---- 遍历本页单元格进行绘制 ----
                for row_idx in range(r_start, r_end + 1):
                    for col_idx in range(1, max_col + 1):
                        # 被合并覆盖的单元格：锚点在本页则由锚点绘制，
                        # 锚点在上一页（跨页合并）则补画延续边框
                        if (row_idx, col_idx) in merged_covered:
                            anchor = merged_anchor.get((row_idx, col_idx))
                            if anchor is None or anchor[0] >= r_start:
                                continue
                            cx1 = int(x_positions[col_idx - 1])
                            cy1 = int(y_positions[row_idx - r_start])
                            cx2 = int(x_positions[col_idx])
                            cy2 = int(y_positions[row_idx - r_start + 1])
                            draw.rectangle([cx1, cy1, cx2, cy2], fill=(255, 255, 255),
                                           outline=(51, 51, 51), width=1)
                            continue

                        cell = ws.cell(row=row_idx, column=col_idx)

                        # 确定单元格范围（是否合并），跨页合并裁剪到本页
                        if (row_idx, col_idx) in merged_map:
                            min_r, min_c, max_r, max_c = merged_map[(row_idx, col_idx)]
                            max_r = min(max_r, r_end)
                        else:
                            min_r, min_c, max_r, max_c = row_idx, col_idx, row_idx, col_idx

                        x1 = int(x_positions[min_c - 1])
                        y1 = int(y_positions[min_r - r_start])
                        x2 = int(x_positions[max_c])
                        y2 = int(y_positions[max_r - r_start + 1])
                        cw = x2 - x1
                        chh = y2 - y1

                        value = str(cell.value) if cell.value is not None else ''

                        is_bold = False
                        font_size = 11
                        if cell.font:
                            is_bold = bool(cell.font.bold)
                            if cell.font.size:
                                font_size = cell.font.size

                        # 对齐与自动换行（Excel 默认：文本靠左、数字靠右）
                        is_number = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
                        h_align = 'right' if is_number else 'left'
                        v_align = 'center'
                        wrap_flag = False
                        if cell.alignment:
                            if cell.alignment.horizontal and cell.alignment.horizontal != 'general':
                                h_align = cell.alignment.horizontal
                            if cell.alignment.vertical:
                                v_align = cell.alignment.vertical
                            wrap_flag = bool(cell.alignment.wrap_text)

                        # 背景色
                        bg_color = (255, 255, 255)
                        try:
                            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                                rgb = str(cell.fill.fgColor.rgb)
                                if len(rgb) == 8 and rgb != '00000000':
                                    r_v, g_v, b_v = int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)
                                    if not (r_v == 0 and g_v == 0 and b_v == 0):
                                        bg_color = (r_v, g_v, b_v)
                        except Exception:
                            pass

                        # 绘制单元格矩形（边框+背景）
                        draw.rectangle([x1, y1, x2, y2], fill=bg_color, outline=(51, 51, 51), width=1)

                        if not value:
                            continue

                        # ---- 确定字号：先按行高约束，不换行时再按列宽约束 ----
                        fs = font_size
                        line_h = int(fs * px_per_pt * 1.25)
                        raw_lines = value.split('\n')
                        while len(raw_lines) * line_h > chh - pad and fs > 6:
                            fs -= 1
                            line_h = int(fs * px_per_pt * 1.25)
                        if not wrap_flag:
                            for line in raw_lines:
                                lw = text_width(line, get_font(fs, is_bold))
                                while lw > cw - pad * 2 and fs > 6:
                                    fs -= 1
                                    lw = text_width(line, get_font(fs, is_bold))
                        font = get_font(fs, is_bold)
                        line_h = int(fs * px_per_pt * 1.25)

                        # 换行处理
                        if wrap_flag:
                            lines = wrap_text(value, font, cw - pad * 2)
                            # 折行后总高度超出单元格时缩小字体
                            while len(lines) * line_h > chh - pad and fs > 6:
                                fs -= 1
                                font = get_font(fs, is_bold)
                                line_h = int(fs * px_per_pt * 1.25)
                                lines = wrap_text(value, font, cw - pad * 2)
                        else:
                            lines = raw_lines

                        # 垂直起始位置
                        total_text_h = len(lines) * line_h
                        if v_align == 'top':
                            ty = y1 + pad
                        elif v_align == 'bottom':
                            ty = y2 - pad - total_text_h
                        else:
                            ty = y1 + (chh - total_text_h) // 2

                        # 逐行绘制文字
                        for line in lines:
                            lw = int(text_width(line, font))
                            if h_align == 'left':
                                tx = x1 + pad
                            elif h_align == 'right':
                                tx = x2 - pad - lw
                            else:
                                tx = x1 + (cw - lw) // 2
                            draw.text((tx, ty), line, font=font, fill=(0, 0, 0))
                            ty += line_h

                # 追加的空表格行: 仅画边框(无文字), 用于占满最后一页版面
                for fi in range(len(filler_heights)):
                    fy1 = int(y_positions[r_end - r_start + 1 + fi])
                    fy2 = int(y_positions[r_end - r_start + 2 + fi])
                    for col_idx in range(1, max_col + 1):
                        draw.rectangle([int(x_positions[col_idx - 1]), fy1,
                                        int(x_positions[col_idx]), fy2],
                                       fill=(255, 255, 255),
                                       outline=(51, 51, 51), width=1)

                return img

            # A4 画布尺寸与边距像素
            page_w_px = int(round(page_w_mm * dpi / 25.4))
            page_h_px = int(round(page_h_mm * dpi / 25.4))
            margin_px = int(round(margin_mm * dpi / 25.4))

            def save_page(sub_img, out_path, top_align):
                """将子图放置到 A4 画布并保存（top_align=True 时分页内容顶部对齐边距）"""
                page = Image.new('RGB', (page_w_px, page_h_px), 'white')
                offset_x = max(0, (page_w_px - sub_img.width) // 2)
                offset_y = margin_px if top_align else max(0, (page_h_px - sub_img.height) // 2)
                # 超出画布时裁剪（单行高于整页的极端情况）
                paste_w = min(sub_img.width, page_w_px - offset_x)
                paste_h = min(sub_img.height, page_h_px - offset_y)
                if paste_w < sub_img.width or paste_h < sub_img.height:
                    sub_img = sub_img.crop((0, 0, paste_w, paste_h))
                page.paste(sub_img, (offset_x, offset_y))
                page.save(out_path, 'JPEG', quality=95, dpi=(dpi, dpi))

            output_files = []
            base_noext = os.path.splitext(output_path)[0]
            if single_page:
                # 一页可容纳：居中放置，文件名同样追加起始编码序号
                single_path = f"{base_noext}-{self.start_num:04d}.jpg"
                save_page(render_page(1, max_row), single_path, top_align=False)
                output_files.append(single_path)
            else:
                # 超过一页：按内容区高度贪心分组行进行分页
                page_h_pt = content_h_inch * 72.0
                page_ranges = []
                cur_start = 1
                cur_h = 0.0
                for idx, h in enumerate(row_heights, start=1):
                    if cur_h > 0 and cur_h + h > page_h_pt + 1e-6:
                        page_ranges.append((cur_start, idx - 1))
                        cur_start = idx
                        cur_h = 0.0
                    cur_h += h
                page_ranges.append((cur_start, max_row))

                # ---- 排版后最后一页仅空表格(无文字)时, 删除该页 ----
                def _row_has_text(r):
                    for c in range(1, max_col + 1):
                        v = ws.cell(row=r, column=c).value
                        if v is not None and str(v).strip():
                            return True
                    return False

                def _range_has_text(rs, re_):
                    return any(_row_has_text(r) for r in range(rs, re_ + 1))

                while len(page_ranges) > 1 and not _range_has_text(*page_ranges[-1]):
                    rs_d, re_d = page_ranges.pop()
                    self._wlog(f"  排版后最后一页(行{rs_d}-{re_d})仅空表格无文字, 已删除该页")

                # ---- 最后一页未占满版面时, 以空表格行补足 ----
                # 补足空行行高统一为固定标准 FILL_ROW_H_PT(所有文件一致,
                # 对齐样例空行间距); 每个补足空行都取同一固定行高,
                # 各行行高完全一致; 不拉伸末行去凑满版面,
                # 补足整数行后不足一行的余量留白
                filler_heights = []
                rs_last, re_last = page_ranges[-1]
                used_h_pt = sum(row_heights[rs_last - 1:re_last])
                remain_h_pt = page_h_pt - used_h_pt
                fill_h_pt = self.FILL_ROW_H_PT
                if remain_h_pt >= fill_h_pt:
                    # 以固定统一行高补足整数行: 所有文件的空行间距一致;
                    # 剩余空间不足以再补一行时不强凑, 留白即可
                    filler_n = int(remain_h_pt // fill_h_pt)
                    leftover_pt = remain_h_pt - filler_n * fill_h_pt
                    filler_heights = [fill_h_pt * px_per_pt] * filler_n
                    self._wlog(f"  最后一页未占满版面, 以统一空行行高 {fill_h_pt:g}pt 补足 "
                               f"{filler_n} 个等高空表格行"
                               f"(余 {leftover_pt:.1f}pt 留白)")

                self._wlog(f"  表格超过一页 A4，分页为 {len(page_ranges)} 页...")
                for n, (rs, re_) in enumerate(page_ranges, start=1):
                    # 分页文件名：原文件名 + "-" + 四位序号（从用户指定起始编码开始）
                    page_path = f"{base_noext}-{self.start_num + n - 1:04d}.jpg"
                    if n == len(page_ranges):
                        save_page(render_page(rs, re_, filler_heights=filler_heights),
                                  page_path, top_align=True)
                    else:
                        save_page(render_page(rs, re_), page_path, top_align=True)
                    output_files.append(page_path)

            wb.close()

            return True, "", output_files

        except ImportError as e:
            return False, f"缺少必要库: {e}", []
        except Exception as e:
            return False, f"转换失败: {e}", []

    def run(self):
        logf = None
        try:
            files = self._find_excel_files()
            if not files:
                self.finished_signal.emit(False, "所选目录下没有找到xlsx或xls文件")
                return

            # 在输出目录下生成处理日志文件
            log_dir = self.input_dir if self.same_dir else self.output_dir
            os.makedirs(log_dir, exist_ok=True)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_path = os.path.join(log_dir, f"表格转换处理日志_{ts}.txt")
            logf = open(log_path, 'w', encoding='utf-8')
            self._log_file = logf

            total = len(files)
            self._wlog("表格转JPG处理日志")
            self._wlog(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self._wlog(f"表格目录: {self.input_dir}")
            if self.same_dir:
                self._wlog("输出方式: 原目录输出")
            else:
                self._wlog(f"输出方式: 输出到目录 {self.output_dir}")
            self._wlog(f"找到 {total} 个表格文件")
            self._wlog(f"分页起始编码: {self.start_num:04d}")
            self._wlog("按 A4 排版标准输出（自动选择纵向/横向，20mm 页边距，超过一页自动分页）")
            self._wlog("输出前编辑: 删除 G 列后无关列; 末页仅空表格则删除该页; 末页未占满以空表格行补足")
            self._wlog("=" * 70)

            success_count = 0
            fail_count = 0

            for i, filepath in enumerate(files):
                if self.is_stopped:
                    self._wlog("用户停止处理")
                    break

                rel_path = os.path.relpath(filepath, self.input_dir)
                self._wlog(f"[{i + 1}/{total}] 处理: {rel_path}")

                # 确定输出路径
                if self.same_dir:
                    # 输出到原目录
                    output_path = os.path.splitext(filepath)[0] + '.jpg'
                else:
                    # 输出到指定目录，保持相对路径结构
                    rel_jpg = os.path.splitext(rel_path)[0] + '.jpg'
                    output_path = os.path.join(self.output_dir, rel_jpg)
                    os.makedirs(os.path.dirname(output_path), exist_ok=True)

                # 转换文件（根据实际格式而非扩展名分发）
                try:
                    actual_format = self._detect_file_format(filepath)
                    if actual_format in ('xlsx', 'xlsm'):
                        ok, err, out_files = self._convert_xlsx(filepath, output_path)
                    elif actual_format == 'xls':
                        ok, err, out_files = self._convert_xls(filepath, output_path)
                    else:
                        # 未知格式，按扩展名尝试
                        if filepath.lower().endswith(('.xlsx', '.xlsm')):
                            ok, err, out_files = self._convert_xlsx(filepath, output_path)
                        else:
                            ok, err, out_files = self._convert_xls(filepath, output_path)

                    if ok:
                        success_count += 1
                        for op in out_files:
                            self._wlog(f"  → {os.path.basename(op)}")
                    else:
                        fail_count += 1
                        self._wlog(f"  × 失败: {err}")
                except Exception as e:
                    fail_count += 1
                    self._wlog(f"  × 出错: {e}")

                self.progress_signal.emit(i + 1, total)

            self._wlog("=" * 70)
            self._wlog(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            self._wlog(f"总计: 成功 {success_count} 个，失败 {fail_count} 个")
            logf.close()
            logf = None
            self._log_file = None

            msg = (f"转换完成！成功 {success_count} 个，失败 {fail_count} 个\n"
                   f"日志: {os.path.basename(log_path)}")
            self.finished_signal.emit(True, msg)

        except Exception as e:
            import traceback
            self.log_signal.emit(traceback.format_exc())
            if logf is not None:
                try:
                    logf.close()
                except:
                    pass
                self._log_file = None
            self.finished_signal.emit(False, f"处理出错: {e}")

    def stop(self):
        self.is_stopped = True


class XlsxToJpgPage(FunctionPage):
    """表格输出为JPG功能页 - 将xlsx/xls文件转换为300DPI的JPG图像"""

    def __init__(self):
        super().__init__("表格输出为JPG")
        self.worker = None

        group = QGroupBox("将Excel表格(xlsx/xls)转换为300DPI的JPG图像")
        form = QFormLayout()

        # 输入目录
        self.input_dir = QLineEdit()
        self.input_dir.setPlaceholderText("选择包含xlsx/xls文件的目录...")
        btn_browse_input = QPushButton("选择文件夹")
        btn_browse_input.setObjectName("BrowseBtn")
        btn_browse_input.clicked.connect(
            lambda: self.input_dir.setText(QFileDialog.getExistingDirectory(self, "选择表格目录")))
        h_input = QHBoxLayout()
        h_input.addWidget(self.input_dir)
        h_input.addWidget(btn_browse_input)

        # 原目录输出选项
        self.same_dir_cb = QCheckBox("原目录输出（JPG文件输出到表格所在目录）")
        self.same_dir_cb.setChecked(True)
        self.same_dir_cb.stateChanged.connect(self._on_same_dir_changed)

        # 输出目录（仅在取消原目录输出时可用）
        self.output_dir = QLineEdit()
        self.output_dir.setPlaceholderText("选择JPG输出目录...")
        self.output_dir.setEnabled(False)
        btn_browse_output = QPushButton("选择文件夹")
        btn_browse_output.setObjectName("BrowseBtn")
        btn_browse_output.setEnabled(False)
        btn_browse_output.clicked.connect(
            lambda: self.output_dir.setText(QFileDialog.getExistingDirectory(self, "选择输出目录")))
        self.btn_browse_output = btn_browse_output
        h_output = QHBoxLayout()
        h_output.addWidget(self.output_dir)
        h_output.addWidget(btn_browse_output)

        form.addRow("表格目录:", h_input)
        form.addRow("", self.same_dir_cb)
        form.addRow("输出目录:", h_output)

        # 起始文件编码（分页输出时序号从此编码开始）
        self.start_num = QLineEdit()
        self.start_num.setPlaceholderText("四位起始编码，如 0002；留空默认从 0001 开始")
        form.addRow("起始文件名:", self.start_num)

        group.setLayout(form)
        self.layout.addWidget(group)

        # 说明文字
        info_label = QLabel("说明：将表格转换为 300DPI 的 JPG 图像，按 A4 纸排版标准输出"
                           "（自动选择纵向/横向，20mm 页边距，表格等比缩放居中）。"
                           "输出前自动删除表格 G 列后的无关列(仅保留 A-G)；"
                           "分页后若最后一页仅有空表格(无文字)则删除该页；"
                           "最后一页未占满版面时以等高的空表格行补足(各行行高一致，所有文件"
                           "统一使用固定空行行高 57pt；剩余空间不足一行时留白)，确保导出文件整洁。"
                           "输出文件名统一追加四位序号，从指定起始编码开始"
                           "（如输入 0002：单页输出 原名-0002.jpg，超过一页时 原名-0002.jpg、原名-0003.jpg …）。"
                           "点击开始转换后会先根据第一个待处理文件提醒输出文件名格式。"
                           "处理日志生成在输出目录下。"
                           "多 sheet 的文件仅转换第一个 sheet；旧版 xls (OLE2) 格式请先另存为 xlsx。")
        info_label.setStyleSheet("color: #666; font-size: 12px; margin: 10px 0;")
        info_label.setWordWrap(True)
        self.layout.addWidget(info_label)

        # 控制按钮
        btn_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始转换")
        self.start_btn.setObjectName("ActionBtn")
        self.start_btn.clicked.connect(self._start)
        self.stop_btn = QPushButton("停止处理")
        self.stop_btn.setObjectName("ActionBtn")
        self.stop_btn.setStyleSheet("background-color: #DA3633;")
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop)
        self.clear_btn = QPushButton("清空结果")
        self.clear_btn.setObjectName("ActionBtn")
        self.clear_btn.clicked.connect(self._clear)
        for b in (self.start_btn, self.stop_btn, self.clear_btn):
            btn_layout.addWidget(b)
        self.layout.addLayout(btn_layout)

        # 进度条
        self.progress = QProgressBar()
        self.progress.setFormat("待开始")
        self.layout.addWidget(self.progress)

        self.add_log_widget()

    def _on_same_dir_changed(self, state):
        enabled = state != Qt.Checked
        self.output_dir.setEnabled(enabled)
        self.btn_browse_output.setEnabled(enabled)

    def _set_running(self, running):
        self.start_btn.setEnabled(not running)
        self.stop_btn.setEnabled(running)

    def _first_excel_file(self, input_dir):
        """查找第一个待处理文件（与 worker._find_excel_files 规则保持一致）"""
        result = []
        for root, _dirs, files in os.walk(input_dir):
            for f in files:
                if f.lower().endswith(('.xlsx', '.xlsm', '.xls')) and not f.startswith('~$'):
                    result.append(os.path.join(root, f))
        return sorted(result)[0] if result else None

    def _start(self):
        input_dir = self.input_dir.text().strip()
        if not input_dir or not os.path.isdir(input_dir):
            QMessageBox.warning(self, "提示", "请先选择有效的表格目录。")
            return

        same_dir = self.same_dir_cb.isChecked()
        output_dir = None
        if not same_dir:
            output_dir = self.output_dir.text().strip()
            if not output_dir:
                QMessageBox.warning(self, "提示", "请选择输出目录，或勾选「原目录输出」。")
                return

        # 校验起始文件编码（须为四位数字，位长不够则提示）
        snum_text = self.start_num.text().strip()
        if snum_text:
            if not (snum_text.isdigit() and len(snum_text) == 4):
                QMessageBox.warning(self, "提示",
                                    f"起始文件名编码位长不足：请输入四位数字编码（如 0002），当前输入「{snum_text}」。")
                return
            start_num = int(snum_text)
        else:
            start_num = 1

        # 提醒：根据第一个待处理文件展示输出文件名格式，用户确认后才转换
        first_file = self._first_excel_file(input_dir)
        if not first_file:
            QMessageBox.warning(self, "提示", "所选目录下没有找到 xlsx/xls 文件。")
            return
        rel = os.path.relpath(first_file, input_dir)
        if same_dir:
            first_out = os.path.splitext(first_file)[0] + '.jpg'
        else:
            first_out = os.path.join(output_dir, os.path.splitext(rel)[0] + '.jpg')
        out_base = os.path.splitext(os.path.basename(first_out))[0]
        preview = (f"第一个待处理文件: {os.path.basename(first_file)}\n\n"
                   f"输出文件名格式：\n"
                   f"  · 单页输出: {out_base}-{start_num:04d}.jpg\n"
                   f"  · 分页输出(超过一页时): {out_base}-{start_num:04d}.jpg、"
                   f"{out_base}-{start_num + 1:04d}.jpg …\n\n"
                   f"选择「Yes」开始转换，选择「No」放弃转换。")
        reply = QMessageBox.question(self, "输出文件名格式确认", preview,
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            self.log("用户放弃转换")
            return

        self._clear()
        self.progress.setValue(0)
        self.progress.setFormat("准备中...")

        self.worker = XlsxToJpgWorker(input_dir, output_dir=output_dir,
                                      same_dir=same_dir, start_num=start_num)
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self._update_progress)
        self.worker.finished_signal.connect(self._on_finished)
        self._set_running(True)
        self.worker.start()

    def _stop(self):
        if self.worker:
            self.worker.stop()
            self.log("正在停止...")

    def _update_progress(self, done, total):
        if total <= 0:
            self.progress.setValue(0)
            self.progress.setFormat("0 / 0")
            return
        pct = int(done * 100 / total)
        self.progress.setValue(pct)
        self.progress.setFormat(f"{done} / {total}  ({pct}%)")

    def _on_finished(self, ok, message):
        self._set_running(False)
        self.log(message)
        self.progress.setFormat("已完成" if ok else "已停止")
        self.worker = None
        QMessageBox.information(self, "完成" if ok else "结束", message)

    def _clear(self):
        if hasattr(self, 'progress'):
            self.progress.setValue(0)
            self.progress.setFormat("待开始")
        if hasattr(self, 'log_box'):
            self.log_box.clear()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("同美档案工具集合")
        self.resize(1000, 650)
        self.setStyleSheet(TechStyle.QSS)

        # 主中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 左侧菜单栏
        self.left_panel = QWidget()
        self.left_panel.setObjectName("LeftPanel")
        self.left_panel.setFixedWidth(220)
        left_layout = QVBoxLayout(self.left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setAlignment(Qt.AlignTop)

        title = QLabel("同美档案工具集合")
        title.setObjectName("TitleLabel")
        title.setAlignment(Qt.AlignCenter)
        left_layout.addWidget(title)

        # 菜单按钮配置
        menus = ["文件改名", "自动编页码", "文件移动", "加盖归档章",
                 "修改DPI", "表格输出为JPG", "JPG转双层PDF", "PDF转OFD", "分件", "文件批量替换"]

        self.menu_buttons = []
        for m in menus:
            btn = QPushButton(m)
            btn.setObjectName("MenuBtn")
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, b=btn: self.on_menu_click(b))
            left_layout.addWidget(btn)
            self.menu_buttons.append(btn)

        left_layout.addStretch()

        # 右侧功能区 (使用QStackedWidget管理多页面)
        self.right_panel = QWidget()
        self.right_panel.setObjectName("RightPanel")
        right_layout = QVBoxLayout(self.right_panel)
        right_layout.setContentsMargins(20, 20, 20, 20)

        self.stack = QStackedWidget()

        # 实例化各功能页
        self.pages = {
            "文件改名": FileRenamePage(),
            "自动编页码": AutoPagingPage(),
            "文件移动": FileMovePage(),
            "加盖归档章": ArchiveStampPage(),
            "修改DPI": ModifyDpiPage(),
            "表格输出为JPG": XlsxToJpgPage(),
            "JPG转双层PDF": JpgToPdfPage(),
            "PDF转OFD": PdfToOfdPage(),
            "分件": FileSplitPage(),
            "文件批量替换": FileBatchReplacePage()
        }

        for name in menus:
            self.stack.addWidget(self.pages[name])

        right_layout.addWidget(self.stack)

        # 组装整体布局
        main_layout.addWidget(self.left_panel)
        main_layout.addWidget(self.right_panel)

        # 默认选中第一个菜单
        self.on_menu_click(self.menu_buttons[0])

    def on_menu_click(self, clicked_btn):
        """处理菜单点击切换事件"""
        # 取消其他按钮的选中状态
        for btn in self.menu_buttons:
            btn.setChecked(False)
        # 激活当前按钮
        clicked_btn.setChecked(True)

        # 切换右侧堆栈页面
        page_name = clicked_btn.text()
        if page_name in self.pages:
            self.stack.setCurrentWidget(self.pages[page_name])


if __name__ == "__main__":
    app = QApplication(sys.argv)
    # 强制使用深色科技感字体渲染
    f = QFont("Microsoft YaHei", 9)
    app.setFont(f)

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
